from django.shortcuts import render, redirect
from main.models import Produit, Mark, Category, Supplier, Stockin, Itemsbysupplier, Client, Represent, Order, Orderitem, Clientprices, Bonlivraison, Facture, Outfacture, Livraisonitem, PaymentClientbl, PaymentClientfc,  PaymentSupplier, Bonsregle, Returnedsupplier, Avoirclient, Returned, Avoirsupplier, Orderitem, Carlogos, Ordersnotif, Connectedusers, Promotion, UserSession, Refstats, Notavailable, Cart, Wishlist, Wich, Notification, Cartitems
from django.contrib.auth import logout
from django.http import JsonResponse, HttpResponse
import openpyxl
# import Count
from django.contrib.auth.models import User
from django.db.models import Count, F, Sum, Q, ExpressionWrapper, Func, fields, IntegerField
from django.db.models.functions import Cast
from django.contrib.sessions.models import Session
from functools import wraps
from django.contrib.auth.decorators import user_passes_test
import json
from django.contrib.auth.models import Group
from django.db import transaction
from datetime import datetime, date, timedelta
from django.utils import timezone
import pandas as pd
from itertools import chain
from django.core.serializers import serialize
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate, login, logout
today = timezone.now().date()
thisyear=timezone.now().year
import requests as req

def isadmin(user):
    if not user.groups.filter(name='admin').exists():
        redirect('main:logoutuser')

# call isadmin here so that it executes before executing all the views

def getproductsbycategory(request):
    # category = Category.objects.get(pk=request.POST.get('category'))
    # products = category.product.filter(category=category)[:10]
    # get ten products from the category
    products = Produit.objects.filter(category__pk=request.POST.get('category')).order_by('code')
    # get marks of the products filtered
    marks = Mark.objects.filter(produit__in=products).distinct().annotate(count=Count('produit'))
    ctx={
        'products':products,
        'home':False,
        'marks':marks,

    }
    return JsonResponse({
        'data':render(request, 'product_search.html', ctx).content.decode('utf-8'),
        'stocktotal':products.aggregate(Sum('stocktotal'))['stocktotal__sum'] or 0,
        'stockfacture':products.aggregate(Sum('stockfacture'))['stockfacture__sum'] or 0,
    })

def adminaddproductpage(request):
    categories=Category.objects.all()
    marques=Mark.objects.all()
    ctx={'categories':categories,
         'marques':marques,
         'commercials':Represent.objects.all(),
         'carlogos':Carlogos.objects.all(),
        }
    return render(request, 'addproduct.html', ctx)



def categoriespage(request):
    ctx={
        'categories':Category.objects.all().order_by('code'),
        'commercials':Represent.objects.all(),
        'title':'Categories'
    }
    return render(request, 'categories.html', ctx)
@csrf_exempt
def createcategory(request):
    print('>>>>>>>> creating category')
    name=request.POST.get('name')
    code=request.POST.get('code')
    affichage=request.POST.get('affichage')
    hideclient=request.POST.get('hideclient')=='True'
    commercialexcluded=request.POST.getlist('commercialexcluded')
    reps=Represent.objects.filter(pk__in=commercialexcluded)
    print(len(commercialexcluded)>0)

    # get image file
    image = request.FILES.get('image')
    # create category
    category=Category.objects.create(name=name, image=image, code=code, affichage=affichage, masqueclients=hideclient)
    if len(commercialexcluded)>0:
        print('getting reps', commercialexcluded, reps)
        category.excludedrep.set(reps)
    return JsonResponse({
        'success':True
    })
@csrf_exempt
def updatecategory(request):
    
    id=request.POST.get('id')
    hideclient=request.POST.get('hideclient')
    commercialexcluded=request.POST.getlist('commercialexcluded')
    reps=Represent.objects.filter(pk__in=commercialexcluded)
    category=Category.objects.get(pk=id)
    category.masqueclients=hideclient
    category.excludedrep.clear()
    category.excludedrep.set(reps)
    category.name=request.POST.get('name')
    category.affichage=request.POST.get('affichage')
    category.code=request.POST.get('code')
    image=request.FILES.get('image')
    if image:
        category.image=image
    category.save()
    return JsonResponse({
        'success':True
    })


@csrf_exempt
def createmarque(request):
    try:
        name=request.POST.get('name')
        # get image file
        image=request.FILES.get('image')
        # create category
        hideclient=request.POST.get('hideclient')
        commercialexcluded=request.POST.getlist('commercialexcluded')
        reps=Represent.objects.filter(pk__in=commercialexcluded)
        print(hideclient, commercialexcluded, reps)
        mrq=Mark.objects.create(name=name, image=image, masqueclients=hideclient)
        mrq.excludedrep.set(reps)
        return JsonResponse({
            'success':True
        })
    except Exception as e:
        print(">>>>>error", e)
@csrf_exempt
def updatemarque(request):

    id=request.POST.get('id')
    print('>>>>>> updating mark', id)
    try:
        image=request.FILES.get('image')
        hideclient=request.POST.get('hideclient')
        commercialexcluded=request.POST.getlist('commercialexcluded')
        reps=Represent.objects.filter(pk__in=commercialexcluded)
        mark=Mark.objects.get(pk=id)
        mark.name=request.POST.get('name')
        mark.masqueclients=hideclient
        mark.excludedrep.clear()
        mark.excludedrep.set(reps)
        if image:
            mark.image=image
        mark.save()
        
        return JsonResponse({
            'success':True
        })
    except Exception as e:
        print(">>>>", e)

def checkref(request):
    ref=request.POST.get('ref').lower().strip()
    product=Produit.objects.filter(ref=ref)
    print(ref)
    if product:
        return JsonResponse({
            'exist':True
        })
    return JsonResponse({
        'exist':False
    })

    # print(ref, category, product)
    # if product:
    # else:
    #     return JsonResponse({
    #         'exist':False
    #     })

def supplierspage(request):
    ctx={
        'suppliers':Supplier.objects.all(),
        'title':'List des fournisseurs'
    }
    return render(request, 'suppliers.html', ctx)

def getsupplierdata(request):
    id=request.POST.get('id')
    supplier=Supplier.objects.get(pk=id)
    return JsonResponse({
        'name':supplier.name,
        'phone':supplier.phone,
        'address':supplier.address,
        'id':supplier.id
    })


def updatesupplier(request):
    id=request.POST.get('pid')
    name=request.POST.get('pname')
    phone=request.POST.get('pphone')
    address=request.POST.get('paddress')
    supplier=Supplier.objects.get(pk=id)
    supplier.name=name
    supplier.phone=phone
    supplier.address=address
    supplier.save()
    ctx={
        'suppliers':Supplier.objects.all(),
        'title':'List des fournisseurs'
    }
    return JsonResponse({
        'html':render(request, 'suppliers.html', ctx).content.decode('utf-8')
    })

def addoneproduct(request):
    try:
        ref=request.GET.get('ref').lower().strip()
        name=request.GET.get('name').strip()
        uniqcode=request.GET.get('uniqcode')
        category=request.GET.get('category')
        commercialsprix=request.GET.get('commercialsprix') or "[]"
        mark=request.GET.get('mark') or None
        logo=request.GET.get('logo', None)
        image=request.GET.get('image')
        sellprice=request.GET.get('sellprice') or 0
        supplier=request.GET.get('supplier') or None
        minstock=request.GET.get('minstock') or 0
        buyprice=request.GET.get('buyprice') or 0
        remise=request.GET.get('remise') or 0
        diametre=request.GET.get('diametre') or ''
        representprice=request.GET.get('repprice') or None
        code=request.GET.get('code') or '00'
        block=request.GET.get('block') or ''
        equivalent=request.GET.get('equiv') or ''
        cars=request.GET.getlist('cars') or ''
        print('>>>>>>>', image)
        netprice=round(float(sellprice)-(float(sellprice)*float(remise)/100), 2)
        # create product
        print('>>>>>>>>>>>>>>>creating product')
        Produit.objects.create(
            ref=ref,
            name=name,
            buyprice=buyprice,
            diametre=diametre,
            sellprice=sellprice,
            remise=remise,
            prixnet=netprice,
            representprice=representprice,
            minstock=minstock,
            equivalent=equivalent,
            cars=cars,
            category_id=category,
            supplier_id=supplier,
            mark_id=mark,
            image=image,
            uniqcode=uniqcode,
            code=code,
            repsprice=commercialsprix,
            block=block,
            stocktotal=0,
            stockfacture=0,
            isactive=True
        )

        
        return JsonResponse({
            'success':True,
        })
    except Exception as e:
        print(e)
        return JsonResponse({
            'error':e
        })


def viewoneproduct(request, id):
    product=Produit.objects.get(pk=id)
    commercial_prices = product.getcommercialsprice()  # Note the parentheses ()
    stockin=Stockin.objects.filter(product=product)
    outbl=Livraisonitem.objects.filter(product=product, isfacture=False).aggregate(Sum('qty'))['qty__sum'] or 0
    outfacture=Outfacture.objects.filter(product=product).aggregate(Sum('qty'))['qty__sum'] or 0
    revbl=Livraisonitem.objects.filter(product=product, isfacture=False).aggregate(Sum('total'))['total__sum'] or 0
    revfacture=Outfacture.objects.filter(product=product).aggregate(Sum('total'))['total__sum'] or 0
    totalout=outbl+outfacture
    totalrev=round(revbl+revfacture, 2)
    print(totalrev)
    stockout=Livraisonitem.objects.filter(product=product, isfacture=False).order_by('-id')[:50]
    stockoutfc=Outfacture.objects.filter(product=product).order_by('-id')[:50]
    avoirs=Returned.objects.filter(product=product)
    qtyin=stockin.aggregate(Sum('quantity'))['quantity__sum'] or 0
    qtyavoir=avoirs.aggregate(Sum('qty'))['qty__sum'] or 0
    releve = chain(*[
    ((outbl, 'outbl') for outbl in stockout),
    ((outfc, 'outfc') for outfc in stockoutfc),
    ])

    # Sort the items by date
    outs = sorted(releve, key=lambda item: item[0].date)
    ctx={
        'outs':outs,
        'title':'Detail de '+product.ref,
        'product':product,
        'cars':product.getcars(),
        'carlogos':Carlogos.objects.all(),
        'categories':Category.objects.all(),
        'marques':Mark.objects.all(),
        'suppliers':Supplier.objects.all(),
        'entries':stockin,
        'sorties':stockout,
        'totalqtyin':qtyin+qtyavoir,
        'totalcout':stockin.aggregate(Sum('total'))['total__sum'] or 0,
        'totalqtyout':totalout,
        'totalcoutout':totalrev,
        'avoirs':avoirs,
        'reps':Represent.objects.all(),
        'repswithprice':commercial_prices
    }
    return render(request, 'viewoneproduct.html', ctx)
@csrf_exempt
def updateproduct(request):
    token = request.headers.get('X-SYNC-TOKEN')
    if token != '8a7f5b2c9d3e4f1g0h6j':
        return JsonResponse({'error': 'Forbidden'}, status=403)
    productid=request.POST.get('productid')
    uniqcode=request.POST.get('uniqcode')
    new=request.POST.get('new')=='True'
    near=request.POST.get('near')=='True'
    logos = request.POST.getlist('logos')
    remise=request.POST.get('remise')
    sellprice=request.POST.get('sellprice')
    netprice=request.POST.get('netprice')
    image = request.FILES.get('image')
    product=Produit.objects.get(uniqcode=uniqcode)
    if product.sellprice:
        if float(sellprice) != float(product.sellprice):
            print('>>>>>>>>price changed', sellprice, product.sellprice, sellprice != product.sellprice)
            reliquas=Wishlist.objects.filter(product=product)
            for i in reliquas:
                i.total=round(float(netprice)*float(i.qty), 2)
                i.save()
            cartitems=Cartitems.objects.filter(product=product)
            for i in cartitems:
                newtotal=round(float(netprice)*float(i.qty), 2)
                newcarttotal=i.cart.total-i.total+newtotal
                i.total=newtotal
                i.save()
                i.cart.total=newcarttotal
                i.cart.save()
    product.carlogos.set(logos)
    product.near=near
    product.isnew=new
    if image:
        product.image=image
    product.equivalent=request.POST.get('equivalent')
    product.representprice=request.POST.get('repprice')
    product.code=request.POST.get('code')
    product.refeq1=request.POST.get('refeq1')
    product.refeq2=request.POST.get('refeq2')
    product.refeq3=request.POST.get('refeq3')
    product.refeq4=request.POST.get('refeq4')
    product.coderef=request.POST.get('coderef')
    product.stocktotal=request.POST.get('stock', 0)
    product.sellprice=sellprice
    product.remise=remise
    product.prixnet=netprice
    product.name=request.POST.get('name')
    product.cars=request.POST.get('cars')
    product.ref=request.POST.get('ref').lower().strip()
    product.category_id=request.POST.get('category_id')
    product.mark_id=request.POST.get('mark_id')
    product.diametre=request.POST.get('diametre')
    product.block=request.POST.get('block')
    product.save()
    return JsonResponse({
        'success':True
    })

def alertstock(request):
    targets = Category.objects.filter(produit__stocktotal__lte=F('produit__minstock')).annotate(
    total_products=Count('produit')
    )
    return render(request, 'alertstock.html', {'title':'Alert Stock', 'categories':targets,
    'suppliers':Supplier.objects.all()})

def getlowbycategory(request):
    category=request.POST.get('category')
    products=Produit.objects.filter(category__id=category, stocktotal__lte=F('minstock'))
    return JsonResponse({
        'data':render(request, 'lowstockproducts.html', {'products':products}).content.decode('utf-8')
    })

def commandsupplier(request):
    productid=request.POST.get('productid')
    supplierid=request.POST.get('supplierid')
    qty=request.POST.get('qty')
    product=Produit.objects.get(pk=productid)
    product.qtycommand=qty
    product.suppliercommand_id=supplierid
    product.iscommanded=True
    product.save()
    return JsonResponse({
        'success':True
    })

def cacelcommand(request):
    productid=request.POST.get('productid')
    product=Produit.objects.get(pk=productid)
    product.qtycommand=0
    product.suppliercommand_id=None
    product.iscommanded=False
    product.save()
    return JsonResponse({
        'success':True
    })

def recevoir(request):
    return render(request, 'recevoir.html', {'title':"Bon d'achat", 'suppliers':Supplier.objects.all(), 'today':timezone.now().date()})

def bonlivraison(request):
    # get the last order_no
    # if there is no order_no then set it to this format 'ym0001'
    # else increment it by one

    # increment it
    year = timezone.now().strftime("%y")
    latest_receipt = Bonlivraison.objects.filter(
        bon_no__startswith=f'BL{year}'
    ).order_by("-bon_no").first()
    if latest_receipt:
        latest_receipt_no = int(latest_receipt.bon_no[-5:])
        receipt_no = f"BL{year}{latest_receipt_no + 1:05}"
    else:
        receipt_no = f"BL{year}00001"
    print('>>>>>>', receipt_no)
    return render(request, 'bonlivraison.html', {
        'title':'Bon de livraison',
        # 'clients':Client.objects.all(),
        # 'products':Produit.objects.all(),
        'commercials':Represent.objects.all(),
    })

def facture(request):
    # get the last order_no
    # if there is no order_no then set it to this format 'ym0001'
    # else increment it by one

    # increment it
    return render(request, 'facture.html', {
        'title':'Facture',
        # 'clients':Client.objects.all(),
        # 'products':Produit.objects.all(),
        'commercials':Represent.objects.all(),
        #'order_no':receipt_no
    })


def suppliercommanproducts(request):
    supplierid=request.POST.get('supplierid')
    products=Produit.objects.filter(suppliercommand_id=supplierid)
    return JsonResponse({
        'data':render(request, 'suppliercommandproducts.html', {'products':products}).content.decode('utf-8')
    })


def searchref(request):
    ref=request.POST.get('ref')
    products=Produit.objects.filter(ref__istartswith=ref)
    return JsonResponse({
        'data':render(request, 'productsbon.html', {'products':products}).content.decode('utf-8')
    })

def addsupply(request):
    supplierid=request.POST.get('supplierid')
    products=request.POST.get('products')
    datebon=datetime.strptime(request.POST.get('datebon'), '%Y-%m-%d')
    datefacture=datetime.strptime(request.POST.get('datefacture'), '%Y-%m-%d')
    nbon=request.POST.get('nbon')
    isfacture= True if request.POST.get('mode')=='facture' else False
    totalbon=request.POST.get('totalbon')
    tva=0
    if isfacture:
        tva=round(float(totalbon)-(float(totalbon)/1.2), 2)

    supplier=Supplier.objects.get(pk=supplierid)
    # print(f'''
    #     bon achat N°: {nbon}
    #     date facture: {datefacture}
    #     date reception: {datebon}
    #     fournisseur: {supplier.name}
    #     sold fournisseur: {float(supplier.rest)+float(totalbon)}
    #     isfacure: {isfacture}
    #     tva: {tva}
    # ''')
    # for i in json.loads(products):
    #     qty=int(i['qty'])
    #     devise=0 if i['devise']=='' else i['devise']
    #     product=Produit.objects.get(pk=i['productid'])
    #     remise=0 if i['remise']=='' else int(i['remise'])
    #     buyprice=0 if i['price']=='' else i['price']
    #     netprice=round(float(buyprice)-(float(buyprice)*float(remise)/100), 2)
        
    #     print(f'''
    #         produit: {product.ref}
    #         nouvau prix achat: {netprice}
    #         devise: {devise}
    #         'qty': {qty}
    #     ''')
    #     if isfacture:
    #         print('add stock facture: ', int(product.stockfacture)+int(i['qty']))
    #     print('addstocktotal ', int(product.stocktotal)+int(i['qty']))

    supplier.rest=float(supplier.rest)+float(totalbon)
    supplier.save()
    bon=Itemsbysupplier.objects.create(
        supplier_id=supplierid,
        total=totalbon,
        date=datebon,
        nbon=nbon,
        isfacture=isfacture,
        tva=tva,
        dateentree=datefacture
    )
    for i in json.loads(products):
        print(int(i['qty']))
        devise=0 if i['devise']=='' else i['devise']
        product=Produit.objects.get(pk=i['productid'])
        remise=0 if i['remise']=='' else int(i['remise'])
        buyprice=0 if i['price']=='' else i['price']
        print(remise)
        print(type(remise))
        netprice=round(float(buyprice)-(float(buyprice)*float(remise)/100), 2)
        if isfacture:
            product.stockfacture=int(product.stockfacture)+int(i['qty'])
            
        product.buyprice=netprice
        product.stocktotal=int(product.stocktotal)+int(i['qty'])
        product.devise= devise
        Stockin.objects.create(
            date=datebon,
            product=product,
            devise=devise,
            quantity=i['qty'],
            price=i['price'],
            remise=remise,
            qtyofprice=i['qty'],
            total=i['total'],
            supplier_id=supplierid,
            nbon=bon,
            facture=isfacture
        )
    # # update cout moyen, it will be calculated by deviding total prices by total qty

        totalprices=Stockin.objects.filter(product=product).aggregate(Sum('total'))['total__sum'] or 0
        totalqty=Stockin.objects.filter(product=product).aggregate(Sum('quantity'))['quantity__sum'] or 0
        print(totalprices, totalqty)
        product.coutmoyen=round(totalprices/totalqty, 2)
        product.save()
    return JsonResponse({
        'html': render(request, 'recevoir.html', {'title':'Recevoir Les produits', 'suppliers':Supplier.objects.all(), 'products':Produit.objects.all()}).content.decode('utf-8')
    })


def addbonlivraison(request):

    #current_time = datetime.now().strftime('%H:%M:%S')
    clientid=request.POST.get('clientid')
    repid=request.POST.get('repid')
    products=request.POST.get('products')
    totalbon=request.POST.get('totalbon')
    orderid=request.POST.get('orderid', None)
    # orderno
    transport=request.POST.get('transport')
    note=request.POST.get('note')
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(f'{datebon}', '%Y-%m-%d')
    client=Client.objects.get(pk=clientid)
    client.soldtotal=round(float(client.soldtotal)+float(totalbon), 2)
    client.soldbl=round(float(client.soldbl)+float(totalbon), 2)
    client.save()
    if orderid is not None:
        cmnd=Order.objects.get(pk=orderid)
        cmnd.isdelivered=True
        cmnd.save()
    # get the last bon no
    year = timezone.now().strftime("%y")
    latest_receipt = Bonlivraison.objects.filter(
        bon_no__startswith=f'BL{year}'
    ).order_by("-bon_no").first()
    if latest_receipt:
        latest_receipt_no = int(latest_receipt.bon_no[-5:])
        receipt_no = f"BL{year}{latest_receipt_no + 1:05}"
    else:
        receipt_no = f"BL{year}00001"
    order=Bonlivraison.objects.create(
        commande_id=orderid,
        client_id=clientid,
        salseman_id=repid,
        total=totalbon,
        date=datebon,
        modlvrsn=transport,
        bon_no=receipt_no,
        note=note
    )
    print('>>>>>>', len(json.loads(products))>0)
    if len(json.loads(products))>0:
        with transaction.atomic():
            for i in json.loads(products):
                product=Produit.objects.get(pk=i['productid'])
                product.stocktotal=int(product.stocktotal)-int(i['qty'])
                product.save()
                Livraisonitem.objects.create(
                    bon=order,
                    remise=i['remise'],
                    name=i['name'],
                    ref=i['ref'],
                    product=product,
                    qty=i['qty'],
                    price=i['price'],
                    total=i['total'],
                    client_id=clientid,
                    date=datebon
                )


    # increment it
    return JsonResponse({
        "success":True
    })

# add facture not generer
def addfacture(request):
    clientid=request.POST.get('clientid')
    repid=request.POST.get('repid')
    products=request.POST.get('products')
    totalbon=request.POST.get('totalbon')
    # orderno
    transport=request.POST.get('transport', '')
    note=request.POST.get('note', '')
    #orderno=request.POST.get('orderno')
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    client=Client.objects.get(pk=clientid)
    client.soldtotal=round(float(client.soldtotal)+float(totalbon), 2)
    client.soldfacture=round(float(client.soldfacture)+float(totalbon), 2)
    client.save()
    tva=round(float(totalbon)-(float(totalbon)/1.2), 2)
    year = timezone.now().strftime("%y")
    latest_receipt = Facture.objects.filter(facture_no__startswith=f'FC{year}').order_by('-facture_no').first()
    if latest_receipt:
        latest_receipt_no = int(latest_receipt.facture_no[-5:])
        receipt_no = f"FC{year}{latest_receipt_no + 1:05}"
    else:
        receipt_no = f"FC{year}00001"
    print('>>>>>>>',latest_receipt)
    facture=Facture.objects.create(
        facture_no=receipt_no,
        total=totalbon,
        tva=tva,
        date=datebon,
        client_id=clientid,
        salseman_id=repid,
        transport=transport,
        note=note
    )
    if len(json.loads(products))>0:
        with transaction.atomic():
            for i in json.loads(products):
                product=Produit.objects.get(pk=i['productid'])
                product.stockfacture=int(product.stockfacture)-int(i['qty'])
                product.save()
                Outfacture.objects.create(
                    facture=facture,
                    remise=i['remise'],
                    name=i['name'],
                    ref=i['ref'],
                    product=product,
                    qty=i['qty'],
                    price=i['price'],
                    total=i['total'],
                    client_id=clientid,
                    date=datebon,
                )

    # year = timezone.now().strftime("%y")
    # latest_receipt = Facture.objects.filter(
    #     facture_no__startswith=f'FC{year}'
    # ).order_by("-facture_no").first()
    # latest_receipt_no = int(latest_receipt.facture_no[-5:])
    # receipt_no = f"FC{year}{latest_receipt_no + 1:05}"

    # increment it
    return JsonResponse({
        'success':True
    })



def supplierinfo(request, id):
    supplier=Supplier.objects.get(pk=id)
    ctx={
        'title':f'Info fournisseur {supplier.name.upper}',
        'supplier':supplier,
        'totalavoirs':Avoirsupplier.objects.filter(supplier=supplier).aggregate(Sum('total'))['total__sum'] or 0,
        'totalpayments':PaymentSupplier.objects.filter(supplier=supplier).aggregate(Sum('amount'))['amount__sum'] or 0,
        'totaltr':Itemsbysupplier.objects.filter(supplier=supplier).aggregate(Sum('total'))['total__sum'] or 0,
        'bons':Itemsbysupplier.objects.filter(supplier=supplier),
        'payments':PaymentSupplier.objects.filter(supplier=supplier)
    }
    return render(request, 'supplierinfo.html', ctx)

def clientinfo(request, id):
    client=Client.objects.get(pk=id)
    ctx={
        'client':client,
        'totalavoirs':Avoirclient.objects.filter(client=client).aggregate(Sum('total'))['total__sum'] or 0,
        'totalpayments':PaymentClientbl.objects.filter(client=client).aggregate(Sum('amount'))['amount__sum'] or 0,
        'totaltr':Bonlivraison.objects.filter(client=client).aggregate(Sum('total'))['total__sum'] or 0,
        'bons':Bonlivraison.objects.filter(client=client),
        'payments':PaymentClientbl.objects.filter(client=client)
    }
    return render(request, 'clientinfo.html', ctx)

def addpaymentssupp(request):
    supplierid=request.POST.get('supplierid')
    pass


def dashboard(request):

    ctx={
        'title':'Dashboard',
        'orders':Order.objects.filter(date__date=date.today()).count(),
        'products':Produit.objects.all().count(),
        'productthismonth':Orderitem.objects.filter(order__date__month=date.today().month).order_by('-qty')[:20],
        'alerts':Produit.objects.filter(stocktotal__lte=F('minstock')).count(),
        'blnotpaid':Bonlivraison.objects.filter(ispaid=False).count(),
        'boncommand':Order.objects.filter(isdelivered=False).count(),
        'soldtotal':round(Client.objects.aggregate(Sum('soldtotal'))['soldtotal__sum'] or 0, 2),



    }
    return render(request, 'pdashboard.html', ctx)

def clientspage(request):
    try:
        lastcode = Client.objects.last()
        print('lastcode', lastcode.code)
        if lastcode:

            codecl = f"{int(lastcode.code) + 1:06}"
        else:
            codecl = f"000001"
    except:
        codecl="000001"
    ctx={
        'clients':Client.objects.all()[:50],
        'title':'List des clients',
        'commerciaux':Represent.objects.all(),
        'lastcode':codecl,
        'soldtotal':round(Client.objects.aggregate(Sum('soldtotal'))['soldtotal__sum'] or 0, 2),
        'soldbl':round(Client.objects.aggregate(Sum('soldbl'))['soldbl__sum'] or 0, 2),
        'soldfacture':round(Client.objects.aggregate(Sum('soldfacture'))['soldfacture__sum'] or 0, 2),
    }
    return render(request, 'clients.html', ctx)



def checkusername(request):
    username=request.POST.get('username')
    if User.objects.filter(username=username).exists():
        return JsonResponse({
            'exist':True
        })
    else:
        return JsonResponse({
            'exist':False
        })

def commercialspage(request):
    ctx={
        'commercials':Represent.objects.all(),
        'title':'List des Commeciaux'
    }
    return render(request, 'commerciaux.html', ctx)

def addcommercial(request):
    # repusername=request.GET.get('repusername')
    # reppassword=request.GET.get('reppassword')
    repname=request.GET.get('repname')
    repphone=request.GET.get('repphone')
    repregion=request.GET.get('repregion')
    repinfo=request.GET.get('repinfo')
    # user=User.objects.create_user(username=repusername, password=reppassword)
    # # Get or create the group
    # group, created = Group.objects.get_or_create(name="salsemen")

    # # Add the user to the group
    # user.groups.add(group)

    # # Save the user
    # user.save()
    Represent.objects.create(
        # user=user,
        name=repname,
        phone=repphone,
        region=repregion,
        info=repinfo,
    )
    
    return JsonResponse({
        'success':True
    })
def checkcodeclient(request):
    code=request.POST.get('code')
    name=request.POST.get('name')
    print(Client.objects.filter(Q(name=name) | Q(code=code)))
    if Client.objects.filter(Q(name=name) | Q(code=code)).exists():
        return JsonResponse({
            'exist':True
        })
    return JsonResponse({
        'exist':False
    })




def addclient(request):
    name=request.GET.get('name')
    phone=request.GET.get('phone')
    address=request.GET.get('address')
    code=request.GET.get('code')
    city=request.GET.get('city')
    ice=request.GET.get('ice')
    region=request.GET.get('region').lower().strip()
    representant=request.GET.get('represent_id')
    print('>>>>>> adding clients')
    # username=request.POST.get('clientusername')
    # password=request.POST.get('clientpassword')
    # user=User.objects.create_user(username=username, password=password)
    client=Client.objects.create(
        city=city,
        ice=ice,
        region=region,
        represent_id=representant,
        code=code,
        name=name,
        phone=phone,
        address=address,
    )
    # group, created = Group.objects.get_or_create(name="clients")
    # user.groups.add(group)
    # user.save()
    # lastcode = Client.objects.last()
    # print('lastcode', lastcode.code)
    # codecl = f"{int(lastcode.code) + 1:06}"
    # ctx={
    #     'clients':Client.objects.all()[:100],
    #     'title':'List des CLient',
    #     'commerciaux':Represent.objects.all(),
    #     'lastcode':codecl
    # }
    return JsonResponse({
        'success':True
    })

def getclientdata(request):
    id=request.POST.get('id')
    client=Client.objects.get(pk=id)
    return JsonResponse({
        'name':client.name,
        'phone':client.phone,
        'address':client.address,
        'id':client.id,
        'code':client.code,
        'city':client.city,
        'region':client.region,
        'ice':client.ice,
        'rep':client.represent_id,
    })


def updateclient(request):
    print('>>>> updating client')
    clientcode=request.GET.get('clientcode')
    client=Client.objects.get(code=clientcode)
    client.name=request.GET.get('name')
    client.phone=request.GET.get('phone')
    client.address=request.GET.get('address')
    client.ice=request.GET.get('ice')
    client.code=request.GET.get('code')
    client.city=request.GET.get('city')
    client.address=request.GET.get('address')
    client.region=request.GET.get('region').lower().strip()
    client.represent_id=request.GET.get('rep')
    client.save()

    return JsonResponse({
        'success':True
    })


def getscommercialdata(request):
    id=request.POST.get('id')
    rep=Represent.objects.get(pk=id)
    return JsonResponse({
        'name':rep.name,
        'phone':rep.phone,
        'region':rep.region,
        'info':rep.info,
        'id':rep.id
    })

def updatecommercial(request):
    id=request.GET.get('updaterepid')
    name=request.GET.get('updaterepname')
    phone=request.GET.get('updaterepphone')
    region=request.GET.get('updaterepregion')
    region=request.GET.get('updaterepregion')
    rep=Represent.objects.get(pk=id)
    rep.name=name
    rep.phone=phone
    rep.region=region
    rep.save()
    
    return JsonResponse({
        'success':True
    })


def onereppage(request, id):
    rep=Represent.objects.get(pk=id)
    ctx={
        'title':f'Page de {rep.name}',
        'rep':rep,
    }
    return render(request, 'onereppage.html', ctx)

def adminpage(request):
    return render(request, 'adminpage.html')

def bonlivraisondetails(request, id):
    order=Bonlivraison.objects.get(pk=id)
    orderitems=Livraisonitem.objects.filter(bon=order, isfacture=False).order_by('product__category')
    print('orderitems', orderitems)
    reglements=PaymentClientbl.objects.filter(bons__in=[order])
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+36] for i in range(0, len(orderitems), 36)]
    ctx={
        'title':f'Bon de livraison {order.bon_no}',
        'order':order,
        'orderitems':orderitems,
        'reglements':reglements
    }
    return render(request, 'bonlivraisondetails.html', ctx)


def facturedetails(request, id):
    order=Facture.objects.get(pk=id)
    orderitems=Outfacture.objects.filter(facture=order)
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+29] for i in range(0, len(orderitems), 29)]

    ctx={
        'title':f'Facture {order.facture_no}',
        'facture':order,
        'orderitems':orderitems,
        'tva':order.tva,
        'ttc':order.total,
        'ht':round(order.total-order.tva, 2),
    }
    print(ctx)
    return render(request, 'facturedetails.html', ctx)

def avoirdetails(request, id):
    order=Avoirclient.objects.get(pk=id)
    orderitems=Returned.objects.filter(avoir=order)
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+36] for i in range(0, len(orderitems), 36)]
    ctx={
        'title':f'avoir {order.no}',
        'avoir':order,
        'orderitems':orderitems,

    }
    return render(request, 'avoirdetails.html', ctx)

def avoirsuppdetails(request, id):
    order=Avoirsupplier.objects.get(pk=id)
    orderitems=Returnedsupplier.objects.filter(avoir=order)
    # split the orderitems into chunks of 10 items
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+36] for i in range(0, len(orderitems), 36)]
    ctx={
        'title':f'avoir {order.no}',
        'avoir':order,
        'orderitems':orderitems,
    }
    return render(request, 'avoirsuppdetails.html', ctx)



def getrepswithprice(request):
    id=request.POST.get('id')
    product=Produit.objects.get(pk=id)
    repsprices=[]
    if product.repsprice:
        repsprices=json.loads(product.repsprice)
        print('get reps with price',json.loads(product.repsprice))
    return JsonResponse({
        'repswithprice': repsprices,
        'representremise':product.representremise
    })


def getclientprice(request):
    pdctid=request.POST.get('id')
    clientid=request.POST.get('clientid')
    target=request.POST.get('target')
    try:
            clientprice=Livraisonitem.objects.filter(bon__client_id=clientid, product_id=pdctid).last()
            price=clientprice.price
            remise=clientprice.remise
            return JsonResponse({
                'price':price,
                'remise':remise
            })
    except:
        return JsonResponse({
            'price':0
        })
    #if target=='bl':
    #    try:
    #        clientprice=Livraisonitem.objects.filter(client_id=clientid, product_id=id).last()
    #        price=clientprice.price
    #        remise=clientprice.remise
    #        return JsonResponse({
    #            'price':price,
    #            'remise':remise
    #        })
    #    except:
    #        return JsonResponse({
    #            'price':0
    #        })
    #else:
    #    try:
    #        clientprice=Outfacture.objects.filter(client_id=clientid, product_id=id).last()
    #        price=clientprice.price
    #        remise=clientprice.remise
    #        return JsonResponse({
    #            'price':price,
    #            'remise':remise
    #        })
    #    except:
    #        return JsonResponse({
    #            'price':0
    #        })

def listbonlivraison(request):
    today = timezone.now().date()
    thisyear=timezone.now().year
    current_time = datetime.now().strftime('%H:%M:%S')
    three_months_ago = timezone.now() - timedelta(days=90)  # Assuming 30 days per month on average

    # Query for Bonlivraison objects that have a 'date' field earlier than three months ago
    depasser = Bonlivraison.objects.filter(date__lt=three_months_ago, ispaid=False).count()
    # get only the last 100 orders of the current year
    bons= Bonlivraison.objects.filter(date__year=timezone.now().year).order_by('-bon_no')[:50]
    total=Bonlivraison.objects.filter(date__year=timezone.now().year).aggregate(Sum('total')).get('total__sum')
    ctx={
        'title':'Bons de livraison',
        'bons':bons,
        'total':total,
        'boncommand':Order.objects.filter(isdelivered=False).count(),
        'depasser':depasser,
        'reps':Represent.objects.all(),
        'today':timezone.now().date()
    }
    print(ctx)
    return render(request, 'listbonlivraison.html', ctx)

def exportbl(request):
    rep=request.GET.get('rep')
    datefrom=request.GET.get('startdate')
    dateto=request.GET.get('enddate')
    region=request.GET.get('region').lower().strip()
    print('>>>>>>', rep, datefrom, dateto)
    if rep and region:
        print('rep and region')
        bons=bons=Bonlivraison.objects.filter(salseman_id=rep,client__region=region, date__range=[datefrom, dateto])
    if rep and not region:
        print('rep and not region')
        bons=bons=Bonlivraison.objects.filter(salseman_id=rep, date__range=[datefrom, dateto])
    if not rep and region:
        print('not rep and region')
        bons=bons=Bonlivraison.objects.filter(client__region=region, date__range=[datefrom, dateto])
    if not region and not rep:
        print('nothing')
        bons=bons=Bonlivraison.objects.filter(date__range=[datefrom, dateto])

    # if rep and datefrom and dateto:
    #     print('date and rep')
    #     # convert date to datetime
    #     datefrom=datetime.strptime(datefrom, '%Y-%m-%d')
    #     dateto=datetime.strptime(dateto, '%Y-%m-%d')
    #     bons=Bonlivraison.objects.filter(salseman_id=rep, date__range=[datefrom, dateto])
    # if rep and not datefrom and not dateto:
    #     print('only rep')
    #     bons=Bonlivraison.objects.filter(salseman_id=rep, date__year=timezone.now().year)
    # if not rep and datefrom and dateto:
    #     print('only date')
    #     datefrom=datetime.strptime(datefrom, '%Y-%m-%d')
    #     dateto=datetime.strptime(dateto, '%Y-%m-%d')
    #     bons=Bonlivraison.objects.filter(date__range=[datefrom, dateto])
    # if not rep and not datefrom and not dateto:
    #     print('nothing')
    #     bons=Bonlivraison.objects.filter(date__year=timezone.now().year)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


    # Create a new Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write column headers
    ws.append(['N° Bon', 'Date', 'Client', 'Code cl.', 'total', 'region', 'ville', 'soldbl', 'rep.', 'status', 'facture', 'transport'])

    # Write product data
    for product in bons:
        ws.append([
            product.bon_no, product.date.strftime("%d/%m/%Y"), product.client.name, product.client.code, product.total, product.client.region, product.client.city, product.client.soldbl, product.salseman.name, 'R0' if product.ispaid else 'N1', 'OUI' if product.isfacture else 'NON', product.modlvrsn])

    response['Content-Disposition'] = f'attachment; filename="bonlivraison.xlsx"'
    # Save the workbook to the response
    wb.save(response)
    return response


def exportfc(request):
    rep=request.GET.get('rep')
    region=request.GET.get('region')
    datefrom=request.GET.get('startdate')
    dateto=request.GET.get('enddate')
    print('>>>>>>', rep, datefrom, dateto)
    if rep and region:
        print('rep and region')
        bons=bons=Facture.objects.filter(salseman_id=rep,client__region=region, date__range=[datefrom, dateto])
    if rep and not region:
        print('rep and not region')
        bons=bons=Facture.objects.filter(salseman_id=rep, date__range=[datefrom, dateto])
    if not rep and region:
        print('not rep and region')
        bons=bons=Facture.objects.filter(client__region=region, date__range=[datefrom, dateto])
    if not region and not rep:
        print('nothing')
        bons=bons=Facture.objects.filter(date__range=[datefrom, dateto])

    # if rep and datefrom and dateto:
    #     print('date and rep')
    #     # convert date to datetime
    #     datefrom=datetime.strptime(datefrom, '%Y-%m-%d')
    #     dateto=datetime.strptime(dateto, '%Y-%m-%d')
    #     bons=Facture.objects.filter(salseman_id=rep, date__range=[datefrom, dateto])
    # if rep and not datefrom and not dateto:
    #     print('only rep')
    #     bons=Facture.objects.filter(salseman_id=rep, date__year=timezone.now().year)
    # if not rep and datefrom and dateto:
    #     print('only date')
    #     datefrom=datetime.strptime(datefrom, '%Y-%m-%d')
    #     dateto=datetime.strptime(dateto, '%Y-%m-%d')
    #     bons=Facture.objects.filter(date__range=[datefrom, dateto])
    # if not rep and not datefrom and not dateto:
    #     print('nothing')
    #     bons=Facture.objects.filter(date__year=timezone.now().year)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


    # Create a new Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write column headers
    ws.append(['N° facture', 'Date', 'Client', 'Code cl.', 'total', 'region', 'ville', 'soldfc', 'rep.', 'status', ])

    # Write product data
    for product in bons:
        ws.append([
            product.facture_no, product.date.strftime("%d/%m/%Y"), product.client.name, product.client.code, product.total, product.client.region, product.client.city, product.client.soldfacture, product.salseman.name, 'R0' if product.ispaid else 'N1'])

    response['Content-Disposition'] = f'attachment; filename="facture.xlsx"'
    # Save the workbook to the response
    wb.save(response)
    return response





def listavoirclient(request):
    bons= Avoirclient.objects.filter(date__year=thisyear).order_by('-date')
    total=bons.aggregate(Sum('total')).get('total__sum')
    ctx={
        'title':'Avoir Client',
        'bons':bons,
        'total':total
    }
    return render(request, 'listavoirclient.html', ctx)


def listavoirsupplier(request):
    print('>>>>>>',)
    bons= Avoirsupplier.objects.all()
    total=bons.aggregate(Sum('total')).get('total__sum')
    ctx={
        'title':'Bons de livraison',
        'bons':bons,
        'total':total
    }
    return render(request, 'listavoirsupplier.html', ctx)

def listfactures(request):
    three_months_ago = timezone.now() - timedelta(days=90)
    depasser = Facture.objects.filter(date__lt=three_months_ago, ispaid=False).count()
    # get only the last 100 orders of the current year
    bons= Facture.objects.filter(date__year=timezone.now().year).order_by('-facture_no')[:50]
    total=round(Facture.objects.filter(date__year=timezone.now().year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    totaltva=round(Facture.objects.filter(date__year=timezone.now().year).aggregate(Sum('tva'))['tva__sum'] or 0, 2)
    ctx={
        'title':'List des factures',
        'bons':bons,
        'reps':Represent.objects.all(),
        'depasserfc':depasser,
        'total':total,
        'totaltva':totaltva,
        'today':timezone.now().date()
    }
    if bons:
        ctx['total']=round(bons.aggregate(Sum('total')).get('total__sum'), 2)
        ctx['totaltva']=round(bons.aggregate(Sum('tva')).get('tva__sum'), 2)
    return render(request, 'listfactures.html', ctx)


def activerproduct(request):
    uniqcode=request.GET.get('uniqcode')
    product=Produit.objects.get(uniqcode=uniqcode)
    product.isactive=True
    product.save()
    return JsonResponse({
        'success':True
    })

def desactiverproduct(request):
    uniqcode=request.GET.get('uniqcode')
    product=Produit.objects.get(uniqcode=uniqcode)
    product.isactive=False
    product.save()
    return JsonResponse({
        'success':True
    })

def generatefacture(request, id):
    livraison=Bonlivraison.objects.get(pk=id)
    items=Livraisonitem.objects.filter(bon=livraison)
    lastdate=Facture.objects.last().date
    year = timezone.now().strftime("%y")
    latest_receipt = Facture.objects.filter(
        facture_no__startswith=f'FC{year}'
    ).order_by("-facture_no").first()
    if latest_receipt:
        latest_receipt_no = int(latest_receipt.facture_no[-5:])
        receipt_no = f"FC{year}{latest_receipt_no + 1:05}"
    else:
        receipt_no = f"FC{year}00001"
    ctx={
        'livraison':livraison,
        'items':items,
        'receipt_no':receipt_no,
        'lastdate':lastdate
    }
    return render(request, 'generatefacture.html', ctx)

# genereate facture from bl
def createfacture(request):
    bon=request.POST.get('bon')
    total=request.POST.get('totalbon')
    datefacture=request.POST.get('datefacture')
    datefacture=datetime.strptime(datefacture, '%Y-%m-%d')
    orderno=request.POST.get('orderno')
    products=json.loads(request.POST.get('products'))
    livraison=Bonlivraison.objects.get(pk=bon)
    livraison.isfacture=True
    livraison.statusfc='f1'

    # watch out for negative total
    livraison.total=round(livraison.total-float(total), 2)
    livraison.save()
    thisclient=Client.objects.get(pk=livraison.client_id)
    # we substract sold bl because we generate from bin livraisso
    thisclient.soldbl=round(thisclient.soldbl-float(total), 2)
    thisclient.soldfacture=round(livraison.client.soldfacture+float(total), 2)
    thisclient.save()
    # list of ids in oredritems
    tva=round(float(total)-(float(total)/1.2), 2)
    facture=Facture.objects.create(
        bon_id=bon,
        facture_no=orderno,
        total=total,
        tva=tva,
        date=datefacture,
        client=livraison.client,
        salseman=livraison.salseman,
    )


    product_ids_to_remove = [i['productid'] for i in products]

    # Delete the matching Orderitem objects in a single transaction
    Livraisonitem.objects.filter(bon=livraison, product_id__in=product_ids_to_remove).update(isfacture=True)
    with transaction.atomic():

        for i in products:
            product=Produit.objects.get(pk=i['productid'])
            product.stockfacture=int(product.stockfacture)-int(i['qty'])
            product.save()
            Outfacture.objects.create(
                facture=facture,
                product=product,
                qty=i['qty'],
                price=i['price'],
                total=i['total'],
                remise=i['remise'],
                ref=i['ref'],
                name=i['name'],
                client=livraison.client,
                date=datefacture
            )
    # if livraison.ispaid:
    #     reglement=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison, amount__gte=float(total)).first()
    #     if reglement is not None:
    #         # give reglement bl to regl fact
    #         reglement.amount=round(float(reglement.amount)-float(total), 2)
    #         reglement.usedinfacture=True
    #         reglement.save()

    #         regfac=PaymentClientfc.objects.create(
    #             client=livraison.client,
    #             amount=float(total),
    #             mode=reglement.mode,
    #             factures=facture,
    #             amountofeachbon=reglement.echance,
    #             npiece=reglement.npiece
    #         )
    #     else:
    #         reglements=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison)
    #         reglfacture=0
    #         for i in reglements:
    #             if reglfacture==float(total):
    #                 break
    #             if i.amount<=round(float(total)-reglfacture, 2):
    #                 regfac=PaymentClientfc.objects.create(
    #                     client=livraison.client,
    #                     amount=i.amount,
    #                     mode=i.mode,
    #                     factures=facture,
    #                     echance=i.echance,
    #                     npiece=i.npiece
    #                 )
    #                 reglfacture+=i.amount
    #                 i.amount=0
    #                 i.usedinfacture=True
    #                 i.save()
    #             else:
    #                 wanted=round(float(total)-reglfacture, 2)
    #                 regfac=PaymentClientfc.objects.create(
    #                     client=livraison.client,
    #                     amount=wanted,
    #                     mode=i.mode,
    #                     factures=facture,
    #                     echance=i.echance,
    #                     npiece=i.npiece
    #                 )
    #                 reglfacture+=wanted
    #                 i.amount=round(float(i.amount)-wanted)
    #                 i.usedinfacture=True
    #                 i.save()

    #     # Facturesregle.objects.create(
    #     #     payment=regfac,
    #     #     bon=facture,
    #     #     amount=float(total)

    #     # )
    #     # finish substraction form reglement bon
    #     facture.ispaid=True
    #     facture.save()
    # elif livraison.rest > 0:
    #     if float(total)==float(livraison.total):
    #         reglements=PaymentClientbl.objects.filter
    #     reglement=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison, amount__gte=float(total)).first()
    #     if reglement is not None:
    #         reglement.amount=round(float(reglement.amount)-float(total), 2)
    #         reglement.usedinfacture=True
    #         reglement.save()
    #         regfac=PaymentClientfc.objects.create(
    #             client=livraison.client,
    #             amount=float(total),
    #             mode=reglement.mode,
    #             factures=facture,
    #             echance=reglement.echance,
    #             npiece=reglement.npiece
    #         )
    #         facture.ispaid=True
    #         facture.save()
    #         # new bon rest
    #         #total reglements

    #     else:
    #         # if onre of the reglements is not gte total of facture
    #         # we need to iterate over regl and sum up the regl
    #         reglements=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison)
    #         reglfacture=0
    #         for i in reglements:
    #             if reglfacture==float(total):
    #                 break
    #             reglfacture+=i.amount
    #             i.amount=0
    #             i.usedinfacture=True
    #             i.save()
    #     totalreglements=PaymentClientbl.objects.filter(client=livraison.client,bons__in=livraison).aggregate(Sum('amount'))['amount__sum']
    #     livraison.rest=round(livraison.total-totalreglements, 2)
    #     livraison.save()
    #     livraison.client.soldbl=round(livraison.client.soldbl-float(totalreglements), 2)
    # else:
    #     livraison.client.soldfacture=round(livraison.client.soldfacture+float(total), 2)
    #     livraison.client.save()
    return JsonResponse({
        'success':True
    })


def degenerer(request):
    bonid=request.POST.get('bonid')
    livraison=Bonlivraison.objects.get(pk=bonid)
    Livraisonitem.objects.filter(bon=livraison).update(isfacture=False)
    # delete facture and outfacture
    facture=Facture.objects.get(bon=livraison)
    outfactures=Outfacture.objects.filter(facture=facture)
    print(livraison, facture, outfactures)

    # products=Produit.objects.get(pk=i.product_id) for i in outfactures

    for i in outfactures:

        product=Produit.objects.get(pk=i.product_id)
        product.stockfacture=int(product.stockfacture)+int(i.qty)
        product.save()
        i.delete()
    livraison.isfacture=False
    livraison.statusfc='b1'
    livraison.total=round(livraison.total+float(facture.total), 2)
    livraison.save()
    facture.delete()
    livraison.client.soldbl=round(livraison.client.soldbl+float(facture.total), 2)
    livraison.client.soldfacture=round(livraison.client.soldfacture-float(facture.total), 2)
    livraison.client.save()
    return JsonResponse({
        'html':render(request, 'bonlivraisonbody.html', {'order':livraison}).content.decode('utf-8')
    })

def modifierlivraison(request, id):
    livraison=Bonlivraison.objects.get(pk=id)
    items=Livraisonitem.objects.filter(bon=livraison, isfacture=False)
    ctx={
        'title':'Modifier '+livraison.bon_no,
        'livraison':livraison,
        'items':items,
        'commercials':Represent.objects.all(),
        # 'products':Produit.objects.all(),
        # 'clients':Client.objects.all(),
    }
    return render(request, 'modifierlivraison.html', ctx)

def modifieravoir(request, id):
    avoir=Avoirclient.objects.get(pk=id)
    items=Returned.objects.filter(avoir=avoir)
    ctx={
        'avoir':avoir,
        'items':items,
        'commercials':Represent.objects.all(),
    }
    return render(request, 'modifieravoir.html', ctx)


def modifierfacture(request, id):
    facture=Facture.objects.get(pk=id)
    items=Outfacture.objects.filter(facture=facture)
    ctx={
        'facture':facture,
        'items':items,
        'products':Produit.objects.all(),
        'commercials':Represent.objects.all(),
        'clients':Client.objects.all(),
    }
    return render(request, 'modifierfacture.html', ctx)


def updatebonlivraison(request):
    id=request.POST.get('bonid')
    livraison=Bonlivraison.objects.get(pk=id)
    client=Client.objects.get(pk=request.POST.get('clientid'))
    totalbon=request.POST.get('totalbon')
    transport=request.POST.get('transport')
    note=request.POST.get('note')
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(f'{datebon}', '%Y-%m-%d')

    thisclient=livraison.client
    if livraison.client==client:
        print('same client', float(thisclient.soldtotal), float(livraison.total), float(totalbon))
        thisclient.soldtotal=round(float(thisclient.soldtotal)-float(livraison.total)+float(totalbon), 2)
        thisclient.soldbl=round(float(thisclient.soldbl)-float(livraison.total)+float(totalbon), 2)
        thisclient.save()
    else:
        print('not same')
        thisclient.soldtotal=round(float(thisclient.soldtotal)-float(livraison.total), 2)
        thisclient.soldbl=round(float(thisclient.soldbl)-float(livraison.total), 2)
        thisclient.save()
        print('old', thisclient.soldtotal)
        client.soldtotal+=float(totalbon)
        client.soldbl+=float(totalbon)
        client.save()
        print('new', client.soldtotal)
    livraison.modlvrsn=transport
    livraison.client=client
    livraison.note=note
    livraison.salseman_id=request.POST.get('repid')
    livraison.total=totalbon
    livraison.date=datebon
    livraison.bon_no=request.POST.get('orderno')
    livraison.save()
    items=Livraisonitem.objects.filter(bon=livraison)
    # update this items
    for i in items:
        product=Produit.objects.get(pk=i.product_id)
        product.stocktotal=int(product.stocktotal)+int(i.qty)
        product.save()
        i.delete()

    print('client:', livraison.client.id)
    for i in json.loads(request.POST.get('products')):
        # ABORTER FOR NOW
        # clientpricehistory=Clientprices.objects.filter(client_id=livraison.client.id, product_id=i['productid']) or None
        # if clientpricehistory:
        #     print('clientpricehistory exist')
        #     clientpricehistory.update(price=i['price'])
        # else:
        #     print('clientpricehistory not exist')
        #     Clientprices.objects.create(client_id=livraison.client.id, product_id=i['productid'], price=i['price'])


        qty=int(i['qty'])
        product=Produit.objects.get(pk=i['productid'])
        product.stocktotal=int(product.stocktotal)-qty

        product.save()

        # create new livraison items
        Livraisonitem.objects.create(
            bon=livraison,
            remise=i['remise'],
            name=i['name'],
            ref=i['ref'],
            product=product,
            qty=qty,
            price=i['price'],
            total=i['total'],
            date=datebon
        )

    return JsonResponse({
        'success':True
    })


def updatebonfacture(request):
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    client=Client.objects.get(pk=request.POST.get('clientid'))
    facture=Facture.objects.get(pk=request.POST.get('factureid'))
    totalbon=request.POST.get('totalbon')
    thisclient=facture.client
    if facture.client==client:
        print('same client', float(thisclient.soldtotal), float(facture.total), float(totalbon))
        thisclient.soldtotal=round(float(thisclient.soldtotal)-float(facture.total)+float(totalbon), 2)
        thisclient.soldfacture=round(float(thisclient.soldfacture)-float(facture.total)+float(totalbon), 2)
        thisclient.save()
    else:
        print('not same')
        thisclient.soldtotal=round(float(thisclient.soldtotal)-float(facture.total), 2)
        thisclient.soldfacture=round(float(thisclient.soldfacture)-float(facture.total), 2)
        thisclient.save()
        print('old', thisclient.soldtotal)
        client.soldtotal+=float(totalbon)
        client.soldfacture+=float(totalbon)
        client.save()
        print('new', client.soldtotal)
    tva=round(float(totalbon)-(float(totalbon)/1.2), 2)
    facture.tva=tva
    facture.client=client
    facture.salseman_id=request.POST.get('repid')
    facture.total=totalbon
    facture.facture_no=request.POST.get('orderno')
    facture.note=request.POST.get('note')
    facture.date=datebon
    facture.save()
    items=Outfacture.objects.filter(facture=facture)
    # update this items
    for i in items:
        product=Produit.objects.get(pk=i.product_id)
        product.stockfacture=int(product.stockfacture)+int(i.qty)
        product.save()
        i.delete()

    print('client:', facture.client.id)
    with transaction.atomic():
        for i in json.loads(request.POST.get('products')):
            # update price in facture
            # clientpricehistory=Clientprices.objects.filter(client_id=facture.client.id, product_id=i['productid']) or None
            # if clientpricehistory:
            #     print('clientpricehistory exist')
            #     clientpricehistory.update(price=i['price'])
            # else:
            #     print('clientpricehistory not exist')
            #     Clientprices.objects.create(client_id=facture.client.id, product_id=i['productid'], price=i['price'])


            qty=int(i['qty'])
            product=Produit.objects.get(pk=i['productid'])
            product.stockfacture=int(product.stockfacture)-qty

            product.save()

            # create new livraison items
            Outfacture.objects.create(
                facture=facture,
                remise=i['remise'],
                name=i['name'],
                ref=i['ref'],
                product=product,
                qty=qty,
                price=i['price'],
                total=i['total'],
                client=client,
                date=datebon,
            )

    return JsonResponse({
        'success':True
    })


def listreglementbl(request):
    reglements=PaymentClientbl.objects.all().order_by('-id')[:50]
    print('lenreg', len(reglements))
    ctx={
        'title':'List des reglements CL BL',
        'reglements':reglements,
        'today':timezone.now().date()
    }
    if reglements:
        ctx['total']=round(reglements.aggregate(Sum('amount'))['amount__sum'], 2)

    return render(request, 'listreglementbl.html', ctx)


def listreglementsupp(request):
    reglements=PaymentSupplier.objects.all().order_by('-id')[:50]
    ctx={
        'title':'List des reglements Fournisseur',
        'reglements':reglements,
        'suppliers':Supplier.objects.all(),
        'today':timezone.now().date()
    }
    if reglements:
        ctx['total']=round(reglements.aggregate(Sum('amount'))['amount__sum'], 2)
    return render(request, 'listreglementsupp.html', ctx)


def listreglementfc(request):
    reglements=PaymentClientfc.objects.filter(date__year=thisyear).order_by('-date')[:50]

    ctx={
        'title':'List des reglements CL fc',
        'reglements':reglements,
        'today':timezone.now().date(),
        'total':round(PaymentClientfc.objects.filter(date__year=thisyear).aggregate(Sum('amount'))['amount__sum'] or 0, 2)


    }
    if reglements:
        ctx['total']=round(reglements.aggregate(Sum('amount'))['amount__sum'], 2)


    return render(request, 'listreglementfc.html', ctx)


def getclientbons(request):
    clientid=request.POST.get('clientid')
    print('>>>', clientid)
    bons=Bonlivraison.objects.filter(client_id=clientid)[:50]
    total=round(Bonlivraison.objects.filter(client_id=clientid).aggregate(Sum('total')).get('total__sum')or 0,  2)
    trs=''
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)" {"checked" if i.reglements.exists() else ""}></td></tr>'


    return JsonResponse({
        'trs':trs,
        'total':total,
        'soldbl':Client.objects.get(pk=clientid).soldbl
    })

def filternonreglr(request):
    clientid=request.GET.get('clientid')
    bons=Bonlivraison.objects.filter(client_id=clientid, ispaid=False).order_by('-id')[:50]
    total=round(Bonlivraison.objects.filter(client_id=clientid, ispaid=False).aggregate(Sum('total')).get('total__sum') or 0, 2)
    trs=''
    for i in bons:
        trs+=f'<tr class="blreglrow nr" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">NR</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)" {"checked" if i.reglements.exists() else ""}></td></tr>'

    return JsonResponse({
        'trs':trs,
        'total':total,
        'soldbl':Client.objects.get(pk=clientid).soldbl
    })

def filternonreglrfc(request):
    clientid=request.GET.get('clientid')
    print(clientid)
    bons=Facture.objects.filter(client_id=clientid, ispaid=False).order_by('-id')[:50]
    total=round(Facture.objects.filter(client_id=clientid, ispaid=False).aggregate(Sum('total')).get('total__sum')or 0, 2)
    trs=''
    for i in bons:
        trs+=f'<tr class="fcreglrow nr" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">NR</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'

    return JsonResponse({
        'trs':trs,
        'total':total,
        'soldfc':Client.objects.get(pk=clientid).soldfacture
    })

def getclientfactures(request):
    clientid=request.POST.get('clientid')
    bons=Facture.objects.filter(client_id=clientid).order_by('-id')[:50]
    trs=''
    for i in bons:
        trs+=f'<tr  class="fcreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td>{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)" {"checked" if i.reglementsfc.exists() else ""}></td></tr>'
    return JsonResponse({
        'trs':trs,
        'totalfactures':round(Facture.objects.filter(client_id=clientid).aggregate(Sum('total'))['total__sum']or 0, 2),
        'soldfactureregl':Client.objects.get(pk=clientid).soldfacture
    })

def reglefactures(request):
    clientid=request.POST.get('clientid')
    client=Client.objects.get(pk=clientid)
    bons=json.loads(request.POST.get('bons'))
    mantant=json.loads(request.POST.get('mantant'))
    mode=json.loads(request.POST.get('mode'))
    npiece=json.loads(request.POST.get('npiece'))
    date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=json.loads(request.POST.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]
    livraisons=Facture.objects.filter(pk__in=bons)
    livraisons.update(ispaid=True)
    livraisons.update(statusreg='f1')
    # livraisonstotals=0
    totalmantant=sum(mantant)
    # for i in livraisons:
    #     if i.rest>0:
    #         livraisonstotals=round(livraisonstotals+i.rest, 2)
    #     else:
    #         livraisonstotals=round(livraisonstotals+i.total, 2)

    # print(totalmantant, livraisonstotals, livraisons)
    # # update client sold
    # # case1: 5000==5000:
    # if totalmantant==livraisonstotals:
    #     print('case1')
    #     livraisons.update(ispaid=True)
    #     livraisons.update(rest=0)
    # # case2: 5000>4500:
    # if totalmantant>livraisonstotals:
    #     print('case2')
    #     diff=totalmantant-livraisonstotals
    #     livraisons.update(ispaid=True)
    #     livraisons.update(rest=0)

    #     # else:
    #     #     livraisons.update(ispaid=True)
    #     # sub diff from client ref
    # # case3: 5000<6000:
    # if totalmantant<livraisonstotals:
        # print('case3')
        # amount=totalmantant
        # for i in livraisons:
        #     if amount<=0:
        #         break
        #     else:
        #         # if facture has rest
        #         if i.rest>0:
        #             if amount>=i.rest:
        #                 i.rest=0
        #                 i.ispaid=True
        #                 print(f'facture {i.facture_no} has rest is paid here')
        #                 i.save()
        #                 amount-=i.rest
        #             else:
        #                 print(f'facture {i.facture_no} has rest has rest here')
        #                 i.rest=round(i.rest-amount, 2)
        #                 i.save()
        #                 break
        #         else:
        #             if amount>=i.total:
        #                 amount-=i.total
        #                 print(f'facture {i.facture_no} is paid here')
        #                 i.ispaid=True
        #                 i.save()
        #             else:
        #                 print(f'facture {i.facture_no} hasrestof {round(i.total-amount, 2)}')
        #                 i.rest=round(i.total-amount, 2)
        #                 i.save()
        #                 break
        # print('rest of amount', amount)
    for m, mod, np, ech in zip(mantant, mode, npiece, echeance):
        tva=round(m-(m/1.2), 2)
        regl=PaymentClientfc.objects.create(
            client_id=clientid,
            amount=m,
            date=date,
            tva=tva,
            echance=ech,
            mode=mod,
            npiece=np,

        )
        regl.factures.set(livraisons)
        # storing factures in facturesregle
        # for i in livraisons:
        #     Facturesregle.objects.create(
        #         payment=regl,
        #         bon=i,
        #         amount=m
        #     )


    client.soldtotal=round(float(client.soldtotal)-float(totalmantant), 2)
    client.soldfacture=round(float(client.soldfacture)-float(totalmantant), 2)
    client.save()
    return JsonResponse({
        "success":True
    })


def reglebons(request):
    clientid=request.POST.get('clientid')
    client=Client.objects.get(pk=clientid)
    bons=json.loads(request.POST.get('bons'))
    mantant=json.loads(request.POST.get('mantant'))
    mode=json.loads(request.POST.get('mode'))
    npiece=json.loads(request.POST.get('npiece'))
    date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=json.loads(request.POST.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]
    livraisons=Bonlivraison.objects.filter(pk__in=bons)
    livraisons.update(ispaid=True)
    livraisons.update(statusreg='r0')
    totalmantant=sum(mantant)
    # for i in livraisons:
    #     if i.rest>0:
    #         livraisonstotals=round(livraisonstotals+i.rest, 2)
    #     else:
    #         livraisonstotals=round(livraisonstotals+i.total, 2)

    # print(totalmantant, livraisonstotals, mantant, mode, npiece, echeance)

    # # update client sold
    # # case1: 5000==5000:
    # amountofeachbon=[]
    # if totalmantant==livraisonstotals:
    #     print('case1')
    #     livraisons.update(ispaid=True)
    #     livraisons.update(rest=0)
    #     for i in livraisons:
    #         amountofeachbon.append(i.total)
    # # case2: 5000>4500:
    # if totalmantant>livraisonstotals:
    #     print('case2')
    #     diff=totalmantant-livraisonstotals
    #     livraisons.update(ispaid=True)
    #     livraisons.update(rest=0)
    #     for i in livraisons:
    #         amountofeachbon.append(i.total)
    #     # else:
    #     #     livraisons.update(ispaid=True)
    #     # sub diff from client ref
    # # case3: 5000<6000:
    # if totalmantant<livraisonstotals:
    #     print('case3')
    #     amount=totalmantant
    #     for i in livraisons:
    #         if amount<=0:
    #             break
    #         else:
    #             if i.rest>0:
    #                 if amount>=i.rest:
    #                     i.rest=0
    #                     i.ispaid=True
    #                     i.save()
    #                     amount-=i.rest
    #                     amountofeachbon.append(amount)
    #                 else:

    #                     i.rest=round(i.rest-amount, 2)
    #                     i.save()
    #                     amountofeachbon.append(i.rest)
    #                     break
    #             else:
    #                 if amount>=i.total:
    #                     amount-=i.total
    #                     i.ispaid=True
    #                     i.save()
    #                     amountofeachbon.append(amount)
    #                 else:
    #                     i.rest=round(i.total-amount, 2)
    #                     amountofeachbon.append(amount)
    #                     i.save()
    #                     break
    #     print('rest of amount', amount)


    for m, mod, np, ech in zip(mantant, mode, npiece, echeance):
        regl=PaymentClientbl.objects.create(
            client_id=clientid,
            amount=m,
            date=date,
            echance=ech,
            mode=mod,
            npiece=np
        )
        regl.bons.set(livraisons)
        # for i in livraisons:
        #     Bonsregle.objects.create(
        #         payment=regl,
        #         bon=i,
        #         amount=m
        #     )

    client.soldtotal=round(float(client.soldtotal)-float(totalmantant), 2)
    client.soldbl=round(float(client.soldbl)-float(totalmantant), 2)
    client.save()
    return JsonResponse({
        "success":True
    })



def checknpiece(request):
    npiece=request.POST.get('npiece')
    if PaymentClientbl.objects.filter(npiece=npiece).exists():
        return JsonResponse({
            'exist':True
        })
    return JsonResponse({
        'exist':False
    })


def viewreglement(request, id):
    reglement=PaymentClientbl.objects.get(pk=id)
    reglements=Bonsregle.objects.filter(payment=reglement)
    ctx={
        'title':'Reglement',
        'reglement':reglement,
        'reglements':reglements
    }
    return render(request, 'viewreglement.html', ctx)

def viewreglementfc(request, id):
    reglement=PaymentClientfc.objects.get(pk=id)
    ctx={
        'title':'Reglement',
        'reglement':reglement,
        'reglfc':True
    }
    return render(request, 'viewreglement.html', ctx)



def situationcl(request):
    ctx={
        'title':'Situation des clients',
        'today':timezone.now().date()
    }
    return render(request, 'situationcl.html', ctx)

def situationsupplier(request):
    ctx={
        'title':'Situation des Fouurnisseurs',
        'suppliers':Supplier.objects.all()
    }
    return render(request, 'situationsupplier.html', ctx)

def situationclfc(request):
    ctx={
        'title':'Situation des clients (Factures)',
        'today':timezone.now().date()
    }
    return render(request, 'situationclfc.html', ctx)

def recevoirexcel(request):
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    df = df.fillna(0)
    trs=''
    totalbon=0
    for d in df.itertuples():
        try:
            ref = d.ref.lower().strip()
        except:
            ref=d.ref
        #name=d.name
        qty=d.qty
        prixachat=float(d.prixachat)
        # prixachat 12,0 make it 12,00
        remise=0 if pd.isna(d.remise) else int(d.remise)
        devise=0 if pd.isna(d.devise) else d.devise
        prixnet=round(prixachat-(prixachat*float(remise)/100), 2)
        if remise==0:
            total=round(prixachat*qty, 2)
        else:
            total=round(prixnet*qty, 2)
        totalbon+=total
        # make total 2 digits after point

        product=Produit.objects.get(ref=ref)
        
        print('product exists')
        trs+=f"""<tr>
            <td class="ref">{ref}</td>
            <td class="name">{product.name}</td>
            <td>{product.buyprice if product.buyprice else 0}</td>
            <td>{product.stocktotal if product.stocktotal else 0}</td>
            <td>
            <input style="width:65px;" type="number" class="devise" value="{devise:.2f}">
            </td>
            <td><input style="width:65px;" type="number" class="calculatebonachat qty" onkeyup="calculate(event)" name="qtybonachat" value="{qty}"></td>
            <td><input style="width:65px;" type="number" class="calculatebonachat price" onkeyup="calculate(event)" name="pricebonachat" value="{prixachat:.2f}"></td>
            <td><input style="width:65px;" type="number" class="calculatebonachat remise" onkeyup="calculate(event)" name="remise" value="{remise}"></td>
            <td class="total">{total:.2f}</td>
            <input type="hidden" name="productid" value="{product.id}">
        </tr>"""
    return JsonResponse({
        'trs':trs,
        'totalbon':totalbon
    })


def avoirclient(request):

    ctx={
            'title':'Avoir client',
            'commercials':Represent.objects.all(),
            #'receipt_no':receipt_no
        }
    return render(request, 'avoirclient.html', ctx)



def addavoirclient(request):
    clientid=request.POST.get('clientid')
    repid=request.POST.get('repid')
    products=request.POST.get('products')
    totalbon=request.POST.get('totalbon')
    mode=request.POST.get('mode')
    isfacture=True if mode=='facture' else False
    # orderno
    #orderno=request.POST.get('orderno')
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    client=Client.objects.get(pk=clientid)

    year = timezone.now().strftime("%y")

    prefix = f'AV{year}'
    try:
        avoirclients = Avoirclient.objects.filter(no__startswith=prefix).last()
        latest_receipt_no = int(avoirclients.no.split('/')[1])
        receipt_no = f"AV{year}/{latest_receipt_no + 1}"
    except:
        receipt_no = f"AV{year}/1"
    print(receipt_no, clientid, repid, totalbon, datebon, isfacture)
    try:
        
        avoir=Avoirclient.objects.create(
            no=receipt_no,
            client_id=clientid,
            representant_id=repid,
            total=totalbon,
            date=datebon,
            avoirfacture=isfacture
        )
        
        for i in json.loads(products):
            product=Produit.objects.get(pk=i['productid'])
            if isfacture:
                product.stockfacture=int(product.stockfacture)+int(i['qty'])
            product.stocktotal=int(product.stocktotal)+int(i['qty'])
            product.save()
            Returned.objects.create(
                avoir=avoir,
                product=product,
                qty=i['qty'],
                remise=0 if i['remise']=="" else i['remise'],
                price=0 if i['price']=="" else i['price'],
                total=i['total'],
            )    
        client.soldtotal=round(float(client.soldtotal)-float(totalbon), 2)
        if isfacture:
            client.soldfacture=round(float(client.soldfacture)-float(totalbon), 2)
        else:
            client.soldbl=round(float(client.soldbl)-float(totalbon), 2)

        client.save()
    except Exception as e:
        print('>>error av cl:', e)

    # increment it
    return JsonResponse({
        'html':render(request, 'avoirclient.html', {
            'title':'Bon de livraison',
            #'clients':Client.objects.all(),
            #'products':Produit.objects.all(),
            'commercials':Represent.objects.all(),
            #'receipt_no':receipt_no
        }).content.decode('utf-8')
    })

def avoirsupplier(request):
    
    ctx={
            'title':'Avoir Fournisseur',
        }
    return render(request, 'avoirsupplier.html', ctx)



def addavoirsupp(request):
    supplierid=request.POST.get('supplierid')
    products=request.POST.get('products')
    totalbon=request.POST.get('totalbon')
    # orderno
    avoirfacture=True if request.POST.get('mode')=='facture' else False
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    supplier=Supplier.objects.get(pk=supplierid)
    year = timezone.now().strftime("%y")

    prefix = f'FAV{year}'
    try:
        avoirsuppliers = Avoirsupplier.objects.filter(no__startswith=prefix).order_by('no').last()
        latest_receipt_no = int(avoirsuppliers.no.split('/')[1])
        receipt_no = f"FAV{year}/{latest_receipt_no + 1}"
    except:
        receipt_no = f"FAV{year}/1"
    avoir=Avoirsupplier.objects.create(
        no=receipt_no,
        supplier_id=supplierid,
        total=totalbon,
        date=datebon,
        avoirfacture=avoirfacture

    )
    supplier.rest-=float(totalbon)
    supplier.save()
    with transaction.atomic():
        for i in json.loads(products):
            product=Produit.objects.get(pk=i['productid'])
            product.stocktotal=int(product.stocktotal)-int(i['qty'])
            if avoirfacture:
                product.stockfacture=int(product.stockfacture)-int(i['qty'])
            product.save()
            Returnedsupplier.objects.create(
                avoir=avoir,
                product=product,
                qty=i['qty'],
                price=i['price'],
                total=i['total'],
            )


    # increment it
    return JsonResponse({
        'success':True
    })


def checkbl(request, id):
    client=Client.objects.get(pk=id)
    print(Bonlivraison.objects.filter(client=client).first())
    if Bonlivraison.objects.filter(client=client).first() is not None:
        return JsonResponse({
            'hasbl':True
        })
    else:
        return JsonResponse({
            'hasbl':False
        })

def checkfacture(request, id):
    client=Client.objects.get(pk=id)
    print(client.name)
    print(Facture.objects.filter(client=client).first())
    if Facture.objects.filter(client=client).first() is not None:
        # get products
        return JsonResponse({
            'hasfacture':True
        })
    else:
        return JsonResponse({
            'hasfacture':False
        })

def relevclient(request):
    clientid=request.POST.get('clientid')
    client=Client.objects.get(pk=clientid)
    startdate=request.POST.get('datefrom')
    enddate=request.POST.get('dateto')
    res=req.get('http://localserver/products/sendrelevclient', {'clientcode':client.code, 'datefrom':startdate, 'dateto':enddate})
    return JsonResponse(res.json()) 
    #startdate = datetime.strptime(startdate, '%Y-%m-%d')
    #enddate = datetime.strptime(enddate, '%Y-%m-%d')
    #avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=False, date__range=[startdate, enddate])
    #reglementsbl=PaymentClientbl.objects.filter(client_id=clientid, date__range=[startdate, enddate])
    #bons=Bonlivraison.objects.filter(client_id=clientid, date__range=[startdate, enddate], total__gt=0)
    # totalcredit=round(avoirs.aggregate(Sum('total'))['total__sum'], 2)+round(reglementsbl.aggregate(Sum('amount'))['amount__sum'], 2)
    # totaldebit=round(bons.aggregate(Sum('total'))['total__sum'], 2)
    # sold=round(totaldebit-totalcredit, 2)

    # chain all the data based on dates
    # first get all dates
    #releve = chain(*[
    #((bon, 'Bonlivraison') for bon in bons),
    #((avoir, 'Avoirclient') for avoir in avoirs),
    #((reglementbl, 'PaymentClientbl') for reglementbl in reglementsbl),
    #])

    # Sort the items by date
    #sorted_releve = sorted(releve, key=lambda item: item[0].date)


    #return JsonResponse({
    #    'html':render(request, 'relevecl.html', {
    #        'releve':[sorted_releve[i:i+32] for i in range(0, len(sorted_releve), 32)],
    #        'client':client,
    #        'startdate':startdate,
    #        'enddate':enddate,

    #    }).content.decode('utf-8')
    #})


def relevsupplier(request):
    supplierid=request.POST.get('supplierid')
    supplier=Supplier.objects.get(pk=supplierid)
    startdate=request.POST.get('datefrom')
    enddate=request.POST.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    avoirs=Avoirsupplier.objects.filter(supplier_id=supplierid, avoirfacture=False, date__range=[startdate, enddate])
    reglementsbl=PaymentSupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate])

    bons=Itemsbysupplier.objects.filter(supplier_id=supplierid, date__range=[startdate, enddate])
    print('rr', supplierid)
    # chain all the data based on dates
    # first get all dates
    releve = chain(*[
    ((bon, 'Itemsbysupplier') for bon in bons),
    ((avoir, 'Avoirsupplier') for avoir in avoirs),
    ((reglementbl, 'Paymentsupplier') for reglementbl in reglementsbl),
    ])

    # Sort the items by date
    sorted_releve = sorted(releve, key=lambda item: item[0].date)


    return JsonResponse({
        'html':render(request, 'relevesupp.html', {
            'releve':sorted_releve,
            'supplier':supplier,
            'startdate':startdate,
            'enddate':enddate,

        }).content.decode('utf-8')
    })



def relevclientfc(request):
    clientid=request.POST.get('clientid')
    client=Client.objects.get(pk=clientid)
    startdate=request.POST.get('datefrom')
    enddate=request.POST.get('dateto')
    res=req.get('http://localserver/products/sendrelevclientfc', {'clientcode':client.code, 'datefrom':startdate, 'dateto':enddate})
    return JsonResponse(res.json())
    #startdate = datetime.strptime(startdate, '%Y-%m-%d')
    #enddate = datetime.strptime(enddate, '%Y-%m-%d')
    #avoirs=Avoirclient.objects.filter(client_id=clientid, avoirfacture=True, date__range=[startdate, enddate])
    #reglementsfc=PaymentClientfc.objects.filter(client_id=clientid, date__range=[startdate, enddate])

    #bons=Facture.objects.filter(client_id=clientid, date__range=[startdate, enddate])

    # chain all the data based on dates
    # first get all dates
    #releve = chain(*[
    #((bon, 'Facture') for bon in bons),
    #((avoir, 'Avoirclient') for avoir in avoirs),
    #((reglementfc, 'PaymentClientfc') for reglementfc in reglementsfc),
    #])

    # Sort the items by date
    #sorted_releve = sorted(releve, key=lambda item: item[0].date)
    #for i in sorted_releve:
    #    print(i)

    #return JsonResponse({
    #    'html':render(request, 'releveclfc.html', {
    #        'releve':[sorted_releve[i:i+26] for i in range(0, len(sorted_releve), 26)],
    #        'client':client,
    #        'startdate':startdate,
    #        'enddate':enddate,

     #   }).content.decode('utf-8')
    #})

def getclientrep(request, id):
    client=Client.objects.get(pk=id)
    print(client.represent_id)
    return JsonResponse({
        'id':client.represent_id
    })

def listbonachat(request):
    thisyear=timezone.now().year
    bons=Itemsbysupplier.objects.filter(date__year=thisyear).order_by('-id')[:50]
    ctx={
        'title':'List des bon achat',
        'bons':bons
    }
    if bons:
        ctx['total']=round(Itemsbysupplier.objects.all().aggregate(Sum('total'))['total__sum'], 2)
        ctx['totaltva']=round(Itemsbysupplier.objects.all().aggregate(Sum('tva'))['tva__sum'], 2)
    return render(request, 'listbonachat.html', ctx)

def listboncommnd(request):
    today = timezone.now().date()
    thisyear=timezone.now().year
    current_time = datetime.now().strftime('%H:%M:%S')
    orders=Order.objects.filter(date__year=thisyear).order_by('-id')[:50]
    ctx={
        'title':'List des bon commnd',
        'bons':orders,
        'today':timezone.now().date()
    }
    if orders:
        ctx['total']=round(Order.objects.all().aggregate(Sum('total'))['total__sum'], 2)
    return render(request, 'listboncommnd.html', ctx)

def bonachatdetails(request, id):
    bon=Itemsbysupplier.objects.get(pk=id)
    items=Stockin.objects.filter(nbon=bon)
    payments=PaymentSupplier.objects.filter(bons__in=[bon])
    print('paym', payments)
    print(items)
    ctx={
        'title':f'Bon achat {bon.nbon}',
        'bon':bon,
        'items':items,
        'payments':payments
    }
    return render(request, 'bonachatdetails.html', ctx)

def modifierbonachat(request, id):
    bon=Itemsbysupplier.objects.get(pk=id)
    items=Stockin.objects.filter(nbon=bon)
    ctx={
        'title':'Modifier bon achat',
        'bon':bon,
        'items':items,
        'suppliers':Supplier.objects.all()
    }
    return render(request, 'modifierbonachat.html', ctx)


def updatebonachat(request):
    id=request.POST.get('bonid')
    bon=Itemsbysupplier.objects.get(pk=id)
    bon.date=datetime.strptime(request.POST.get('datebon'), '%Y-%m-%d')
    bon.nbon=request.POST.get('orderno')
    isfacture= True if request.POST.get('mode')=='facture' else False
    totalbon=request.POST.get('totalbon')
    supplier=Supplier.objects.get(pk=request.POST.get('supplierid'))
    thissupplier=bon.supplier
    if bon.supplier==supplier:
        print('same supplier', float(thissupplier.rest), float(bon.total), float(totalbon))
        thissupplier.rest=round(float(thissupplier.rest)-float(bon.total)+float(totalbon), 2)
        thissupplier.save()
    else:
        print('not same')
        thissupplier.rest=round(float(thissupplier.rest)-float(bon.total), 2)
        thissupplier.save()
        print('old', thissupplier.rest)
        supplier.rest=round(float(supplier.rest)+float(totalbon), 2)
        supplier.save()
        print('new', supplier.rest)
    # bon.supplier.rest=float(bon.supplier.rest)-float(bon.total)
    # bon.supplier.save()
    items=Stockin.objects.filter(nbon=bon)
    # update this items
    for i in items:
        product=i.product
        print('removing from total')
        product.stocktotal=int(product.stocktotal)-int(i.quantity)
        if bon.isfacture:
            print('removing from facture')
            product.stockfacture=int(product.stockfacture)-int(i.quantity)
        product.save()
        i.delete()

    bon.supplier=supplier
    bon.total=totalbon
    bon.nbon=request.POST.get('orderno')
    bon.isfacture=isfacture
    bon.save()

    with transaction.atomic():
        for i in json.loads(request.POST.get('products')):


            qty=0 if i['qty']=="" else int(i['qty'])
            product=Produit.objects.get(pk=i['productid'])
            print('>>>>>>>adding fc')
            product.stocktotal=int(product.stocktotal)+qty
            if isfacture:
                print('>>>>>>>adding fc')
                product.stockfacture=int(product.stockfacture)+qty
            product.save()
            # create new livraison items
            Stockin.objects.create(
                nbon=bon,
                supplier=supplier,
                remise=0 if i['remise']=="" else i['remise'],
                devise=0 if i['devise']=="" else i['devise'],
                product=product,
                date=datetime.strptime(request.POST.get('datebon'), '%Y-%m-%d'),
                quantity=qty,
                price=0 if i['price']=="" else i['price'],
                total=0 if i['total']=="" else i['total'],
                facture=isfacture
            )
            totalprices=Stockin.objects.filter(product=product).aggregate(Sum('total'))['total__sum'] or 0
            totalqty=Stockin.objects.filter(product=product).aggregate(Sum('quantity'))['quantity__sum'] or 0
            print(totalprices, totalqty)
            product.coutmoyen=round(totalprices/totalqty, 2)
            product.save()

    return JsonResponse({
        'success':True
    })

def getsuppbons(request):
    supplierid=request.POST.get('supplierid')
    bons=Itemsbysupplier.objects.filter(supplier_id=supplierid, ispaid=False)
    trs=''
    for i in bons:
        trs+=f'<tr><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.nbon}</td><td>{i.supplier.name}</td><td>{i.total}</td><td class="text-danger">{i.rest if i.rest>0 else "---"}</td> <td><input type="checkbox" value="{i.id}" name="bonsachattopay" onchange="checkreglementbox(event)"></td></tr>'

    return JsonResponse({
        'trs':trs
    })

def reglebonsachat(request):
    supplierid=request.POST.get('supplierid')
    supplier=Supplier.objects.get(pk=supplierid)
    bons=json.loads(request.POST.get('bons'))
    mantant=json.loads(request.POST.get('mantant'))
    mode=json.loads(request.POST.get('mode'))
    npiece=json.loads(request.POST.get('npiece'))
    date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=json.loads(request.POST.get('echeance'))
    echeance=[datetime.strptime(i, '%Y-%m-%d') if i!='' else None for i in echeance]

    livraisons=Itemsbysupplier.objects.filter(pk__in=bons)
    livraisons.update(ispaid=True)
    totalmantant=sum(mantant)
    # livraisonstotals=0
    # for i in livraisons:
    #     if i.rest>0:
    #         livraisonstotals=round(livraisonstotals+i.rest, 2)
    #     else:
    #         livraisonstotals=round(livraisonstotals+i.total, 2)

    # # update client sold
    # # case1: 5000==5000:
    # if totalmantant==livraisonstotals:
    #     print('case1')
    #     livraisons.update(ispaid=True)
    #     livraisons.update(rest=0)
    # # case2: 5000>4500:
    # if totalmantant>livraisonstotals:
    #     print('case2')
    #     diff=totalmantant-livraisonstotals
    #     livraisons.update(ispaid=True)
    #     livraisons.update(rest=0)
    #     # else:
    #     #     livraisons.update(ispaid=True)
    #     # sub diff from client ref
    # # case3: 5000<6000:
    # if totalmantant<livraisonstotals:
    #     print('case3')
    #     amount=totalmantant
    #     for i in livraisons:
    #         if amount<=0:
    #             break
    #         else:
    #             if i.rest>0:
    #                 if amount>=i.rest:
    #                     i.rest=0
    #                     i.ispaid=True
    #                     i.save()
    #                     amount-=i.rest
    #                 else:

    #                     i.rest=round(i.rest-amount, 2)
    #                     i.save()
    #                     break
    #             else:
    #                 if amount>=i.total:
    #                     amount-=i.total
    #                     i.ispaid=True
    #                     i.save()
    #                 else:
    #                     i.rest=round(i.total-amount, 2)
    #                     i.save()
    #                     break
    #     print('rest of amount', amount)

    for m, mod, np, ech in zip(mantant, mode, npiece, echeance):
        regl=PaymentSupplier.objects.create(
            supplier_id=supplierid,
            amount=m,
            date=date,
            echeance=ech,
            mode=mod,
            npiece=np,
        )
        regl.bons.set(livraisons)
        # for i in livraisons:
        #     Bonsregle.objects.create(
        #         payment=regl,
        #         bon=i,
        #         amount=m
        #     )

    supplier.rest=round(float(supplier.rest)-float(totalmantant), 2)
    supplier.save()
    return JsonResponse({
        "success":True
    })


def journalachat(request):
    items=Stockin.objects.order_by('-id')[:50]
    ctx={
        'title':'Journal Achat',
        'items':items,
        'today':timezone.now().date(),
        'totaljach':round(Stockin.objects.aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalqtyjach':round(Stockin.objects.aggregate(Sum('quantity'))['quantity__sum'] or 0, 2),
    }
    return render(request, 'journalachat.html', ctx)

def laodjournalachat(request):
    page = int(request.GET.get('page', 1))

    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    products = Stockin.objects.all()[start:end]
    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalacha-row">
            <td>{i.date}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page
    })

def journalachatfc(request):
    items=Stockin.objects.filter(facture=True, date__year=thisyear)[:20]
    ctx={
        'title':'Journal Achat Facture',
        'items':items,
        'today':timezone.now().date(),
        'total':round(Stockin.objects.filter(facture=True, date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalqty':round(Stockin.objects.filter(facture=True, date__year=thisyear).aggregate(Sum('quantity'))['quantity__sum'] or 0, 2),
    }
    return render(request, 'journalachatfc.html', ctx)

def loadjournalachatfc(request):
    page = int(request.GET.get('page', 1))

    per_page = 20  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    products = Stockin.objects.filter(facture=True)[start:end]
    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalachafc-row">
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page
    })

def loadjournalachat(request):
    page = int(request.GET.get('page', 1))

    per_page = 100  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    products = Stockin.objects.order_by('-id')[start:end]
    trs=''
    for i in products:
        trs+=f'''
        <tr class="jach-row">
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page
    })


def journalvente(request):
    items=Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).order_by('-date')[:50]
    ctx={
        'title':'Journal vente',
        'items':items,
        'totalqty':Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).aggregate(Sum('qty'))['qty__sum'] or 0,
        'totaltotal':round(Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'today':timezone.now().date()
    }
    print(ctx)
    return render(request, 'journalvente.html', ctx)

def yeardatajournalv(request):
    year=request.GET.get('year')
    print(year)
    items=Livraisonitem.objects.filter(isfacture=False, date__year=year).order_by('-date')[:50]
    trs=''
    totalmarge=0
    for i in items:
        try:
            marge_value = round((i.product.prixnet - (i.product.coutmoyen if i.product.coutmoyen else 0)) * i.qty, 2)
        except:
            marge_value = 0
        totalmarge+=marge_value
        trs+=f'''
        <tr class="journalvente-row" year="{year}">
            <td>{i.date.strftime('%d/%m/%Y') if i.date else ''}</td>
            <td>{i.bon.bon_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjv">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjv">{i.qty}</td>
            <td class="totaljv">{i.total}</td>
            <td>{i.bon.client.name}</td>
            <td>{i.bon.salseman.name}</td>
            <td class="text-success margejv">
                {marge_value}
            </td>
        </tr>
        '''

    return JsonResponse({
        'trs':trs,
        'totalqty':Livraisonitem.objects.filter(isfacture=False, date__year=year).aggregate(Sum('qty'))['qty__sum'] or 0,
        'total':round(Livraisonitem.objects.filter(isfacture=False, date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalmarge':round(totalmarge, 2)
    })

def yeardatajournalvfc(request):
    year=request.GET.get('year')
    print(year)
    items=Outfacture.objects.filter(date__year=year).order_by('-date')[:50]
    trs=''
    totalmarge=0
    for i in items:
        try:
            marge_value = round((i.product.prixnet - (i.product.coutmoyen if i.product.coutmoyen else 0)) * i.qty, 2)
        except:
            marge_value = 0
        totalmarge+=marge_value
        trs+=f'''
        <tr class="journalventefc-row" year="{year}">
            <td>{i.date.strftime('%d/%m/%Y') if i.date else ''}</td>
            <td>{i.facture.facture_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjv">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjv">{i.qty}</td>
            <td class="totaljv">{i.total}</td>
            <td class="totaljv"></td>
            <td>{i.facture.client.name}</td>
            <td>{i.facture.salseman.name}</td>
            <td class="text-success margejv">
                {marge_value}
            </td>
        </tr>
        '''

    return JsonResponse({
        'trs':trs,
        'totalqty':Outfacture.objects.filter(date__year=year).aggregate(Sum('qty'))['qty__sum'] or 0,
        'total':round(Outfacture.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalmarge':round(totalmarge, 2)
    })

def journalventefc(request):
    items=Outfacture.objects.filter(date__year=thisyear).order_by('-date')[:50]
    ctx={
        'title':'Journal vente Facture',
        'items':items,
        'totalqty':Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('qty'))['qty__sum'] or 0,
        'total':round(Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0,2),
        'today':timezone.now().date()
    }
    return render(request, 'journalventefc.html', ctx)

def loadjournalvente(request):
    page = int(request.GET.get('page', 1))
    year = request.GET.get('year')
    term = request.GET.get('term')
    startdate = request.GET.get('startdate')
    enddate = request.GET.get('enddate')
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    if term != '0':

        # Split the term into individual words separated by '*'
        search_terms = term.split('+')
        # Create a list of Q objects for each search term and combine them with &

        q_objects = Q()
        for term in search_terms:
            if term:
                q_objects &= (Q(client__name__iregex=term)|Q(ref__iregex=term)|Q(name__iregex=term)|Q(total__iregex=term)|Q(bon__bon_no__iregex=term))
        if year=='0':
            # means the year is not selected, so the records of the current year
            products = Livraisonitem.objects.filter(isfacture=False).filter(q_objects).order_by('-date')[start:end]
            total=round(Livraisonitem.objects.filter(isfacture=False).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
            totalqty=Livraisonitem.objects.filter(isfacture=False).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
        else:
            products = Livraisonitem.objects.filter(isfacture=False, date__year=year).filter(q_objects).order_by('-date')[start:end]
            total=round(Livraisonitem.objects.filter(isfacture=False, date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
            totalqty=Livraisonitem.objects.filter(isfacture=False, date__year=year).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
        for i in products:
            trs+=f'''
            <tr class="journalvente-row" year={year} term={term}>
                <td>{i.date.strftime('%d/%m/%Y')}</td>
                <td>{i.bon.bon_no}</td>
                <td>{i.product.ref.upper()}</td>
                <td>{i.product.name}</td>
                <td>{i.price}</td>
                <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
                <td style="color:blue" class="">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
                <td class="text-danger coutmoyenjv">{i.product.buyprice if i.product.buyprice else 0}</td>
                <td class="text-danger qtyjv">{i.qty}</td>
                <td class="totaljv">{i.total}</td>
                <td>{i.bon.client.name}</td>
                <td>{i.bon.salseman.name}</td>
                <td class="text-success margejv">

                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(products) == per_page,
        })
    if startdate!='0' and enddate!='0':
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        bons=Livraisonitem.objects.filter(isfacture=False, date__range=[startdate, enddate]).order_by('-date')[:50]
        trs=''
        for i in products:
            trs+=f'''
            <tr class="journalvente-row" startdate={startdate} enddate={enddate}>
                <td>{i.date.strftime('%d/%m/%Y')}</td>
                <td>{i.bon.bon_no}</td>
                <td>{i.product.ref.upper()}</td>
                <td>{i.product.name}</td>
                <td>{i.price}</td>
                <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
                <td style="color:blue" class="">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
                <td class="text-danger coutmoyenjv">{i.product.buyprice if i.product.buyprice else 0}</td>
                <td class="text-danger qtyjv">{i.qty}</td>
                <td class="totaljv">{i.total}</td>
                <td>{i.bon.client.name}</td>
                <td>{i.bon.salseman.name}</td>
                <td class="text-success margejv">

                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(products) == per_page,
        })
    if year=='0':
        # means the year i not selected, so the records of the current year
        products = Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).order_by('-date')[start:end]
    else:
        products = Livraisonitem.objects.filter(isfacture=False, date__year=year).order_by('-date')[start:end]
    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalvente-row" year={year}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.bon.bon_no}</td>
            <td>{i.product.ref.upper()}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger coutmoyenjv">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjv">{i.qty}</td>
            <td class="totaljv">{i.total}</td>
            <td>{i.bon.client.name}</td>
            <td>{i.bon.salseman.name}</td>
            <td class="text-success margejv">

            </td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page
    })

def loadjournalventefc(request):
    page = int(request.GET.get('page', 1))
    term=request.GET.get('term')
    year=request.GET.get('year')
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    if term != '0':
        regex_search_term = term.replace('+', '*')

        # Split the term into individual words separated by '*'
        search_terms = regex_search_term.split('*')
        # Create a list of Q objects for each search term and combine them with &

        q_objects = Q()
        for term in search_terms:
            if term:
                q_objects &= (Q(client__name__iregex=term)|Q(ref__iregex=term)|Q(name__iregex=term)|Q(total__iregex=term)|Q(facture__facture_no__iregex=term))
        if year=='0':
            # means the year i not selected, so the records of the current year
            products = Outfacture.objects.filter(date__year=thisyear).filter(q_objects).order_by('-date')[:50]
            total=round(Outfacture.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
            totalqty=Outfacture.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
        else:
            products = Outfacture.objects.filter(date__year=year).filter(q_objects).order_by('-date')[:50]
            total=round(Outfacture.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
            totalqty=Outfacture.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
        trs=''
        for i in products:
            trs+=f'''
            <tr class="journalventefc-row" year={year} term={term}>
                <td>{i.date.strftime('%d/%m/%Y')}</td>
                <td>{i.facture.facture_no}</td>
                <td>{i.product.ref}</td>
                <td>{i.product.name}</td>
                <td>{i.price}</td>
                <td class="prnetjvfc">{i.product.prixnet if i.product.prixnet else 0}</td>
                <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
                <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
                <td class="text-danger qtyjvfc">{i.qty}</td>
                <td class="totaljvfc">{i.total}</td>
                <td></td>
                <td>{i.facture.client.name}</td>
                <td>{i.facture.salseman.name}</td>
                <td class="text-success margejvfc">

                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'total':total,
            'totalqty':totalqty
        })

    if startdate != '0' and enddate != '0':
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        bons=Outfacture.objects.filter(date__range=[startdate, enddate]).order_by('-date')[:50]
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="journalventefc-row" startdate={startdate} enddate={enddate}>
                <td>{i.date.strftime('%d/%m/%Y')}</td>
                <td>{i.facture.facture_no}</td>
                <td>{i.product.ref}</td>
                <td>{i.product.name}</td>
                <td>{i.price}</td>
                <td class="prnetjvfc">{i.product.prixnet if i.product.prixnet else 0}</td>
                <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
                <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
                <td class="text-danger qtyjvfc">{i.qty}</td>
                <td class="totaljvfc">{i.total}</td>
                <td></td>
                <td>{i.facture.client.name}</td>
                <td>{i.facture.salseman.name}</td>
                <td class="text-success margejvfc">

                </td>
            </tr>
            '''
        ctx={
            'trs':trs,
        }
        if bons:
            ctx['total']=round(Outfacture.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'], 2)
            ctx['totalqty']=Outfacture.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('qty')).get('qty__sum')
        return JsonResponse(ctx)
    if year=='0':
        # means the year i not selected, so the records of the current year
        products = Outfacture.objects.filter(date__year=thisyear).order_by('-date')[:50]
        total=round(Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('qty'))['qty__sum'] or 0
    else:
        products = Outfacture.objects.filter(date__year=year).order_by('-date')[:50]
        total=round(Outfacture.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Outfacture.objects.filter(date__year=year).aggregate(Sum('qty'))['qty__sum'] or 0

    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalventefc-row" year={year}>
            <td>{i.date.strftime("%d/%m/%Y")}</td>
            <td>{i.facture.facture_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjvfc">{i.product.prixnet}</td>
            <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen}</td>
            <td class="text-danger">{i.product.buyprice}</td>
            <td class="text-danger qtyjvfc">{i.qty}</td>
            <td>{i.total}</td>
            <td></td>
            <td>{i.client.name}</td>
            <td>{i.facture.salseman.name}</td>
            <td class="text-success margejvfc">

            </td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page,
        'total':(Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalqty':Outfacture.objects.filter(date__year=thisyear).aggregate(Sum('qty'))['qty__sum'] or 0,
    })

# product search selects2 for bons
def searchproduct(request):
    # get url pams
    print('rrrr')
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (
                Q(ref__icontains=term) | 
                Q(name__icontains=term) | 
                Q(mark__name__icontains=term) | 
                Q(category__name__icontains=term) |
                Q(equivalent__icontains=term) | 
                Q(refeq1__icontains=term) | 
                Q(refeq2__icontains=term) | 
                Q(refeq3__icontains=term) | 
                Q(refeq4__icontains=term)
            )
    # check if term in product.ref or product.name
    products=Produit.objects.filter(q_objects).order_by('-stocktotal')

    results=[]
    for i in products:
        results.append({
            'id':f'{i.ref}§{i.name}§{i.buyprice}§{i.stocktotal}§{i.stockfacture}§{i.id}§{i.sellprice}§{i.remise}§{i.prixnet}§{i.representprice}',
            'text':f'{i.ref.upper()} - {i.name.upper()} ({"(OUI)" if i.stocktotal > 0 else "(NON)"})',
            'stock':i.stocktotal,
            'stockfacture':i.stockfacture
        })
    return JsonResponse({'results': results})

def filterbldate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    print(startdate, enddate)
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Bonlivraison.objects.filter(date__range=[startdate, enddate]).order_by('-bon_no')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="ord {"text-danger" if i.ispaid else ''} bl-row" startdate={startdate} enddate={enddate} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')">
            <td>{ i.bon_no }</td>
            <td>{ i.date.strftime("%d/%m/%Y")}</td>
            <td>{ i.client.name }</td>
            <td>{ i.client.code }</td>
            <td>{ i.total}</td>
            <td>{ i.client.region}</td>
            <td>{ i.client.city}</td>
            <td>{ i.client.soldbl}</td>
            <td>{ i.salseman }</td>
            <td class="d-flex justify-content-between">
              <div>
              {'R0' if i.ispaid else 'N1' }

              </div>
              <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

            </td>
            <td class="text-danger">
            {'OUI' if i.isfacture else 'NON'}

            </td>

            <td>
              {i.commande.order_no if i.commande else '--'}
            </td>
            <td>
              {i.modlvrsn}
            </td>
          </tr>
        '''
    ctx={
        'trs':trs
    }
    if bons:
        ctx['total']=round(Bonlivraison.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total')).get('total__sum'), 2)
    return JsonResponse(ctx)

def searchclientrep(request):
    term=request.GET.get('term')
    rep=request.user.represent
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(name__icontains=term) | 
                Q(code__icontains=term) | 
                Q(region__icontains=term) | 
                Q(city__icontains=term) | 
                Q(address__icontains=term))
    clients=Client.objects.filter(represent=rep).filter(q_objects)
    # if '+' in term:
    #     term=term.split('+')
    #     for i in term:
    #         clients=Client.objects.filter(
    #             Q(name__icontains=i) |
    #             Q(code__icontains=i) |
    #             Q(region__icontains=i) |
    #             Q(city__icontains=i)
    #         )
    # else:
    #     clients=Client.objects.filter(
    #         Q(name__icontains=term) |
    #         Q(code__icontains=term) |
    #         Q(region__icontains=term) |
    #         Q(city__icontains=term)
    #     )
    results=[]
    for i in clients:
        results.append({
            'id':i.id,
            'text':f'{i.name} - {i.city}'
        })
    return JsonResponse({'results': results})

def searchsupplier(request):
    term=request.GET.get('term')
    print(term)
    suppliers=Supplier.objects.filter(Q(name__icontains=term) | Q(phone__icontains=term)| Q(address__icontains=term))
    print('suppliers', suppliers)
    results=[]
    for i in suppliers:
        results.append({
            'id':i.id,
            'text':i.name
        })
    return JsonResponse({'results': results})

def getclientfactureprice(request):
    id=request.POST.get('id')
    clientid=request.POST.get('clientid')
    print(id, 'rr', clientid)
    try:
        clientprice=Outfacture.objects.filter(client_id=clientid, product_id=id).last()
        price=clientprice.price
        remise=clientprice.remise
        return JsonResponse({
            'price':price,
            'remise':remise
        })
    except:
        return JsonResponse({
            'price':0,
            'remise':0
        })


def updatereglebons(request):
    reglementid=request.POST.get('reglementid')
    mantant=request.POST.get('mantant')
    mode=request.POST.get('mode')
    npiece=request.POST.get('npiece')
    date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=request.POST.get('echeance')
    echeance=datetime.strptime(echeance, '%Y-%m-%d') if echeance!='' else None
    newbons=json.loads(request.POST.get('bons'))
    reglement=PaymentClientbl.objects.get(pk=reglementid)

    thisclient=reglement.client
    print('soldbl', float(thisclient.soldbl),float(reglement.amount),float(mantant))
    thisclient.soldbl=round(float(thisclient.soldbl)+float(reglement.amount)-float(mantant), 2)
    thisclient.soldtotal=round(float(thisclient.soldtotal)+float(reglement.amount)-float(mantant), 2)
    thisclient.save()
    oldbons=reglement.bons.all()
    livraisons=Bonlivraison.objects.filter(pk__in=newbons)
    oldbons.update(ispaid=False)
    oldbons.update(statusreg='n1')
    print(oldbons)
    print(livraisons)
    livraisons.update(ispaid=True)
    livraisons.update(statusreg='r0')
    reglement.bons.clear()

    # # update regleemnt amount
    reglement.date=date
    reglement.amount=mantant
    reglement.mode=mode
    reglement.npiece=npiece
    reglement.echance=echeance

    reglement.bons.set(livraisons)
    reglement.save()
    return JsonResponse({
        'success':True
    })
    # thisclient.soldbl=round(float(thisclient.soldbl)-float(reglement.amount)+float(mantant), 2)
    # substract the old total from client soldbl


def updatereglesupp(request):
    reglementid=request.POST.get('reglementid')
    mantant=request.POST.get('mantant')
    mode=request.POST.get('mode')
    npiece=request.POST.get('npiece')
    date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=request.POST.get('echeance')
    echeance=datetime.strptime(echeance, '%Y-%m-%d') if echeance!='' else None
    newbons=json.loads(request.POST.get('bons'))
    reglement=PaymentSupplier.objects.get(pk=reglementid)

    thissupplier=reglement.supplier
    thissupplier.rest=round(float(thissupplier.rest)+float(reglement.amount)-float(mantant), 2)
    thissupplier.save()
    oldbons=reglement.bons.all()
    livraisons=Itemsbysupplier.objects.filter(pk__in=newbons)
    oldbons.update(ispaid=False)
    print(oldbons)
    print(livraisons)
    livraisons.update(ispaid=True)
    reglement.bons.clear()

    # # update regleemnt amount
    reglement.date=date
    reglement.amount=mantant
    reglement.mode=mode
    reglement.npiece=npiece
    reglement.echance=echeance

    reglement.bons.set(livraisons)
    reglement.save()
    return JsonResponse({
        'success':True
    })
    # thisclient.soldbl=round(float(thisclient.soldbl)-float(reglement.amount)+float(mantant), 2)
    # substract the old total from client soldbl

def getreglementbl(request, id):
    reglement=PaymentClientbl.objects.get(pk=id)

    bons=reglement.bons.all()
    # bons without bons in reglement
    livraisons=Bonlivraison.objects.filter(client=reglement.client).exclude(pk__in=[bon.pk for bon in bons]).order_by('-id')[:50]
    # livraisons=Bonlivraison.objects.filter(client=reglement.client)
    #we need bons to calculate total bl
    bonstocalculate=Bonlivraison.objects.filter(client=reglement.client)
    trs=''
    for i in livraisons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="loadblinupdateregl" reglemntid="{id}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'date':reglement.date.strftime('%Y-%m-%d'),
        'echance':reglement.echance.strftime('%Y-%m-%d') if reglement.echance else '',
        'mode':reglement.mode,
        'npiece':reglement.npiece,
        'bons':list(bons.values()),
        'client':reglement.client.name,
        'clientid':reglement.client.id,
        'mantant':reglement.amount,
        'livraisons':trs,
        'total':round(bonstocalculate.aggregate(Sum('total'))['total__sum'] or 0, 2),
        'soldclient':reglement.client.soldbl
    })



def getreglementfc(request, id):
    reglement=PaymentClientfc.objects.get(pk=id)
    # facture of this reglement
    bons=reglement.factures.all()

    livraisons=Facture.objects.filter(client=reglement.client).exclude(pk__in=[bon.pk for bon in bons]).order_by('-id')[:50]
    bonstocalculate=Facture.objects.filter(client=reglement.client)
    trs=''
    for i in livraisons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="loadblinupdatereglfc" reglemntid="{id}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'date':reglement.date.strftime('%Y-%m-%d'),
        'echance':reglement.echance.strftime('%Y-%m-%d') if reglement.echance else '',
        'mode':reglement.mode,
        'npiece':reglement.npiece,
        'bons':list(bons.values()),
        'client':reglement.client.name,
        'clientid':reglement.client.id,
        'mantant':reglement.amount,
        'livraisons':trs,
        # 'livraisons':list(livraisons.values()),
        'soldclientfc':reglement.client.soldfacture,
        'total':round(bonstocalculate.aggregate(Sum('total'))['total__sum'] or 0, 2),


    })

def updatereglefactures(request):
    reglementid=request.POST.get('reglementid')
    mantant=request.POST.get('mantant')
    mode=request.POST.get('mode')
    npiece=request.POST.get('npiece')
    date=datetime.strptime(request.POST.get('date'), '%Y-%m-%d')
    echeance=request.POST.get('echeance')
    echeance=datetime.strptime(echeance, '%Y-%m-%d') if echeance!='' else None
    newbons=json.loads(request.POST.get('bons'))
    reglement=PaymentClientfc.objects.get(pk=reglementid)
    thisclient=reglement.client
    print('soldbl', float(thisclient.soldfacture),float(reglement.amount),float(mantant))
    thisclient.soldfacture=round(float(thisclient.soldfacture)+float(reglement.amount)-float(mantant), 2)
    thisclient.soldtotal=round(float(thisclient.soldtotal)+float(reglement.amount)-float(mantant), 2)
    thisclient.save()
    oldbons=reglement.factures.all()
    oldbons.update(ispaid=False)
    oldbons.update(statusreg='b1')
    reglement.factures.clear()
    livraisons=Facture.objects.filter(pk__in=newbons)
    livraisons.update(ispaid=True)
    livraisons.update(statusreg='f1')



    # # update regleemnt amount
    reglement.date=date
    reglement.amount=mantant
    reglement.mode=mode
    reglement.npiece=npiece
    reglement.echance=echeance
    reglement.factures.set(livraisons)
    # reglement.bons.set(livraisons)
    reglement.save()
    return JsonResponse({
        'success':True
    })
    # thisclient.soldbl=round(float(thisclient.soldbl)-float(reglement.amount)+float(mantant), 2)
    # substract the old total from client soldbl

def getlastsuppprice(request):
    id=request.POST.get('id')
    supplierid=request.POST.get('supplierid')
    print(id, 'rr', supplierid)
    try:
        lastprice=Stockin.objects.filter(product_id=id, supplier_id=supplierid).last()
        print(lastprice)
        price=lastprice.price
        remise=lastprice.remise

        return JsonResponse({
            'price':price,
            'remise':remise,
            'facture':lastprice.facture
        })
    except:
        return JsonResponse({
            'price':0
        })


def boncommandedetails(request, id):

    order=Order.objects.get(pk=id)
    orderitems=Orderitem.objects.filter(order=order)
    orderitems=list(orderitems)
    orderitems=[orderitems[i:i+36] for i in range(0, len(orderitems), 36)]
    ctx={
        'title':f'Bon de Commande {order.order_no}',
        'order':order,
        'orderitems':orderitems,
    }
    return render(request, 'boncommandedetails.html', ctx)


def genererbonlivraison(request, id):
    order=Order.objects.get(pk=id)
    items=Orderitem.objects.filter(order=order)
    # we need date of last invoice

    #year = timezone.now().strftime("%y")
    #latest_receipt = Bonlivraison.objects.filter(
    #    bon_no__startswith=f'BL{year}'
    #).order_by("-id").first()
    #if latest_receipt:
    #    latest_receipt_no = int(latest_receipt.bon_no[-5:])
    #    receipt_no = f"BL{year}{latest_receipt_no + 1:05}"
    #else:
    #    receipt_no = f"BL{year}00001"
    ctx={
        'order':order,
        'items':items,
        #'receipt_no':receipt_no,
        #'clients':Client.objects.all(),
        'reps':Represent.objects.all(),
    }
    return render(request, 'genererbonlivraison.html', ctx)


def createclientaccount(request):
    print('>>>>>>')
    clientcode=request.GET.get('clientcode')
    client= Client.objects.get(code=clientcode)
    olduser=client.user
    username=request.GET.get('username')
    password=request.GET.get('password')
    #check for username
    print('>>>>>>', clientcode, username, password)
    user=User.objects.filter(username=username).first()
    if user:
        print('username already', username)
        return JsonResponse({
            'success':False,
            'error':'Username exist déja'
        })
    # create user
    user=User.objects.create_user(username=username, password=password)
    # assign user to client
    group, created = Group.objects.get_or_create(name="clients")
    user.groups.add(group)
    user.save()
    client.user=user
    client.save()
    if olduser:
        cart=Cart.objects.filter(user=olduser).first()
        wich=Wich.objects.filter(user=olduser).first()
        if cart:
            cart.user=user
            cart.save()
        if wich:
            wich.user=user
            wich.save()
        olduser.delete()
    return JsonResponse({
        'success':True
    })


def createrepaccount(request):
    repid=request.GET.get('repid')
    rep= Represent.objects.get(pk=repid)
    username=request.GET.get('username')
    password=request.GET.get('password')
    #check for username
    user=User.objects.filter(username=username).first()
    if user:
        return JsonResponse({
            'success':False,
            'error':'Username exist déja'
        })

    # create user
    user=User.objects.create_user(username=username, password=password)
    # assign user to rep
    olduser=rep.user
    group, created = Group.objects.get_or_create(name="salsemen")
    user.groups.add(group)
    user.save()
    rep.user=user
    rep.save()
    if olduser:
        cart=Cart.objects.filter(user=olduser).first()
        wich=Wich.objects.filter(user=olduser).first()
        if cart:
            cart.user=user
            cart.save()
        if wich:
            wich.user=user
            wich.save()
        olduser.delete()
    return JsonResponse({
        'success':True
    })

def carlogospage(request):
    ctx={
        # maintain same names
        'categories':Carlogos.objects.all(),
        'title':'Voitures logo'
    }
    return render(request, 'carlogos.html', ctx)

def createlogo(request):
    name=request.POST.get('name')
    # get image file
    image=request.FILES.get('image')
    # create category
    Carlogos.objects.create(name=name, image=image)
    return JsonResponse({
        'success':True
    })

def updatelogo(request):
    image=request.FILES.get('updatelogoimage') or None
    id=request.POST.get('id')
    carlogo=Carlogos.objects.get(pk=id)
    carlogo.name=request.POST.get('updatelogoname')
    if image:
        carlogo.image=image
    carlogo.save()
    ctx={
        'categories':Carlogos.objects.all(),
        'title':'Voiture logo'
    }
    return JsonResponse({
        'html':render(request, 'carlogos.html', ctx).content.decode('utf-8')
    })


def getnotpaid(request):
    # get bon livraison not paid more than 3 months

    three_months_ago = timezone.now() - timedelta(days=90)  # Assuming 30 days per month on average

    # Query for Bonlivraison objects that have a 'date' field earlier than three months ago
    bons = Bonlivraison.objects.filter(date__lt=three_months_ago, ispaid=False)


    return JsonResponse({
        'html':render(request, 'bllist.html', {'bons':bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total')).get('total__sum'), 2)
    })

def getnotpaidfc(request):
    # get bon livraison not paid more than 3 months

    three_months_ago = timezone.now() - timedelta(days=90)  # Assuming 30 days per month on average

    # Query for Bonlivraison objects that have a 'date' field earlier than three months ago
    bons = Facture.objects.filter(date__lt=three_months_ago, ispaid=False)


    return JsonResponse({
        'html':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total')).get('total__sum') or 0, 2)
    })


def filterfcdate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Facture.objects.filter(date__range=[startdate, enddate]).order_by('-date')[:50]
    print('total', Facture.objects.filter(date__range=[startdate, enddate]).count(), Facture.objects.count())
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" startdate={startdate} enddate={enddate} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
            <td>{ i.facture_no }</td>
            <td>{ i.date.strftime("%d/%m/%Y")}</td>
            <td>{ i.total}</td>
            <td>{ i.tva}</td>
            <td>{ i.client.name }</td>
            <td>{ i.client.code }</td>
            <td>{ i.client.region}</td>
            <td>{ i.client.city}</td>
            <td>{ i.client.soldfacture}</td>
            <td>{ i.salseman }</td>
            <td class="d-flex justify-content-between">
              <div>
              {'R0' if i.ispaid else 'N1' }

              </div>
              <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

            </td>
            <td class="text-danger">

            </td>

            <td>
              {i.bon.bon_no if i.bon else "--"}
            </td>
          </tr>
        '''
    ctx={
        'trs':trs
    }
    if bons:
        ctx['total']=round(bons.aggregate(Sum('total')).get('total__sum'), 2)
        ctx['totaltva']=round(bons.aggregate(Sum('tva')).get('tva__sum'), 2)
    return JsonResponse(ctx)
    # return JsonResponse({
    #     'html':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8'),
    #     'total':round(bons.aggregate(Sum('total')).get('total__sum'), 2),
    #     'totaltva':round(bons.aggregate(Sum('tva')).get('tva__sum'), 2),

    # })

def filterachatdate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Itemsbysupplier.objects.filter(date__range=[startdate, enddate])
    return JsonResponse({
        'html':render(request, 'achatlist.html', {'bons':bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('total')).get('total__sum'), 2)
    })

def updateclientpassword(request):
    clientid=request.POST.get('clientid')
    password=request.POST.get('password')
    try:
        user=Client.objects.get(pk=clientid).user
        user.set_password(password)
        user.save()
        return JsonResponse({
            'success':True
        })
    except Exception as e:
        return JsonResponse({
            'success':False,
            'error':e
        })

def updatereppassword(request):
    repid=request.POST.get('repid')
    password=request.POST.get('password')
    try:
        user=Represent.objects.get(pk=repid).user
        print(user, 'user', password)
        user.set_password(password)
        user.save()
        # logout(user)

        return JsonResponse({
            'success':True
        })
    except Exception as e:
        return JsonResponse({
            'success':False,
            'error':e
        })



def filterreglbldate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=PaymentClientbl.objects.filter(date__range=[startdate, enddate])
    print(bons)
    return JsonResponse({
        'html':render(request, 'reglbllist.html', {'bons':bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('amount')).get('amount__sum'), 2),

    })


def filterreglfcdate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=PaymentClientfc.objects.filter(date__range=[startdate, enddate])
    return JsonResponse({
        'html':render(request, 'reglfclist.html', {'bons':bons}).content.decode('utf-8'),
        'total':round(bons.aggregate(Sum('amount')).get('amount__sum'), 2),

    })



def sortupbl(request):
    bons=Bonlivraison.objects.all().order_by('date')
    return JsonResponse({
        'html':render(request, 'bllist.html', {'bons':bons}).content.decode('utf-8')
    })

def sortdownbl(request):
    bons=Bonlivraison.objects.all().order_by('-date')
    return JsonResponse({
        'html':render(request, 'bllist.html', {'bons':bons}).content.decode('utf-8')
    })

def sortupfc(request):
    bons=Facture.objects.all().order_by('date')
    return JsonResponse({
        'html':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8')
    })

def sortdownfc(request):
    bons=Facture.objects.all().order_by('-date')
    return JsonResponse({
        'html':render(request, 'fclist.html', {'bons':bons}).content.decode('utf-8')
    })

def excelclients(request):
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    df = df.fillna('')
    for d in df.itertuples():
        name=d.name
        code=d.code
        try:
            # this needs to be lowercase
            region=d.region.lower().strip()
        except:
            region=d.region
        city=d.city
        phone=str(d.phone)
        rep=d.rep
        ice=str(d.ice)
        address=None if pd.isna(d.address) else d.address
        try:
            client=Client.objects.get(Q(name=name) | Q(code=code))
            with open('error.txt', 'a') as f:
                f.write(f'{name} - {code} exist déja \n')
        except Client.DoesNotExist:
            print('client not exist')
            client=Client.objects.create(
                represent_id=rep,
                code=code,
                name=name,
                city=city,
                ice=ice,
                region=region,
                phone=phone,
                address=address
            )
    return JsonResponse({
        'success':True
    })

def excelpdcts(request):
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    entries=0
    for d in df.itertuples():
        try:
            ref = d.ref.lower().strip()
        except:
            ref=d.ref
        #reps=json.dumps(d.rep)
        #name = d.name
        #mark = None if pd.isna(d.mark) else d.mark
        #refeq = '' if pd.isna(d.refeq) else d.refeq
        #status = False if pd.isna(d.status) else True
        #coderef = '' if pd.isna(d.code) else d.code
        #diam = '' if pd.isna(d.diam) else d.diam
        #qty = 0 if pd.isna(d.qty) else d.qty
        #buyprice = 0 if pd.isna(d.buyprice) else d.buyprice
        #devise = 0 if pd.isna(d.devise) else d.devise
        # car = None if pd.isna(d.car) else d.car
        #prixbrut = 0 if pd.isna(d.prixbrut) else d.prixbrut
        #ctg = None if pd.isna(d.ctg) else d.ctg
        #order = '' if pd.isna(d.order) else d.order
        #img = None if pd.isna(d.img) else d.img
        remise = 0 if pd.isna(d.remise) else d.remise
        #prixnet=0 if pd.isna(d.prixnet) else d.prixnet
        try:
            print('entering', ref)
            product=Produit.objects.get(ref=ref)
            product.representremise=int(remise)
            product.save()
            entries+=1

        except Exception as e:
            print('>>',e, ref)
            with open('error.txt', 'a') as ff:
                ff.write(f'>> {e} {ref}')
            # product=Produit.objects.create(
            #     ref=ref,
            #     equivalent=refeq,
            #     isactive=False,
            #     coderef=coderef,
            #     name=name,
            #     mark_id=mark,
            #     sellprice=prixbrut,
            #     prixnet=prixnet,
            #     remise=remise,
            #     category_id=ctg,
            #     image=img,
            #     stockfacture=qty,
            #     #diametre=diam,
            #     buyprice=buyprice,
            #     devise=devise
            # )
    
    print('>>>', entries)
    return JsonResponse({
        'success':True
    })


def deactivateaccount(request):
    userid=request.GET.get('userid')
    user=User.objects.get(pk=userid)
    user.is_active=False
    user.save()
    # delete user session in django session
    #UserSession.objects.filter(user=user).delete()
    # Clear the user's session
    #Session.objects.filter(session_key__in=UserSession.objects.filter(user=user).values('session_key')).delete()

    return JsonResponse({
        'success':True
    })

def activateaccount(request):
    userid=request.GET.get('userid')
    user=User.objects.get(pk=userid)
    user.is_active=True
    user.save()
    return JsonResponse({
        'success':True
    })

def stock(request):
    categories=Category.objects.all()
    products=Produit.objects.all()[:50]
    ctx={'categories':categories,
        'title':'Liste des Articles',
        'products':products,
        'stocktotal':Produit.objects.all().aggregate(Sum('stocktotal'))['stocktotal__sum']or 0,
        'stockfacture':Produit.objects.all().aggregate(Sum('stockfacture'))['stockfacture__sum']or 0,


    }
    return render(request, 'admin/products.html', ctx)


def getreglementsupp(request, id):
    reglement=PaymentSupplier.objects.get(pk=id)
    ctx={
        'title':'Reglement',
        'reglement':reglement,
    }
    bons=reglement.bons.all()
    # bons without bons in reglement
    livraisons=Itemsbysupplier.objects.filter(supplier=reglement.supplier).exclude(pk__in=[bon.pk for bon in bons])
    # livraisons=Itemsbysupplier.objects.filter(supplier=reglement.supplier)
    #we need bons to calculate total bl
    bonstocalculate=Itemsbysupplier.objects.filter(supplier=reglement.supplier)
    trs=''
    for i in livraisons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementssupp.exists() else ""}" class="loadblinupdatereglsupp" reglemntid="{id}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.nbon}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementssupp.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'date':reglement.date.strftime('%Y-%m-%d'),
        'echance':reglement.echeance.strftime('%Y-%m-%d') if reglement.echeance else '',
        'mode':reglement.mode,
        'npiece':reglement.npiece,
        'bons':list(bons.values()),
        'supplier':reglement.supplier.name,
        'mantant':reglement.amount,
        'livraisons':list(livraisons.values()),
        'total':round(bonstocalculate.aggregate(Sum('total'))['total__sum'], 2),
        'soldsupplier':reglement.supplier.rest
    })

def updatebonavoir(request):
    id=request.POST.get('bonid')
    avoir=Avoirclient.objects.get(pk=id)
    client=Client.objects.get(pk=request.POST.get('clientid'))
    # we need avoir n° cause delete avoir will delete id, id is used in avoir n°
    avoirno=avoir.no
    avoiritems=Returned.objects.filter(avoir=avoir)
    totalbon=request.POST.get('totalbon')
    newmode=request.POST.get('mode')
    isfacture=True if newmode=='facture' else False
    print("isfacture", isfacture)
    thisclient=avoir.client
    # regle stock
    # # regle soldclient
    # if avoir.avoirfacture:
    #     thisclient.soldtotal=round(float(thisclient.soldtotal)+float(avoir.total), 2)
    #     thisclient.soldfacture=round(float(thisclient.soldfacture)+float(avoir.total), 2)
    #     thisclient.save()
    #     for i in avoiritems:
    #         i.product.stocktotal+=i.qty
    #         i.product.stockfacture+=i.qty
    #         i.product.save()
    #         i.delete()
    # else:
    #     thisclient.soldtotal=round(float(thisclient.soldtotal)+float(avoir.total), 2)
    #     thisclient.soldbl=round(float(thisclient.soldbl)+float(avoir.total), 2)
    #     thisclient.save()
    #     for i in avoiritems:
    #         i.product.stocktotal+=i.qty
    #         i.product.save()
    #         i.delete()

    # # delete old avoir
    # # create new avoir
    if avoir.client==client:
        thisclient.soldtotal=round(float(thisclient.soldtotal)+float(avoir.total)-float(totalbon), 2)
        if avoir.avoirfacture:
            thisclient.soldfacture=round(float(thisclient.soldfacture)+float(avoir.total), 2)
        else:
            thisclient.soldbl=round(float(thisclient.soldbl)+float(avoir.total), 2)
        if isfacture:
            thisclient.soldfacture=round(float(thisclient.soldfacture)-float(totalbon), 2)
        else:
            thisclient.soldbl=round(float(thisclient.soldbl)-float(totalbon), 2)

        thisclient.save()
    else:
        # not the same client
        thisclient.soldtotal=round(float(thisclient.soldtotal)+float(avoir.total), 2)
        # add sold to old client
        if avoir.avoirfacture:
            thisclient.soldfacture=round(float(thisclient.soldfacture)+float(avoir.total), 2)
        else:
            thisclient.soldbl=round(float(thisclient.soldbl)+float(avoir.total), 2)
        thisclient.save()
        # add sold to new client
        client.soldtotal=round(float(client.soldtotal)- float(totalbon), 2)
        if isfacture:
            client.soldfacture=round(float(client.soldfacture)- float(totalbon), 2)
        else:
            client.soldbl=round(float(client.soldbl)- float(totalbon), 2)
        client.save()
        print('new', client.soldtotal)
    items=Returned.objects.filter(avoir=avoir)
    for i in items:
        product=Produit.objects.get(pk=i.product_id)
        product.stocktotal=int(product.stocktotal)-int(i.qty)
        if avoir.avoirfacture:
            product.stockfacture=int(product.stockfacture)-int(i.qty)
        product.save()
        i.delete()
    avoir.client=client
    avoir.representant_id=request.POST.get('repid')
    avoir.total=totalbon
    datebon=request.POST.get('datebon')
    datebon=datetime.strptime(datebon, '%Y-%m-%d')
    avoir.date=datebon
    avoir.no=request.POST.get('orderno')
    if isfacture:
        avoir.avoirfacture=True
        client.soldbl=5
    else:
        avoir.avoirfacture=False
        client.soldbl=5
    avoir.save()
    # update this items


    print('client:', avoir.client.id)
    with transaction.atomic():
        for i in json.loads(request.POST.get('products')):
            product=Produit.objects.get(pk=i['productid'])
            product.stocktotal=int(product.stocktotal)+int(i['qty'])
            if isfacture:
                product.stockfacture=int(product.stockfacture)+int(i['qty'])
            product.save()
            Returned.objects.create(
                avoir=avoir,
                product=product,
                qty=i['qty'],
                remise=i['remise'],
                price=i['price'],
                total=i['total'],
            )

    return JsonResponse({
        'success':True
    })

def notifyadmin(request):
    oldnotif=Ordersnotif.objects.filter(isread=True)
    oldnotif.delete()
    newnotif=Ordersnotif.objects.filter(isread=False)
    response= JsonResponse({
        "length": newnotif.count(),
    })
    response['Access-Control-Allow-Origin'] = 'http://localserver'
    return response

def disablenotif(request):
    newnotif=Ordersnotif.objects.filter(isread=False)
    newnotif.update(isread=True)
    return JsonResponse({
        'success':True
    })

def listecheance(request):
    # get payments that are cheque or effet in mode
    reglbl=PaymentClientbl.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False, echance__lte=today)
    reglfc=PaymentClientfc.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False, echance__lte=today)
    print(reglbl)
    print(reglfc)
    echeance = chain(*[
    ((rbl, 'bl') for rbl in reglbl),
    ((rfc, 'fc') for rfc in reglfc),
    ])
    totalbl=reglbl.aggregate(Sum('amount'))['amount__sum'] or 0
    totalfc=reglfc.aggregate(Sum('amount'))['amount__sum'] or 0
    total=round(totalbl+totalfc, 2)
    # Sort the items by date
    sorted_echeance = sorted(echeance, key=lambda item: item[0].date)
    print(sorted_echeance)
    ctx={
        'title':'List des echeances Actuel',
        'echeances':sorted_echeance,
        'total':total
    }
    return render(request, 'listecheance.html', ctx)


def echeancetoday(request):
    reglbl=PaymentClientbl.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False,echance=today).count()
    reglfc=PaymentClientfc.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False,echance=today).count()
    print('>>>>>',PaymentClientbl.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False,echance=today), PaymentClientfc.objects.filter(Q(mode="cheque") | Q(mode="effet"), ispaid=False,echance=today), today)
    return JsonResponse({
        'length':reglbl+reglfc
    })
def tabs(request):
    return render(request, 'tabs.html')


def getconnectedusers(request):
    # more than 5 minutes means the user is not connected
    five_minutes_ago = timezone.now() - timedelta(minutes=10)
    notconnected=Connectedusers.objects.filter(lasttime__lt=five_minutes_ago)
    connected=Connectedusers.objects.filter(lasttime__gt=five_minutes_ago)
    length=connected.count()
    trs=''
    for i in connected:
        if i.user.groups.all().first().name=='clients':
            trs+=f"""
            <tr>
            <td>{i.user.client.name}</td>
            <td>client</td>
            <td>{i.activity}</td>
            <td>{(i.lasttime).strftime('%Y-%m-%d %H:%M:%S')}</td>
            </tr>
            """
        else:
             trs+=f"""
            <tr>
            <td>{i.user.represent.name}</td>
            <td>commercial({i.user.represent.client.name if i.user.represent.client else ''})</td>
            <td>{i.activity}</td>
            <td>{(i.lasttime).strftime('%Y-%m-%d %H:%M:%S')}</td>
            </tr>
            """
    print('connected', connected)
    print('notconnected', notconnected)
    response= JsonResponse({
        'length':length,
        'trs':trs
    })
    response['Access-Control-Allow-Origin'] = 'http://localserver'
    return response


def payreglbl(request):
    reglid=request.GET.get('reglid')
    regl=PaymentClientbl.objects.get(pk=reglid)
    regl.ispaid=True
    regl.save()
    return JsonResponse({
        'success':True
    })

def payreglfc(request):
    reglid=request.GET.get('reglid')
    regl=PaymentClientfc.objects.get(pk=reglid)
    print(reglid, regl)
    regl.ispaid=True
    regl.save()
    return JsonResponse({
        'success':True
    })


def boncommandes(request):
    length=Order.objects.filter(isdelivered=False).count()
    return JsonResponse({
        'length':length
    })


def listeconnected(request):
    five_minutes_ago = timezone.now() - timedelta(minutes=10)
    
    notconnected=Connectedusers.objects.filter(lasttime__lt=five_minutes_ago).order_by('-lasttime')
    connected=Connectedusers.objects.filter(lasttime__gt=five_minutes_ago)
    listconnected=[]
    listactive=[]
    for i in connected:
        if i.user.groups.all().first().name=='clients':
            listconnected.append([i.user.client.name, i.activity, i.lasttime])
        else:
            listconnected.append([i.user.username, i.activity, i.lasttime])
    for i in notconnected:
        if i.user.groups.all().first().name=='clients':
            listactive.append([i.user.client.name, i.activity, i.lasttime])
        else:
            listactive.append([i.user.username, i.activity, i.lasttime])
    ctx={
        'title':'List utilisateurs Active',
        'connected':listconnected,
        'active':listactive
    }

    response= JsonResponse(ctx)
    response['Access-Control-Allow-Origin'] = 'http://localserver'
    return response

def updatepdctdata(requests):
    id=requests.GET.get('id')
    ref=requests.GET.get('ref')
    stocktotal=requests.GET.get('stocktotal')
    #cars=requests.GET.get('cars')
    try:
        pdct=Produit.objects.get(pk=id)
        pdct.stocktotal=stocktotal
        #pdct.cars=cars
        pdct.save()
    except:
        print('not working')
    return JsonResponse({
        'success':True
    })

def promotionspage(request):
    ctx={
        'promotions':Promotion.objects.all(),
        'title':'List des promotions'
    }
    return render(request, 'promotions.html', ctx)

def createpromotion(request):
    name=request.GET.get('name')
    # get image file
    image=request.GET.get('image')
    # create category
    Promotion.objects.create(info=name, image=image)
    
    return JsonResponse({
        'success':True
    })

def updatepromotion(request):
    image=request.FILES.get('image') or None
    id=request.POST.get('id')
    promotion=Promotion.objects.get(pk=id)
    promotion.info=request.POST.get('name')
    if image:
        promotion.image=image
    promotion.save()
    return JsonResponse({
        'success':True
    })


def searchproductsforstock(request):
    term=request.GET.get('term')
    if(term==''):
        products=Produit.objects.all()[:50]
        trs=''
        for i in products:
            trs+=f'''
        <tr ondblclick="createtab('pdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
            style="background:{'#f3d6d694;' if not i.isactive else '' }"
                data-product-id="{ i.id }" class="product-row">
                  <td style="padding: 5px; font-weight: bold;" >
                      {i.ref.upper()}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.name}
                  </td>

                  <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
                      {i.buyprice if i.buyprice else 0}
                  </td>
                  <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
                      {i.sellprice}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center">
                      {i.remise}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center">
                      {i.prixnet}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
                      {i.stocktotal}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
                    <span class="stockfacture invisible">{i.stockfacture}</span>
                </td>

                  <td style="padding: 5px; font-weight: bold;">
                    {i.diametre}
                </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-success">
                    {i.block}
                </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[0] if i.getequivalent() else ''}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[1] if i.getequivalent() else ''}

                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[2] if i.getequivalent() else ''}

                  </td>

                    <td style="padding: 5px; font-weight: bold;">
                        {i.mark}
                    </td>
                    <td style="padding: 5px; font-weight: bold;">
                        {i.code}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-danger"><span class="percentage invisible"> {round(i.getpercentage(), 2)}</span></td>

              </tr>
        '''
        return JsonResponse({
            'trs':trs,
            'stocktotal':Produit.objects.all().aggregate(Sum('stocktotal'))['stocktotal__sum']or 0,
            'stockfacture':Produit.objects.all().aggregate(Sum('stockfacture'))['stockfacture__sum']or 0,
        })
    term = request.GET.get('term').lower()

    # Remove non-alphanumeric characters and convert to lowercase


    # Split the cleaned term into individual words separated by '*'
    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:

            # term = ''.join(char for char in term if char.isalnum())
            q_objects &= (Q(ref__icontains=term) | Q(coderef__icontains=term) | Q(name__icontains=term) | Q(category__name__icontains=term) |  Q(mark__name__icontains=term) |  Q(equivalent__icontains=term)  |  Q(refeq1__icontains=term) |  Q(refeq2__icontains=term)  |  Q(block__icontains=term) | Q(refeq1__icontains=term) | Q(refeq1__icontains=term) | Q(sellprice__icontains=term)  | Q(buyprice__icontains=term)  | Q(cars__icontains=term)  | Q(diametre__icontains=term))
    products=Produit.objects.filter(q_objects)[:50]
    trs=''
    for i in products:
        trs+=f'''
        <tr ondblclick="createtab('pdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
              style="background:{'#f3d6d694;' if not i.isactive else '' }"
              class="product-row" data-product-id="{ i.id }"
              >
                  <td style="padding: 5px; font-weight: bold;" >
                      {i.ref.upper()}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.name}
                  </td>

                  <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
                      {i.buyprice if i.buyprice else 0}
                  </td>
                  <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
                      {i.sellprice}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center">
                      {i.remise}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center">
                      {i.prixnet}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
                      {i.stocktotal}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
                    <span class="stockfacture invisible">{i.stockfacture}</span>
                  </td>

                  <td style="padding: 5px; font-weight: bold;">
                    {i.diametre}
                </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-success">
                    {i.block}
                </td>
                  <td style="padding: 5px; font-weight: bold;">
                    {i.coderef}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[0] if i.getequivalent() else ''}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[1] if i.getequivalent() else ''}

                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[2] if i.getequivalent() else ''}

                  </td>
                <td style="padding: 5px; font-weight: bold;">
                      {i.mark}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.code}
                  </td>
                  <td style="padding: 5px; font-weight: bold;" class="text-danger"><span class="percentage invisible"> {round(i.getpercentage(), 2)}</span></td>
        '''
    return JsonResponse({
        'trs':trs,
        'stocktotal':Produit.objects.filter(q_objects).aggregate(Sum('stocktotal'))['stocktotal__sum']or 0,
        'stockfacture':Produit.objects.filter(q_objects).aggregate(Sum('stockfacture'))['stockfacture__sum']or 0,
    })

def loadstock(request):
    page = int(request.GET.get('page', 1))
    term = request.GET.get('term')
    notactive = request.GET.get('notactive')
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    if term=='0':
        print('>>>>>>>>>>>>', term=='0')
        if notactive=='1':
            print('from notactive')
            products = Produit.objects.filter(isactive=False)[start:end]
            trs=''
            for i in products:
                trs+=f"""
                    <tr ondblclick="createtab('pdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
                        style="background:{'#f3d6d694;' if not i.isactive else '' }"
                            data-product-id="{ i.id }" class="product-row notactive">
                            <td style="padding: 5px; font-weight: bold;" class="pe-2">
                                {i.ref.upper()}
                            </td>
                            <td style="padding: 5px; font-weight: bold;">
                                {i.name}
                            </td>

                            <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
                                {i.buyprice if i.buyprice else 0}
                            </td>
                            <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
                                {i.sellprice if i.sellprice else 0}
                            </td>
                            <td style="padding: 5px; font-weight: bold;" class="text-center">
                                {i.remise}
                            </td>
                            <td style="padding: 5px; font-weight: bold;" class="text-center">
                                {i.prixnet}
                            </td>
                            <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
                                {i.stocktotal}
                            </td>
                            <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
                                <span class="stockfacture invisible">{i.stockfacture}</span>
                            </td>

                            <td style="padding: 5px; font-weight: bold;">
                                {i.diametre}
                            </td>
                            <td style="padding: 5px; font-weight: bold;" class="text-success">
                                {i.block}
                            </td>
                            <td style="padding: 5px; font-weight: bold;">
                                {i.coderef}
                            </td>
                            <td style="padding: 5px; font-weight: bold;">
                                {i.getequivalent()[0] if i.getequivalent() else ''}
                            </td>
                            <td style="padding: 5px; font-weight: bold;">
                                {i.getequivalent()[1] if i.getequivalent() else ''}

                            </td>
                            <td style="padding: 5px; font-weight: bold;">
                                {i.getequivalent()[2] if i.getequivalent() else ''}

                            </td>
                            <td style="padding: 5px; font-weight: bold;">
                                {i.mark}
                            </td>
                             <td style="padding: 5px; font-weight: bold;"  class="text-danger">
                                {i.code}
                            </td>

                        </tr>
                """

            return JsonResponse({
                'trs':trs,
                'has_more': len(products) == per_page
            })
        products = Produit.objects.all()[start:end]
        trs=''
        for i in products:
            trs+=f'''
            <tr ondblclick="createtab('pdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
                style="background:{'#f3d6d694;' if not i.isactive else '' }"
                    data-product-id="{ i.id }" class="product-row">
                    <td style="padding: 5px; font-weight: bold;" class="pe-2">
                        {i.ref.upper()}
                    </td>
                    <td style="padding: 5px; font-weight: bold;">
                        {i.name}
                    </td>

                    <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
                        {i.buyprice if i.buyprice else 0}
                    </td>
                    <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
                        {i.sellprice}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center">
                        {i.remise}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center">
                        {i.prixnet}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
                        {i.stocktotal}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
                        <span class="stockfacture invisible">{i.stockfacture}</span>
                    </td>

                    <td style="padding: 5px; font-weight: bold;">
                        {i.diametre}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-success">
                        {i.block}
                    </td>
                    <td style="padding: 5px; font-weight: bold;">
                        {i.coderef}
                    </td>

                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[0] if i.getequivalent() else ''}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[1] if i.getequivalent() else ''}

                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[2] if i.getequivalent() else ''}

                  </td>

                    <td style="padding: 5px; font-weight: bold;">
                        {i.mark}
                    </td>
                    <td style="padding: 5px; font-weight: bold;"  class="text-danger">
                        {i.code}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-danger"><span class="percentage invisible"> {round(i.getpercentage(), 2)}</span></td>
                </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(products) == per_page
        })
    else:
        print('>>>>>>>>>>>>', term=='0')

        regex_search_term = term.replace('+', '*')

        # Split the term into individual words separated by '*'
        search_terms = regex_search_term.split('*')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:
            if term:

                q_objects &= (Q(ref__iregex=term) | Q(name__iregex=term) | Q(category__name__iregex=term) |  Q(mark__name__iregex=term))
        products=Produit.objects.filter(q_objects)[start:end]
        trs=''
        for i in products:
            trs+=f'''
            <tr ondblclick="createtab('pdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
                style="background:{'#f3d6d694;' if not i.isactive else '' }"
                    data-product-id="{ i.id }" class="product-row ">
                    <td style="padding: 5px; font-weight: bold;" >
                        {i.ref.upper()}
                    </td>
                    <td style="padding: 5px; font-weight: bold;">
                        {i.name}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center prachat">
                        {i.buyprice if i.buyprice else 0}
                    </td>
                    <td style="padding: 5px; font-weight: bold; font-size: 14px; color: var(--orange);" class="text-center">
                        {i.sellprice}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center">
                        {i.remise}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center">
                        {i.prixnet}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center text-danger stock">
                        {i.stocktotal}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-center stockfacture" style="color: blue;">
                        <span class="stockfacture invisible">{i.stockfacture}</span>
                    </td>

                    <td style="padding: 5px; font-weight: bold;">
                        {i.diametre}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-success">
                        {i.block}
                    </td>
                    <td style="padding: 5px; font-weight: bold;">
                        {i.coderef}
                    </td>
                    <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[0] if i.getequivalent() else ''}
                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[1] if i.getequivalent() else ''}

                  </td>
                  <td style="padding: 5px; font-weight: bold;">
                      {i.getequivalent()[2] if i.getequivalent() else ''}

                  </td>

                    <td style="padding: 5px; font-weight: bold;">
                        {i.mark}
                    </td>
                    <td style="padding: 5px; font-weight: bold;"  class="text-danger">
                        {i.code}
                    </td>
                    <td style="padding: 5px; font-weight: bold;" class="text-danger"><span class="percentage invisible"> {round(i.getpercentage(), 2)}</span></td>
                </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(products) == per_page
        })
def loadlistbl(request):
    page = int(request.GET.get('page', 1))
    year =request.GET.get('year')
    startdate =request.GET.get('startdate')
    enddate =request.GET.get('enddate')
    term =request.GET.get('term')
    per_page = 50  # Adjust as needed
    print(term, year, startdate, enddate)
    trs=''
    start = (page - 1) * per_page
    end = page * per_page
    if term != '0':
        regex_search_term = term.replace('+', '*')

        # Split the term into individual words separated by '*'
        search_terms = regex_search_term.split('*')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:
            if term:
                q_objects &= (Q(client__name__iregex=term) | Q(salseman__name__iregex=term) | Q(bon_no__iregex=term) | Q(total__iregex=term))
        if year=='0':
            bons=Bonlivraison.objects.filter(q_objects).filter(date__year=thisyear).order_by('-bon_no')[start:end]
            total=round(Bonlivraison.objects.filter(q_objects).filter(date__year=thisyear).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        else:
            bons=Bonlivraison.objects.filter(q_objects).filter(date__year=year).order_by('-bon_no')[start:end]
            total=round(Bonlivraison.objects.filter(q_objects).filter(date__year=year).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        for i in bons:
            trs+=f'''
            <tr  
            style="background: {"lightgreen;" if i.isdelivered else ""} color:{"blue" if i.isfacture else ""} "  
            class="ord {"text-danger" if i.ispaid else ''} bl-row" 
            year={year} 
            orderid="{i.id}" 
            ondblclick="createtab('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')"
            term="{term}">
                <td>{ i.bon_no }</td>
                    <td>{ i.date.strftime("%d/%m/%Y")}</td>
                    <td>{ i.client.name }</td>
                    <td>{ i.client.code }</td>
                    <td style="color: blue;">{ i.total}</td>
                    <td>{ i.client.region}</td>
                    <td>{ i.client.city}</td>
                    <td>{ "%.2f" % i.client.soldbl} </td>
                    <td>{ i.salseman }</td>
                    <td class="d-flex justify-content-between">
                    <div>
                    {'R0' if i.ispaid else 'N1' }

                    </div>
                    <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                    </td>
                    <td>
                    {'OUI' if i.isfacture else 'NON'}
                    </td>

                    <td>
                    
                    </td>
                    <td>
                    {i.modlvrsn}
                    </td>
                    <td>
                    {i.note}
                    </td>
                    <td>
                      <button class="btn btn-sm btn-info" onclick="makedelivered('{i.id}', event)"></button>
                    </td>
                    
              </tr>
            '''
        print('>>>load bl term')
        return JsonResponse({
            'trs':trs,
            'has_more': len(bons) == per_page
        })
    if startdate != '0' and enddate != '0':
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        bons=Bonlivraison.objects.filter(date__range=[startdate, enddate]).order_by('-bon_no')[start:end]
        total=round(Bonlivraison.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'] or 0, 2)
        for i in bons:
            trs+=f'''
            <tr 
            style="background: {"lightgreen;" if i.isdelivered else ""} color:{"blue" if i.isfacture else ""} " 
            class="ord {"text-danger" if i.ispaid else ''} bl-row" 
            year={year} orderid="{i.id}" 
            ondblclick="createtab('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')"
            startdate={startdate} enddate={enddate}>
                <td>{ i.bon_no }</td>
                    <td>{ i.date.strftime("%d/%m/%Y")}</td>
                    <td>{ i.client.name }</td>
                    <td>{ i.client.code }</td>
                    <td style="color: blue;">{ i.total}</td>
                    <td>{ i.client.region}</td>
                    <td>{ i.client.city}</td>
                    <td>{ "%.2f" % i.client.soldbl }</td>
                    <td>{ i.salseman }</td>
                    <td class="d-flex justify-content-between">
                    <div>
                    {'R0' if i.ispaid else 'N1' }

                    </div>
                    <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                    </td>
                    <td>
                    {'OUI' if i.isfacture else 'NON'}

                    </td>

                    <td>
                    
                    </td>
                    <td>
                    {i.modlvrsn}
                    </td>
                    <td>
                    {i.note}
                    </td>
                    <td>
                      <button class="btn btn-sm btn-info" onclick="makedelivered('{i.id}', event)"></button>
                    </td>

              </tr>
            '''
        print('>>>load bl date f')
        return JsonResponse({
            'trs':trs,
            'has_more': len(bons) == per_page
        })
    if year=="0":
        bons= Bonlivraison.objects.filter(date__year=thisyear).order_by('-bon_no')[start:end]
        total=round(Bonlivraison.objects.filter(date__year=thisyear).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
    else:
        bons= Bonlivraison.objects.filter(date__year=year).order_by('-bon_no')[start:end]
        total=round(Bonlivraison.objects.filter(date__year=year).order_by('-bon_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
    
    for i in bons:
        trs+=f'''
        <tr style="background: {"lightgreen;" if i.isdelivered else ""} color:{"blue" if i.isfacture else ""} " class="ord {"text-danger" if i.ispaid else ''} bl-row" year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')">
            <td>{ i.bon_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td style="color: blue;">{ i.total}</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ "%.2f" % i.client.soldbl}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td>
                {'OUI' if i.isfacture else 'NON'}

                </td>

                <td>
                
                </td>
                <td>
                {i.modlvrsn}
                </td>
                <td>
                {i.note}
                </td>
                <td>
                  <button class="btn btn-sm btn-info" onclick="makedelivered('{i.id}', event)"></button>
                </td>
          </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })

def loadlistbc(request):
    page = int(request.GET.get('page', 1))
    year =request.GET.get('year')
    startdate =request.GET.get('startdate')
    enddate =request.GET.get('enddate')
    term =request.GET.get('term')
    print(year, startdate, enddate, term)
    per_page = 50  # Adjust as needed
    trs=''

    start = (page - 1) * per_page
    end = page * per_page
    if term != '0':
        print('>>>>> in term')
        regex_search_term = term.replace('+', '*')

        # Split the term into individual words separated by '*'
        search_terms = regex_search_term.split('*')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:
            if term:
                q_objects &= (Q(client__name__iregex=term) | Q(salseman__name__iregex=term) | Q(bon_no__iregex=term) | Q(total__iregex=term))
        if year=='0':
            bons=Order.objects.filter(q_objects).filter(date__year=thisyear).order_by('-id')[start:end]
            total=round(Order.objects.filter(q_objects).filter(date__year=thisyear).order_by('-id').aggregate(Sum('total'))['total__sum'] or 0, 2)
        else:
            bons=Order.objects.filter(q_objects).filter(date__year=year).order_by('-id')[start:end]
            total=round(Order.objects.filter(q_objects).filter(date__year=year).order_by('-id').aggregate(Sum('total'))['total__sum'] or 0, 2)
        for i in bons:
            trs+=f'''
            <tr class="orderrow {'text-danger' if not i.isdelivered else ''}" year={year} term={term} startdate={startdate} enddate={enddate} orderid="{i.code}" ondblclick="createtab('command{i.id}', 'Commande {i.order_no}', '/products/boncommandedetails/{i.id}')">
                <td>{ i.order_no }</td>
                <td>{ i.date.strftime('%d/%m/%Y') }</td>
                <td>{ i.client.name if i.client else '' }</td>
                <td>{ i.client.code if i.client else '' }</td>
                <td>{ i.total}</td>
                <td>{ i.client.region if i.client else '' }</td>
                <td>{ i.client.city if i.client else '' }</td>
                <td>{ i.client.soldbl if i.client else '' }</td>
                <td>{ i.salseman }</td>
                <td>
                {"Non" if i.isclientcommnd else "OUI"}
                </td>
                <td>
                {"R1" if i.isdelivered else  "R0" }
                </td>

            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(bons) == per_page,
            'total':total
        })
    if startdate != '0' and enddate != '0':
        print('>>>>> in date fil')
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        bons=Order.objects.filter(date__range=[startdate, enddate]).order_by('-id')[start:end]
        total=round(Order.objects.filter(date__range=[startdate, enddate]).order_by('-id').aggregate(Sum('total'))['total__sum'] or 0, 2)
        for i in bons:
            trs+=f'''
            <tr class="orderrow {'text-danger' if not i.isdelivered else ''}" year={year} term={term} startdate={startdate} enddate={enddate} orderid="{i.code}" ondblclick="createtab('command{i.id}', 'Commande {i.order_no}', '/products/boncommandedetails/{i.id}')">
                <td>{ i.order_no }</td>
                <td>{ i.date.strftime('%d/%m/%Y') }</td>
                <td>{ i.client.name if i.client else '' }</td>
                <td>{ i.client.code if i.client else '' }</td>
                <td>{ i.total}</td>
                <td>{ i.client.region if i.client else '' }</td>
                <td>{ i.client.city if i.client else '' }</td>
                <td>{ i.client.soldbl if i.client else '' }</td>
                <td>{ i.salseman }</td>
                <td>
                {"Non" if i.isclientcommnd else "OUI"}
                </td>
                <td>
                {"R1" if i.isdelivered else  "R0" }
                </td>

            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(bons) == per_page,
            'total':total
        })
    if year=="0":
        print('>>>>> in year')

        bons= Order.objects.filter(date__year=thisyear).order_by('-id')[start:end]
        total=round(Order.objects.filter(date__year=thisyear).order_by('-id').aggregate(Sum('total'))['total__sum'] or 0, 2)
    else:
        print('>>>>> in year')

        bons= Order.objects.filter(date__year=year).order_by('-id')[start:end]
        total=round(Order.objects.filter(date__year=year).order_by('-id').aggregate(Sum('total'))['total__sum'] or 0, 2)
    for i in bons:
        trs+=f'''
        <tr class="orderrow {'text-danger' if not i.isdelivered else ''}" year={year} term={term} startdate={startdate} enddate={enddate} orderid="{i.code}" ondblclick="createtab('command{i.id}', 'Commande {i.order_no}', '/products/boncommandedetails/{i.id}')">
            <td>{ i.order_no }</td>
            <td>{ i.date.strftime('%d/%m/%Y') }</td>
            <td>{ i.client.name if i.client else '' }</td>
            <td>{ i.client.code if i.client else '' }</td>
            <td>{ i.total}</td>
            <td>{ i.client.region if i.client else '' }</td>
            <td>{ i.client.city if i.client else '' }</td>
            <td>{ i.client.soldbl if i.client else '' }</td>
            <td>{ i.salseman }</td>
            <td>
            {"Non" if i.isclientcommnd else "OUI"}
            </td>
            <td>
            {"R1" if i.isdelivered else  "R0" }
            </td>

        </tr>
        '''


    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })



def searchforlistbl(request):
    term=request.GET.get('term')
    year=request.GET.get('year')
    print('>>>>>>>>>>>', year)
    # we dont need this
    if(term==''):

        bons=Bonlivraison.objects.filter(date__year=thisyear)[:50]
        total=round(Bonlivraison.objects.filter(date__year=thisyear).aggregate(Sum('total'))['total__sum'] or 0, 2)
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''} bl-row" orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')">
                <td>{ i.bon_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.total}</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldbl}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td class="text-danger">

                </td>
                <td class="text-danger">
                {'OUI' if i.isfacture else 'NON'}

                </td>

                <td>
                
                
                </td>
                <td>
                {i.note}
                </td>
                <td>
                {i.modlvrsn}
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs
        })
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (
            Q(client__name__iregex=term) |
            Q(salseman__name__iregex=term) |
            Q(bon_no__iregex=term) |
            Q(total__iregex=term) |
            Q(client__code__iregex=term) |
            Q(client__region__iregex=term) |
            Q(client__city__iregex=term)|
            Q(statusreg=term.lower())|
            Q(statusfc=term.lower())
            )
    if year=='0':
        bons=Bonlivraison.objects.filter(q_objects).filter(date__year=thisyear).order_by('-id')[:50]
        total=round(Bonlivraison.objects.filter(q_objects).filter(date__year=thisyear).order_by('-id').aggregate(Sum('total'))['total__sum'] or 0, 2)
    else:
        bons=Bonlivraison.objects.filter(q_objects).filter(date__year=year).order_by('-id')[:50]
        total=round(Bonlivraison.objects.filter(q_objects).filter(date__year=year).order_by('-id').aggregate(Sum('total'))['total__sum'] or 0, 2)
    trs=''
    for i in bons:
        trs+=f'''
            <tr 
            style="background: {"lightgreen;" if i.isdelivered else ""} color:{"blue" if i.isfacture else ""} " 
            class="ord {"text-danger" if i.ispaid else ''} bl-row" 
            year={year} term={term} 
            orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')"
            >
                <td>{ i.bon_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td style="color: blue;">{ i.total}</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldbl}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td>
                {'OUI' if i.isfacture else 'NON'}

                </td>

                <td>
                
                </td>
                <td>
                {i.note}
                </td>
                <td>
                {i.modlvrsn}
                </td>
            </tr>
            '''

    return JsonResponse({
        'trs':trs,
        'total':total
    })


def loadlistfc(request):
    page = int(request.GET.get('page', 1))
    year =request.GET.get('year')
    startdate =request.GET.get('startdate')
    enddate =request.GET.get('enddate')
    term =request.GET.get('term')
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    if term != '0':
        regex_search_term = term.replace('+', '*')

        # Split the term into individual words separated by '*'
        search_terms = regex_search_term.split('*')
        print(search_terms)

        # Create a list of Q objects for each search term and combine them with &
        q_objects = Q()
        for term in search_terms:
            if term:
                q_objects &= (Q(client__name__iregex=term) | Q(salseman__name__iregex=term) | Q(bon_no__iregex=term) | Q(total__iregex=term))
        if year=='0':
            bons=Facture.objects.filter(q_objects).filter(date__year=thisyear).order_by('-facture_no')[start:end]
            total=round(Facture.objects.filter(q_objects).filter(date__year=thisyear).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
            totaltva=round(Facture.objects.filter(q_objects).filter(date__year=thisyear).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
        else:
            bons=Facture.objects.filter(q_objects).filter(date__year=year).order_by('-facture_no')[start:end]
            total=round(Facture.objects.filter(q_objects).filter(date__year=year).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
            totaltva=round(Facture.objects.filter(q_objects).filter(date__year=year).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''}
             fc-row"
                style="color:{"blue" if i.bon else ""} "
              year={year} term={term} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td class="text-danger">

                </td>

                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(bons) == per_page,
            'total':total,
            'totaltva':totaltva,
        })
    if startdate != '0' and enddate != '0':
        print('>>>>>>>>>>in start end dat')
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        print(startdate, enddate)
        bons=Facture.objects.filter(date__range=[startdate, enddate]).order_by('-facture_no')[start:end]
        total=round(Facture.objects.filter(date__range=[startdate, enddate]).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        totaltva=round(Facture.objects.filter(date__range=[startdate, enddate]).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''}
             fc-row"
                style="color:{"blue" if i.bon else ""} "
              startdate={startdate} enddate={enddate} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td class="text-danger">

                </td>

                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(bons) == per_page,
            'total':total,
            'totaltva':totaltva,
        })
    if year=="0":
        bons= Facture.objects.filter(date__year=timezone.now().year).order_by('-facture_no')[start:end]
        total=round(Facture.objects.filter(date__year=timezone.now().year).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        totaltva=round(Facture.objects.filter(date__year=timezone.now().year).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
        print('year',bons)
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''}
             fc-row"
                style="color:{"blue" if i.bon else ""} "
              year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td class="text-danger">

                </td>

                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(bons) == per_page,
            'total':total,
            'totaltva':totaltva,
        })
    else:
        bons= Facture.objects.filter(date__year=year).order_by('-facture_no')[start:end]
        total=round(Facture.objects.filter(date__year=year).order_by('-facture_no').aggregate(Sum('total'))['total__sum'] or 0, 2)
        totaltva=round(Facture.objects.filter(date__year=year).order_by('-facture_no').aggregate(Sum('tva'))['tva__sum'] or 0, 2)
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td class="text-danger">

                </td>

                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs,
            'has_more': len(bons) == per_page,
            'total':total,
            'totaltva':totaltva,
        })


def searchforlistfc(request):
    term=request.GET.get('term')
    year=request.GET.get('year')
    if(term==''):
        bons=Facture.objects.all()[:100]
        trs=''
        for i in bons:
            trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td class="text-danger">

                </td>

                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
            </tr>
            '''
        return JsonResponse({
            'trs':trs
        })
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if '&' in term:
            date_range = term.split('-')
            start_date = datetime.strptime(date_range[0].strip(), '%d/%m/%Y')
            end_date = datetime.strptime(date_range[1].strip(), '%d/%m/%Y')
            q_objects &= (
                Q(client__name__iregex=term)|
                Q(client__code__iregex=term)|
                Q(salseman__name__iregex=term)|
                Q(facture_no__iregex=term)|
                Q(bon__bon_no__iregex=term)|
                Q(client__region__iregex=term)|
                Q(total__iregex=term)|
                Q(date__range=[start_date, end_date])
                )
        else:
            q_objects &= (
                Q(client__name__iregex=term)|
                Q(client__code__iregex=term)|
                Q(salseman__name__iregex=term)|
                Q(facture_no__iregex=term)|
                Q(bon__bon_no__iregex=term)|
                Q(client__region__iregex=term)|
                Q(total__iregex=term)
            )
    if year=='0':
        bons=Facture.objects.filter(q_objects).filter(date__year=thisyear).order_by('-id')[:50]
        total=round(Facture.objects.filter(q_objects).filter(date__year=thisyear).order_by('-id').aggregate(Sum('total'))['total__sum'] or 0, 2)
    else:
        bons=Facture.objects.filter(q_objects).filter(date__year=year).order_by('-id')[:50]
        total=round(Facture.objects.filter(q_objects).filter(date__year=year).order_by('-id').aggregate(Sum('total'))['total__sum'] or 0, 2)
    trs=''
    for i in bons:
        trs+=f'''
            <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" term={term} year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
                <td>{ i.facture_no }</td>
                <td>{ i.date.strftime("%d/%m/%Y")}</td>
                <td>{ i.total}</td>
                <td>{ i.tva}</td>
                <td>{ i.client.name }</td>
                <td>{ i.client.code }</td>
                <td>{ i.client.region}</td>
                <td>{ i.client.city}</td>
                <td>{ i.client.soldfacture}</td>
                <td>{ i.salseman }</td>
                <td class="d-flex justify-content-between">
                <div>
                {'R0' if i.ispaid else 'N1' }

                </div>
                <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

                </td>
                <td class="text-danger">

                </td>

                <td>
                {i.bon.bon_no if i.bon else "--"}
                </td>
            </tr>
            '''
    return JsonResponse({
        'trs':trs,
        'total':total,

    })


def createnewclientaccount(request):
    clientcode=request.GET.get('clientcode')
    client= Client.objects.get(code=clientcode)
    olduser=client.user
    username=request.GET.get('username')
    password=request.GET.get('password')
    print('>>>>>>>>>', clientcode)
    #check for username
    

    # create user
    user=User.objects.create_user(username=username, password=password)
    # assign user to client
    group, created = Group.objects.get_or_create(name="clients")
    user.groups.add(group)
    user.save()
    client.user=user
    client.save()
    cart=Cart.objects.filter(user=olduser).first()
    wich=Wich.objects.filter(user=olduser).first()
    if cart:
        cart.user=user
        cart.save()
    if wich:
        wich.user=user
        wich.save()
    olduser.delete()
    return JsonResponse({
        'success':True
    })

def createnewrepaccount(request):
    repid=request.GET.get('repid')
    rep= Represent.objects.get(pk=repid)
    olduser=rep.user
    
    username=request.GET.get('username')
    password=request.GET.get('password')
    user=User.objects.filter(username=username).first()
    if user:
        return JsonResponse({
            'success':False,
            'error':'Username exist déja'
        })
    user=User.objects.create_user(username=username, password=password)
    # assign user to rep
    group, created = Group.objects.get_or_create(name="salsemen")
    user.groups.add(group)
    user.save()
    rep.user=user
    rep.save()
    cart=Cart.objects.filter(user=olduser).first()
    wich=Wich.objects.filter(user=olduser).first()
    if cart:
        cart.user=user
        cart.save()
    if wich:
        wich.user=user
        wich.save()
    olduser.delete()
    return JsonResponse({
        'success':True
    })




def yeardatabl(request):
    year=request.GET.get('year')
    # get all bls of that year
    bls=Bonlivraison.objects.filter(date__year=year).order_by('-id')[:50]
    trs=''
    for i in bls:
        trs+=f'''
        <tr class="ord {"text-danger" if i.ispaid else ''} bl-row" year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Bon livraison {i.bon_no}', '/products/bonlivraisondetails/{i.id}')">
            <td>{ i.bon_no }</td>
            <td>{ i.date.strftime("%d/%m/%Y")}</td>
            <td>{ i.client.name }</td>
            <td>{ i.client.code }</td>
            <td>{ i.total}</td>
            <td>{ i.client.region}</td>
            <td>{ i.client.city}</td>
            <td>{ i.client.soldbl}</td>
            <td>{ i.salseman }</td>
            <td class="d-flex justify-content-between">
              <div>
              {'R0' if i.ispaid else 'N1' }

              </div>
              <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

            </td>
            <td class="text-danger">
            {'OUI' if i.isfacture else 'NON'}

            </td>

            <td>
              {i.commande.order_no if i.commande else '--'}
            </td>
            <td>
              {i.modlvrsn}
            </td>
          </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':round(Bonlivraison.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    })

def yeardatabc(request):
    year=request.GET.get('year')
    print(year)
    # get all bls of that year
    bons=Order.objects.filter(date__year=year).order_by('-id')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="orderrow {'text-danger' if not i.isdelivered else ''}" year={year} orderid="{i.code}" ondblclick="createtab('command{i.id}', 'Commande {i.order_no}', '/products/boncommandedetails/{i.id}')">
            <td>{ i.order_no }</td>
            <td>{ i.date.strftime('%d/%m/%Y') }</td>
            <td>{ i.client.name if i.client else '' }</td>
            <td>{ i.client.code if i.client else '' }</td>
            <td>{ i.total}</td>
            <td>{ i.client.region if i.client else '' }</td>
            <td>{ i.client.city if i.client else '' }</td>
            <td>{ i.client.soldbl if i.client else '' }</td>
            <td>{ i.salseman }</td>
            <td>
            {"Non" if i.isclientcommnd else "OUI"}
            </td>
            <td>
              {"R1" if i.isdelivered else  "R0" }
            </td>

          </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':round(Order.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2)
    })



def yeardatafc(request):
    year=request.GET.get('year')
    print(year)
    # get all bls of that year
    bls=Facture.objects.filter(date__year=year).order_by("-id")[:50]
    trs=''
    for i in bls:
        trs+=f'''
        <tr class="ord {"text-danger" if i.ispaid else ''} fc-row" year={year} orderid="{i.id}" ondblclick="createtab('bonl{i.id}', 'Facture {i.facture_no}', '/products/facturedetails/{i.id}')">
            <td>{ i.facture_no }</td>
            <td>{ i.date.strftime("%d/%m/%Y")}</td>
            <td>{ i.total}</td>
            <td>{ i.tva}</td>
            <td>{ i.client.name }</td>
            <td>{ i.client.code }</td>
            <td>{ i.client.region}</td>
            <td>{ i.client.city}</td>
            <td>{ i.client.soldbl}</td>
            <td>{ i.salseman }</td>
            <td class="d-flex justify-content-between">
              <div>
              {'R0' if i.ispaid else 'N1' }

              </div>
              <div style="width:15px; height:15px; border-radius:50%; background:{'green' if i.ispaid else 'orange' };" ></div>

            </td>
            <td class="text-danger">

            </td>

            <td>
              {i.bon.bon_no if i.bon else "--"}
            </td>
          </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':round(bls.aggregate(Sum('total'))['total__sum'] or 0, 2)
    })



def loadreglbl(request):
    page = int(request.GET.get('page', 1))
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    print('start', start, end)
    reblbls = PaymentClientbl.objects.all().order_by('-id')[start:end]
    return JsonResponse({
        'trs':render(request, 'reglbllist.html', {'bons':reblbls}).content.decode('utf-8'),
        'has_more': len(reblbls) == per_page
    })


def laodreglfc(request):
    page = int(request.GET.get('page', 1))
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page

    reblbls = PaymentClientfc.objects.all()[start:end]
    # trs=''
    # for i in reblbls:
    #     tooltip_content = ''.join([bon.bon_no for bon in i.bons.all()])
    #     trs += f'''
    #     <tr style="{ "background: yellow;" if (i.echance and i.echance == today) else ("color: red;" if (i.echance and i.echance < today) else "") }" class="reglbl-row">
    #         <td>
    #             { i.mode } <br>
    #             <!-- <select name="updatemodereglbl{ i.id }" id=""></select> -->
    #         </td>
    #         <td>
    #             { i.npiece } <br>
    #             <input type="text" class="d-none updatenpiecereglbl{ i.id }">
    #         </td>
    #         <td>
    #             { i.amount }<br>
    #             <!-- <input type="text" class="d-none updateamountreglbl{ i.id }"> -->
    #         </td>
    #         <td>
    #             { i.date }<br>
    #             <input type="date" class="d-none updatetdatereglbl{ i.id }">
    #         </td>
    #         <td>
    #             { i.client.name }<br>
    #         </td>
    #         <td>
    #             { i.client.code }<br>
    #         </td>
    #         <td>
    #             { i.echance} <br>
    #         </td>
    #         <td class="d-flex justify-content-between">
    #             <button type="button" class="btn btn-secondary btn-sm" data-toggle="tooltip" data-placement="top" data-tooltip="Bon Nos: {tooltip_content}">
    #             </button>
    #             <button class="btn btn-success btn-sm" onclick="updatereglementbl({ i.id })">✐</button>
    #         </td>
    #     </tr>
    # '''

    return JsonResponse({
        'trs':render(request, 'reglfclist.html', {'bons':reblbls}).content.decode('utf-8'),
        'has_more': len(reblbls) == per_page
    })



def loadclients(request):
    page = int(request.GET.get('page', 1))
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page

    products = Client.objects.all()[start:end]
    trs=''
    for i in products:
        trs+=f'''
        <tr class="client-row">
            <td>
                <button class="btn editsuppbtn border" id="{i.id}" data-toggle="modal" data-target="#editclientmodal" onclick="populateclientfields({i.id})">
                    ✐
                </button>
            </td>
            <td onclick="createtab('client{i.id}', 'Client: {i.name} ', '/products/client/{i.id}')">{i.name} </td>
            <td>{i.code}</td>
            <td>{i.phone}</td>
            <td>{i.city}</td>
            <td>{i.region}</td>
            <td>

                {i.represent.name if i.represent else ''}


            </td>
            <td>{i.soldtotal}</td>
            <td style="background: yellowgreen;">{i.soldbl}</td>
            <td style="background: aliceblue;">{i.soldfacture}</td>
            <td>{i.ice}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == per_page
    })


def exportproducts(request):
    categoryid=request.GET.get('categoryid')
    if categoryid=='0':
        products=Produit.objects.all()
        filename='Produit_tous'+today.strftime('%d/%m/%y')+'.xlsx'
    else:
        category=Category.objects.get(pk=categoryid)
        products=Produit.objects.filter(category__id=categoryid)
        filename='Produit_'+category.name+today.strftime('%d%m%y')+'.xlsx'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


    # Create a new Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write column headers
    ws.append(['ref', 'name', 'category', 'buyprice', 'sellprice', 'remise', 'prixnet', 'stocktotal', 'stockfacture', 'mark', 'diametre', 'block', 'equivalent', 'refeq1', 'refeq2'])

    # Write product data
    for product in products:
        ws.append([
            product.ref, product.name,
            product.category.name if product.category else '',  # Extract category name
            product.buyprice, product.sellprice,
            product.remise, product.prixnet, product.stocktotal, product.stockfacture, product.mark.name if product.mark else '',
            product.diametre, product.block, product.equivalent, product.refeq1, product.refeq2
        ])

    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    # Save the workbook to the response
    wb.save(response)

    return response

def datepdct(request):
    datefrom=request.GET.get('datefrom')
    dateto=request.GET.get('dateto')
    print('dates>>>>>>', datefrom, dateto)
    datefrom=datetime.strptime(datefrom, '%Y-%m-%d')
    dateto=datetime.strptime(dateto, '%Y-%m-%d')
    stockin=Stockin.objects.filter(date__range=[datefrom, dateto])
    stockout=Bonlivraison.objects.filter(date__range=[datefrom, dateto])
    totalqtyin=stockin.aggregate(Sum('quantity'))['quantity__sum'] or 0
    totalqtyout=stockout.aggregate(Sum('qty'))['qty__sum'] or 0
    totalcoutin=stockin.aggregate(Sum('total'))['total__sum'] or 0
    totalrevenu=stockout.aggregate(Sum('total'))['total__sum'] or 0
    trin=''
    trout=''
    for i in stockin:
        trin+=f'''
            <tr>
                <td>
                    {i.date.strftime('%d/%m/%Y')}
                </td>
                <td>
                    {i.n_bon}
                </td>
                <td>
                    {i.supplier.name}
                </td>
                <td>
                    {i.quantity}
                </td>
                <td>
                    {i.price}
                </td>
                <td>
                    {i.total}
                </td>
            </tr>
        '''
    for i in stockout:
        trout+=f'''
            <tr>
                <td>
                    {i.date.strftime('%d/%m/%Y')}
                </td>
                <td>
                    {i.bon.bon_no}
                </td>
                <td>
                    {i.bon.client.name}
                </td>
                <td>
                    {i.qty}
                </td>
                <td>
                    {i.price}
                </td>
                <td>
                    {i.remise}%
                </td>
                <td>
                    {i.total}
                </td>
            </tr>
        '''
    return JsonResponse({
        'trin':trin,
        'trout':trout,
        'totalqtyin':totalqtyin,
        'totalqtyout':totalqtyout,
        'totalcoutin':totalcoutin,
        'totalrevenu':totalrevenu
    })


def showdeactivated(request):
    products=Produit.objects.filter(isactive=False)[:100]
    trs=''
    for i in products:
        trs+=f"""
            <tr ondblclick="createtab('pdct{i.id}', 'Produit {i.ref}', '/products/viewoneproduct/{i.id}')"
                style="background:{'#f3d6d694;' if not i.isactive else '' }"
                    data-product-id="{ i.id }" class="product-row notactive">
                    <td >
                        {i.ref.upper()}
                    </td>
                    <td>
                        {i.name}
                    </td>
                    <td>
                        {i.category}
                    </td>
                    <td class="text-center prachat">
                        {i.buyprice if i.buyprice else 0}
                    </td>
                    <td class="text-center">
                        {i.sellprice}
                    </td>
                    <td class="text-center">
                        {i.remise}
                    </td>
                    <td class="text-center">
                        {i.prixnet}
                    </td>
                    <td class="text-center text-danger stock">
                        {i.stocktotal}
                    </td>
                    <td class="text-center stockfacture" style="color: blue;">
                        <span class="stockfacture invisible">{i.stockfacture}</span>
                    </td>

                    <td>
                        {i.diametre}
                    </td>
                    <td class="text-success">
                        {i.block}
                    </td>
                    <td>
                        {i.coderef}
                    </td>
                    <td>

                    </td>
                    <td>

                    </td>
                    <td>

                    </td>
                    <td>
                        {i.mark}
                    </td>

                </tr>
        """

    return JsonResponse({
        'trs':trs,
        'has_more': len(products) == 100
    })

def searchforlistclient(request):
    term=request.GET.get('term')
    if(term==''):
        products=Client.objects.all()[:50]
        trs=''
        for i in products:
            trs+=f'''
            <tr class="client-row">
                <td>
                    <button class="btn editsuppbtn border" id="{i.id}" data-toggle="modal" data-target="#editclientmodal" onclick="populateclientfields({i.id})">
                        ✐
                    </button>
                </td>
                <!-- <a href="onclick="createtab('client{i.id}', 'client: {i.name}', '/products/clientlier/{i.id}')""></a> -->
                <td onclick="createtab('client{i.id}', 'Client: {i.name}', '/products/client/{i.id}')">{i.name}</td>
                <td >{i.code}</td>
                <td>{i.phone}</td>
                <td>{i.city}</td>
                <td>{i.region.upper}</td>
                <td>
                    {i.represent.name if i.represent else ''}
                </td>
                <td>{i.soldtotal}</td>
                <td style="background: yellowgreen;">{i.soldbl}</td>
                <td style="background: aliceblue;">{i.soldfacture}</td>
                <td>{i.ice}</td>
            </tr>
        '''
        return JsonResponse({
            'trs':trs
        })
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:

            # term = ''.join(char for char in term if char.isalnum())
            q_objects &= (Q(city__icontains=term) | 
                Q(name__icontains=term) | 
                Q(ice__icontains=term) | 
                Q(phone__icontains=term) | 
                Q(region__icontains=term) | 
                Q(code__icontains=term) | 
                Q(represent__name__icontains=term) | 
                Q(address__icontains=term))
    products=Client.objects.filter(q_objects)
    trs=''
    for i in products:
        trs+=f'''
        <tr class="">
            <td>
                <button term={term} class="btn editsuppbtn border" id="{i.id}" data-toggle="modal" data-target="#editclientmodal" onclick="populateclientfields({i.id})">
                    ✐
                </button>
            </td>
            <!-- <a href="onclick="createtab('client{i.id}', 'client: {i.name}', '/products/clientlier/{i.id}')""></a> -->
            <td onclick="createtab('client{i.id}', 'Client: {i.name}', '/products/client/{i.id}')">{i.name}</td>
            <td >{i.code}</td>
            <td>{i.phone}</td>
            <td>{i.city}</td>
            <td>{i.region.upper() if i.region else ''}</td>
            <td>
                {i.represent.name if i.represent else ''}
            </td>
            <td>{i.soldtotal}</td>
            <td style="background: yellowgreen;">{i.soldbl}</td>
            <td style="background: aliceblue;">{i.soldfacture}</td>
            <td>{i.ice}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs
    })


def laodblregl(request):
    page = int(request.GET.get('page', 1))
    clientid=request.GET.get('clientid')
    #nr: non reglé
    nr=request.GET.get('nr')
    per_page = 100  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    print('>>>>>>>>>', nr)
    if nr=='0':
        print('all')
        bons=Bonlivraison.objects.filter(client_id=clientid)[start:end]
        for i in bons:
            trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)" {"checked" if i.reglements.exists() else ""}></td></tr>'
    else:
        print('non regl')
        bons=Bonlivraison.objects.filter(client_id=clientid, ispaid=False)[start:end]
        for i in bons:
            trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow nr" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)"></td></tr>'



    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })


def laodblinupdateregl(request):
    page = int(request.GET.get('page', 1))

    reglementid=request.GET.get('reglementid')
    print('>>>>>>', reglementid)
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    reglement=PaymentClientbl.objects.get(pk=reglementid)

    bons=reglement.bons.all()
    # bons without bons in reglement
    bons=Bonlivraison.objects.filter(client=reglement.client).exclude(pk__in=[bon.pk for bon in bons]).order_by('-id')[start:end]
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="loadblinupdateregl" reglemntid="{reglementid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)"></td></tr>'



    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })

def laodfcinupdateregl(request):
    page = int(request.GET.get('page', 1))

    reglementid=request.GET.get('reglementid')
    print('>>>>>>', reglementid)
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    reglement=PaymentClientfc.objects.get(pk=reglementid)

    bons=reglement.factures.all()
    # bons without bons in reglement
    bons=Facture.objects.filter(client=reglement.client).exclude(pk__in=[bon.pk for bon in bons]).order_by('-id')[start:end]
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="loadblinupdatereglfc" reglemntid="{reglementid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'



    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })

# the facture load in the add reglement
def laodfcregl(request):
    page = int(request.GET.get('page', 1))
    clientid=request.GET.get('clientid')
    #nr: non reglé
    nr=request.GET.get('nr')
    per_page = 50  # Adjust as needed

    start = (page - 1) * per_page
    end = page * per_page
    trs=''
    if nr=='0':
        bons=Facture.objects.filter(client_id=clientid).order_by('-id')[start:end]
        for i in bons:
            trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="fcreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)" {"checked" if i.reglementsfc.exists() else ""}></td></tr>'
    else:
        bons=Facture.objects.filter(client_id=clientid, ispaid=False).order_by('-id')[start:end]
        for i in bons:
            trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="fcreglrow nr" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'



    return JsonResponse({
        'trs':trs,
        'has_more': len(bons) == per_page
    })

def searchclientbls(request):
    clientid=request.GET.get('clientid')
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Use list comprehension for building search terms
    search_terms = [term for term in regex_search_term.split('*') if term]

    # Use __icontains for case-insensitive searches
    q_objects = Q()
    q_objects &= (Q(salseman__name__icontains=term) | Q(bon_no__icontains=term) | Q(total__icontains=term))

    # Combine filter conditions in one line
    bons = Bonlivraison.objects.filter(client_id=clientid).filter(q_objects)[:10]

    trs=''
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="bonstopay" onchange="checkreglementbox(event)" {"checked" if i.reglements.exists() else ""}></td></tr>'
    return JsonResponse({
        'trs':trs
    })


def searchclientfcs(request):
    clientid=request.GET.get('clientid')
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(salseman__name__iregex=term) | Q(facture_no__iregex=term) | Q(total__iregex=term))
    bons=Facture.objects.filter(client_id=clientid).filter(q_objects)[:10]
    print(bons)
    trs=''
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="blreglrow" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.client.name}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)" {"checked" if i.reglementsfc.exists() else ""}></td></tr>'
    return JsonResponse({
        'trs':trs
    })

# search fc in update reglement
def searchclientfcsupdatereg(request):
    clientid=request.GET.get('clientid')
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(salseman__name__iregex=term) | Q(facture_no__iregex=term) | Q(total__iregex=term))
    bons=Facture.objects.filter(client_id=clientid).filter(q_objects)[:10]
    print(bons)
    trs=''
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglementsfc.exists() else ""}" class="" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.facture_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglementsfc.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'trs':trs
    })

def searchclientblsupdatereg(request):
    clientid=request.GET.get('clientid')
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(salseman__name__iregex=term) | Q(bon_no__iregex=term) | Q(total__iregex=term))
    bons=Bonlivraison.objects.filter(client_id=clientid).filter(q_objects)[:10]
    print(bons, clientid)
    trs=''
    for i in bons:
        trs+=f'<tr style="background: {"rgb(221, 250, 237);" if i.reglements.exists() else ""}" class="" clientid="{clientid}"><td>{i.date.strftime("%d/%m/%Y")}</td><td>{i.bon_no}</td><td>{i.total}</td><td class="text-danger">{"RR" if i.reglements.exists() else "NR"}</td> <td><input type="checkbox" value="{i.id}" name="facturestopay" onchange="checkreglementbox(event)"></td></tr>'
    return JsonResponse({
        'trs':trs
    })



def searchregl(request):
    term=request.GET.get('term')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(client__name__icontains=term) | Q(mode__icontains=term) | Q(npiece__icontains=term) | Q(amount__icontains=term))
    regls=PaymentClientbl.objects.filter(q_objects)[:50]
    return JsonResponse({
        'trs':render(request, 'reglbllist.html', {'bons':regls}).content.decode('utf-8'),

    })


def searchreglfc(request):
    term=request.GET.get('term')
    year=request.GET.get('year')
    q_objects = Q()
    if term=='':
        if year=='0':
            regls=PaymentClientfc.objects.filter(date__year=thisyear).filter(q_objects).order_by('-date')[:50]
        else:
            regls=PaymentClientfc.objects.filter(date__year=year).filter(q_objects).order_by('-date')[:50]
        return JsonResponse({
            'trs':render(request, 'reglfclist.html', {'bons':regls, 'today':timezone.now().date()}).content.decode('utf-8'),
        })
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &

    for term in search_terms:
        if term:
            q_objects &= (Q(client__name__iregex=term) | Q(mode__iregex=term) | Q(npiece__iregex=term) | Q(amount__iregex=term))
    if year=='0':
        regls=PaymentClientfc.objects.filter(date__year=thisyear).filter(q_objects).order_by('-date')[:50]
    else:
        regls=PaymentClientfc.objects.filter(date__year=year).filter(q_objects).order_by('-date')[:50]
    return JsonResponse({
        'trs':render(request, 'reglfclist.html', {'bons':regls, 'today':timezone.now().date()}).content.decode('utf-8'),

    })

def brahim(request):
    # if post method
    if request.method=='POST':
        # get user and password
        username=request.POST.get('username')
        password=request.POST.get('password')
        # check if user exist
        user=authenticate(username=username, password=password)
        if user:
            group=user.groups.all().first().name
            if group == 'admin':
                login(request, user)
                return redirect('main:ibra')
    if request.user.groups.all():
        group=request.user.groups.all().first().name
        if group == 'admin':
            return redirect('main:ibra')
    return render(request, 'brahim.html')

def yeardatareglfc(request):
    year=request.GET.get('year')
    regls=PaymentClientfc.objects.filter(date__year=year).order_by('-date')[:50]
    return JsonResponse({
        'trs':render(request, 'reglfclist.html', {'bons':regls, 'today':timezone.now().date(), 'year':{year}}).content.decode('utf-8'),
    })

def filterbcdate(request):
    startdate=request.GET.get('startdate')
    enddate=request.GET.get('enddate')
    print(startdate, enddate)
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Order.objects.filter(date__range=[startdate, enddate]).order_by('-id')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="ord orderrow {'text-danger' if not i.isdelivered else ''}"  startdate={startdate} enddate={enddate} orderid="{i.id}" ondblclick="createtab('command{i.id}', 'Commande {i.order_no}', '/products/boncommandedetails/{i.id}')">
            <td>{ i.order_no }</td>
            <td>{ i.date.strftime('%d/%m/%Y') }</td>
            <td>{ i.client.name if i.client else '' }</td>
            <td>{ i.client.code if i.client else '' }</td>
            <td>{ i.total}</td>
            <td>{ i.client.region if i.client else '' }</td>
            <td>{ i.client.city if i.client else '' }</td>
            <td>{ i.client.soldbl if i.client else '' }</td>
            <td>{ i.salseman }</td>
            <td>
            {'Non' if i.isclientcommnd else 'Oui'}

            </td>
            <td>
              {'R1' if i.isdelivered else 'R0'}
            </td>
          </tr>
        '''
    ctx={
        'trs':trs
    }
    if bons:
        ctx['total']=round(Order.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total')).get('total__sum'), 2)
    return JsonResponse(ctx)


def deletebonachat(request):

    bon=Itemsbysupplier.objects.get(pk=request.GET.get('id'))
    items=Stockin.objects.filter(nbon=bon)
    for i in items:
        i.product.stocktotal=int(i.product.stocktotal)-int(i.quantity)
        if bon.isfacture:
            i.product.stockfacture=int(i.product.stockfacture)-int(i.quantity)

    items.delete()
    bon.delete()
    return JsonResponse({
        'success':True
    })



def searchforjv(request):
    term=request.GET.get('term')
    year=request.GET.get('year')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    # Create a list of Q objects for each search term and combine them with &

    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(client__name__iregex=term)|Q(ref__iregex=term)|Q(name__iregex=term)|Q(total__iregex=term)|Q(bon__bon_no__iregex=term))
    if year=='0':
        # means the year i not selected, so the records of the current year
        products = Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).filter(q_objects).order_by('-date')[:50]
        total=round(Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Livraisonitem.objects.filter(isfacture=False, date__year=thisyear).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
    else:
        products = Livraisonitem.objects.filter(isfacture=False, date__year=year).filter(q_objects).order_by('-date')[:50]
        total=round(Livraisonitem.objects.filter(isfacture=False, date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Livraisonitem.objects.filter(isfacture=False, date__year=year).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
    trs=''
    print('>>>>>>', products)
    for i in products:
        trs+=f'''
        <tr class="journalvente-row" year={year} term={term}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.bon.bon_no}</td>
            <td>{i.product.ref.upper()}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjv">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjv">{i.qty}</td>
            <td class="totaljv">{i.total}</td>
            <td>{i.bon.client.name}</td>
            <td>{i.bon.salseman.name}</td>
            <td class="text-success margejv">

            </td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':total,
        'totalqty':totalqty
    })

def searchforjvfc(request):
    term=request.GET.get('term')
    year=request.GET.get('year')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    # Create a list of Q objects for each search term and combine them with &

    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(client__name__iregex=term)|Q(ref__iregex=term)|Q(name__iregex=term)|Q(total__iregex=term)|Q(facture__facture_no__iregex=term))
    if year=='0':
        # means the year i not selected, so the records of the current year
        products = Outfacture.objects.filter(date__year=thisyear).filter(q_objects).order_by('-date')[:50]
        total=round(Outfacture.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Outfacture.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
    else:
        products = Outfacture.objects.filter(date__year=year).filter(q_objects).order_by('-date')[:50]
        total=round(Outfacture.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Outfacture.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('qty'))['qty__sum'] or 0
    trs=''
    for i in products:
        trs+=f'''
        <tr class="journalventefc-row" year={year} term={term}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.facture.facture_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjvfc">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjvfc">{i.qty}</td>
            <td class="totaljvfc">{i.total}</td>
            <td></td>
            <td>{i.facture.client.name}</td>
            <td>{i.facture.salseman.name}</td>
            <td class="text-success margejvfc">

            </td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':total,
        'totalqty':totalqty
    })

def filterjvdate(request):
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Livraisonitem.objects.filter(isfacture=False, date__range=[startdate, enddate]).order_by('-date')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="journalvente-row" startdate={startdate} enddate={enddate}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.bon.bon_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjv">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjv">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjv">{i.qty}</td>
            <td class="totaljv">{i.total}</td>
            <td>{i.bon.client.name}</td>
            <td>{i.bon.salseman.name}</td>
            <td class="text-success margejv">

            </td>
        </tr>
        '''
    ctx={
        'trs':trs
    }
    if bons:
        ctx['total']=round(Livraisonitem.objects.filter(isfacture=False, date__range=[startdate, enddate]).aggregate(Sum('total')).get('total__sum'), 2)
        ctx['qty']=round(Livraisonitem.objects.filter(isfacture=False, date__range=[startdate, enddate]).aggregate(Sum('qty')).get('qty__sum'), 2)
    return JsonResponse(ctx)

def filterjvfcdate(request):
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Outfacture.objects.filter(date__range=[startdate, enddate]).order_by('-date')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="journalventefc-row" startdate={startdate} enddate={enddate}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.facture.facture_no}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td class="prnetjvfc">{i.product.prixnet if i.product.prixnet else 0}</td>
            <td style="color:blue" class="coutmoyenjvfc">{i.product.coutmoyen if i.product.coutmoyen else 0}</td>
            <td class="text-danger">{i.product.buyprice if i.product.buyprice else 0}</td>
            <td class="text-danger qtyjvfc">{i.qty}</td>
            <td class="totaljvfc">{i.total}</td>
            <td></td>
            <td>{i.facture.client.name}</td>
            <td>{i.facture.salseman.name}</td>
            <td class="text-success margejvfc">

            </td>
        </tr>
        '''
    ctx={
        'trs':trs,
    }
    if bons:
        ctx['total']=round(Outfacture.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'], 2)
        ctx['totalqty']=Outfacture.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('qty')).get('qty__sum')
    return JsonResponse(ctx)

def filterjachdate(request):
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Stockin.objects.filter(date__range=[startdate, enddate]).order_by('-date')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="jach-row" startdat={startdate} enddate={enddate}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''
    ctx={
        'trs':trs,
    }
    if bons:
        ctx['total']=round(Stockin.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'], 2)
        ctx['totalqty']=Stockin.objects.filter(date__range=[startdate, enddate]).aggregate(Sum('quantity'))['quantity__sum']
    return JsonResponse(ctx)

def searchforjach(request):
    term=request.GET.get('term')
    year=request.GET.get('year')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    # Create a list of Q objects for each search term and combine them with &

    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(supplier__name__iregex=term)|Q(product__ref__iregex=term)|Q(product__name__iregex=term)|Q(total__iregex=term)|Q(nbon__nbon__iregex=term))
    if year=='0':
        print('thisyear')
        # means the year i not selected, so the records of the current year
        products = Stockin.objects.filter(date__year=thisyear).filter(q_objects).order_by('-date')[:50]
        total=round(Stockin.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Stockin.objects.filter(date__year=thisyear).filter(q_objects).aggregate(Sum('quantity'))['quantity__sum'] or 0
    else:
        products = Stockin.objects.filter(date__year=year).filter(q_objects).order_by('-date')[:50]
        total=round(Stockin.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Stockin.objects.filter(date__year=year).filter(q_objects).aggregate(Sum('quantity'))['quantity__sum'] or 0
    trs=''
    for i in products:
        trs+=f'''
        <tr class="jach-row" year={year} term={term}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':total,
        'totalqty':totalqty
    })

def yeardatajach(request):
    year=request.GET.get('year')
    print(year)
    items=Stockin.objects.filter(date__year=year).order_by('-date')[:50]
    trs=''
    totalmarge=0
    for i in items:
        try:
            marge_value = round((i.product.prixnet - (i.product.coutmoyen if i.product.coutmoyen else 0)) * i.qty, 2)
        except:
            marge_value = 0
        totalmarge+=marge_value
        trs+=f'''
        <tr class="jach-row" year={year}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''

    return JsonResponse({
        'trs':trs,
        'totalqty':Stockin.objects.filter(date__year=year).aggregate(Sum('quantity'))['quantity__sum'] or 0,
        'total':round(Stockin.objects.filter(date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalmarge':round(totalmarge, 2)
    })

def filterjachfcdate(request):
    startdate=request.GET.get('datefrom')
    enddate=request.GET.get('dateto')
    startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = datetime.strptime(enddate, '%Y-%m-%d')
    bons=Stockin.objects.filter(facture=True, date__range=[startdate, enddate]).order_by('-date')[:50]
    trs=''
    for i in bons:
        trs+=f'''
        <tr class="jachfc-row" startdat={startdate} enddate={enddate}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''
    ctx={
        'trs':trs,
    }
    if bons:
        ctx['total']=round(Stockin.objects.filter(facture=True, date__range=[startdate, enddate]).aggregate(Sum('total'))['total__sum'], 2)
        ctx['totalqty']=Stockin.objects.filter(facture=True, date__range=[startdate, enddate]).aggregate(Sum('quantity'))['quantity__sum']
    return JsonResponse(ctx)

def searchforjachfc(request):
    term=request.GET.get('term')
    year=request.GET.get('year')
    regex_search_term = term.replace('+', '*')

    # Split the term into individual words separated by '*'
    search_terms = regex_search_term.split('*')
    # Create a list of Q objects for each search term and combine them with &

    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (Q(supplier__name__iregex=term)|Q(product__ref__iregex=term)|Q(product__name__iregex=term)|Q(total__iregex=term)|Q(nbon__nbon__iregex=term))
    if year=='0':
        print('thisyear')
        # means the year i not selected, so the records of the current year
        products = Stockin.objects.filter(facture=True, date__year=thisyear).filter(q_objects).order_by('-date')[:50]
        total=round(Stockin.objects.filter(facture=True, date__year=thisyear).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Stockin.objects.filter(facture=True, date__year=thisyear).filter(q_objects).aggregate(Sum('quantity'))['quantity__sum'] or 0
    else:
        products = Stockin.objects.filter(facture=True, date__year=year).filter(q_objects).order_by('-date')[:50]
        total=round(Stockin.objects.filter(facture=True, date__year=year).filter(q_objects).aggregate(Sum('total'))['total__sum'] or 0, 2)
        totalqty=Stockin.objects.filter(facture=True, date__year=year).filter(q_objects).aggregate(Sum('quantity'))['quantity__sum'] or 0
    trs=''
    for i in products:
        trs+=f'''
        <tr class="jachfc-row" year={year} term={term}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''
    return JsonResponse({
        'trs':trs,
        'total':total,
        'totalqty':totalqty
    })

def yeardatajachfc(request):
    year=request.GET.get('year')
    print(year)
    items=Stockin.objects.filter(facture=True, date__year=year).order_by('-date')[:50]
    trs=''
    totalmarge=0
    for i in items:
        try:
            marge_value = round((i.product.prixnet - (i.product.coutmoyen if i.product.coutmoyen else 0)) * i.qty, 2)
        except:
            marge_value = 0
        totalmarge+=marge_value
        trs+=f'''
        <tr class="jachfc-row" year={year}>
            <td>{i.date.strftime('%d/%m/%Y')}</td>
            <td>{i.product.ref}</td>
            <td>{i.product.name}</td>
            <td>{i.price}</td>
            <td>{i.supplier.name}</td>
            <td>{i.devise}</td>
            <td class="qtyjournalachat">{i.quantity}</td>
            <td class="totaljournalachat">{i.total}</td>
        </tr>
        '''

    return JsonResponse({
        'trs':trs,
        'totalqty':Stockin.objects.filter(facture=True, date__year=year).aggregate(Sum('quantity'))['quantity__sum'] or 0,
        'total':round(Stockin.objects.filter(facture=True, date__year=year).aggregate(Sum('total'))['total__sum'] or 0, 2),
        'totalmarge':round(totalmarge, 2)
    })

def updaterepdata(request):
    caneditprice=request.GET.get('caneditprice')
    repid=request.GET.get('repid')
    rep=Represent.objects.get(pk=repid)
    rep.caneditprice=caneditprice
    rep.save()

    return JsonResponse({
        'success':True
    })

def makebondelivered(request):
    id=request.GET.get('id')
    bon=Bonlivraison.objects.get(pk=id)
    bon.isdelivered=True
    bon.save()
    return JsonResponse({
        'success':True
    })

def getitemsforlistbl(request):
    term=request.GET.get('term')
    search_terms = term.split('+')
    print(search_terms)

    # Create a list of Q objects for each search term and combine them with &
    q_objects = Q()
    for term in search_terms:
        if term:
            q_objects &= (
                Q(ref__icontains=term) | 
                Q(name__icontains=term) | 
                Q(mark__name__icontains=term) | 
                Q(category__name__icontains=term) |
                Q(equivalent__icontains=term) | 
                Q(refeq1__icontains=term) | 
                Q(refeq2__icontains=term) | 
                Q(refeq3__icontains=term) | 
                Q(refeq4__icontains=term) | 
                Q(diametre__icontains=term)| 
                Q(cars__icontains=term)
            )
    # check if term in product.ref or product.name
    products=Produit.objects.filter(q_objects)
    trs=[f'''<tr>
    <td><img src={i.image.url if i.image else ''}></td>
    <td>{i.ref.upper()}</td>
    <td>{i.name.upper()}</td>
    <td style="color: #ff6409;
    font-weight: bold;">{i.stocktotal}</td>
    <td style="color:blue;font-weight: bold;">{i.sellprice}</td>
    <td>{i.remise}</td>
    <td>{i.prixnet}</td>
    <td>{i.diametre}</td>
    </tr>
    ''' for i in products]
    return JsonResponse({
        'trs':trs
    })
def refspage(request):
    refs=Refstats.objects.all().order_by('-lastdate')
    refserver=[]
    for i in refs:
        if i.user.groups.all().first().name=='clients':
            refserver.append([i.user.client.name, i.ref, i.times, i.lastdate, i.id])
        elif i.user.groups.all().first().name=='salsemen':
            refserver.append([i.user.represent.name, i.ref, i.times, i.lastdate, i.id])
    return JsonResponse({
        'refs':refserver
    })
    #return render(request, 'refspage.html', {'refs':refs})

def updateadmindata(request):
    from django.contrib.auth.hashers import make_password
    username=request.POST.get('adminusername')
    password=request.POST.get('adminpassword')
    # Check if the username already exists
    existing_user = User.objects.filter(username=username).exclude(id=request.user.id).first()
    print('>>>>>>>',existing_user)
    if existing_user:
        return JsonResponse({
            'success': False,
            'error': 'Username already exists.'
        })
    else:
        # Update user data
        print('>>> updating')
        user = User.objects.get(pk=request.user.id)
        user.username = username
        user.password = make_password(password)  # Ensure the password is securely hashed
        user.save()
       
        return JsonResponse({
            'success': True,
        })

def notavailable(request):
    ctx={
    'products':Notavailable.objects.all()
    }
    return render(request, 'notavailable.html', ctx)

def cartpage(request):
    ctx={
        'carts':Cart.objects.all().order_by('-total')
    }
    return render(request, 'cartspage.html', ctx)

def excelnotav(request):
    myfile = request.FILES['excelFile']
    df = pd.read_excel(myfile)
    entries=0
    for d in df.itertuples():
        try:
            ref = d.ref.lower().strip()
        except:
            ref=d.ref
        #reps=json.dumps(d.rep)
        name = d.name
        mark = None if pd.isna(d.mark) else d.mark
        #refeq = '' if pd.isna(d.refeq) else d.refeq
        #status = False if pd.isna(d.status) else True
        #coderef = '' if pd.isna(d.code) else d.code
        #diam = '' if pd.isna(d.diam) else d.diam
        #qty = 0 if pd.isna(d.qty) else d.qty
        #buyprice = 0 if pd.isna(d.buyprice) else d.buyprice
        #devise = 0 if pd.isna(d.devise) else d.devise
        # car = None if pd.isna(d.car) else d.car
        #prixbrut = 0 if pd.isna(d.prixbrut) else d.prixbrut
        #ctg = None if pd.isna(d.ctg) else d.ctg
        #order = '' if pd.isna(d.order) else d.order
        #img = None if pd.isna(d.img) else d.img
        #remise = 0 if pd.isna(d.remise) else d.remise
        #prixnet=0 if pd.isna(d.prixnet) else d.prixnet
        product=Produit.objects.filter(ref=ref).first()
        if not product:
            
            exis=Notavailable.objects.filter(ref=ref).first()
            if not exis:
                product=Notavailable.objects.create(
                    ref=ref,
                #     equivalent=refeq,
                #     isactive=False,
                #     coderef=coderef,
                    name=name,
                    mark_id=mark,
                #     sellprice=prixbrut,
                #     prixnet=prixnet,
                #     remise=remise,
                #     category_id=ctg,
                    image=f'/products_imags/tette.jpg',
                #     stockfacture=qty,
                #     #diametre=diam,
                #     buyprice=buyprice,
                #     devise=devise
                )
            else:
                print(ref, 'exist in hors stock')
        else:
            print(ref, 'exist in products')
    
    return JsonResponse({
        'success':True
    })
# this to send carts to local server
def getcarts(request):
    carts=Cart.objects.all().order_by('-total').exclude(total=0)
    data=[]
    for i in carts:
        if i.user.groups.all().first().name=='clients':
            data.append([i.user.id, i.user.username, i.total, i.user.client.name])
        elif i.user.groups.all().first().name=='salsemen':
            data.append([i.user.id, i.user.username, i.total, i.user.represent.name])
    return JsonResponse({
        'carts':data
    })

def getwishs(request):
    carts=Wich.objects.all().exclude(total=0)
    data=[]
    for i in carts:
        data.append([i.user.id, i.user.username, i.total, i.user.client.name])
    return JsonResponse({
        'carts':data
    })

def addnotification(request):
    notification=request.GET.get('notification')
    Notification.objects.create(notification=notification)
    return JsonResponse({
        'success':True
        })

def updatenotification(request):
    notificationid=request.GET.get('notificationid')
    notification=request.GET.get('notification')
    notif=Notification.objects.get(pk=notificationid)
    notif.notification=notification
    notif.save()
    return JsonResponse({
        'success':True
    })

def allowcatalog(request):
    clientcode=request.GET.get('clientcode')
    client=Client.objects.get(code=clientcode)
    client.accesscatalog= not client.accesscatalog
    client.save()
    return JsonResponse({
        'success':True
    })
def minidashboard(request):
    return render(request, 'minidashboard.html')

def getcommandnumber(request):
    orders=Order.objects.filter(senttoserver=False)
    length=orders.count()
    if len(orders)==0:
        return JsonResponse({
            'success':False,
            'length':0,
            'message':'Aucune commande à envoyer'
        })
    orderstosend=[]
    #orderitemsstosend=[]
    for order in orders:
        orderdata={
            'id':order.id,
            'order_no':order.order_no,
            'isclientcommnd':order.isclientcommnd,
            'note':order.note,
            'clientcode':order.client.code,
            'salsemanid':order.salseman.id if order.salseman else None,
            'date':order.date,
            'items':[]
        }
        orderstosend.append(orderdata)
        orderitems=Orderitem.objects.filter(order=order)
        thisordertotal=0
        for item in orderitems:
            thisordertotal+=round(item.product.sellprice*item.qty, 2)
            orderitemsdata={
                'ordernumber':item.order.order_no,
                #uniqcode will be the connection
                'uniqcode':item.product.uniqcode,
                'qty':item.qty,
                'price':item.price,
                'total':item.total,
            }
            orderdata['items'].append(orderitemsdata)
            #orderitemsstosend.append(orderitemsdata)
        orderdata['total']=thisordertotal
    orders.update(senttoserver=True)
    return JsonResponse({
        'success':True,
        'length':length,
        'orders':orderstosend,
        #'items':orderitemsstosend
    })
    

def sendcommandstoserver(request):
    orders=Order.objects.filter(senttoserver=False)
    if len(orders)==0:
        return JsonResponse({
            'success':False,
            'message':'Aucune commande à envoyer'
        })
    orderstosend=[]
    #orderitemsstosend=[]
    for order in orders:
        orderdata={
            'id':order.id,
            'order_no':order.order_no,
            'isclientcommnd':order.isclientcommnd,
            'note':order.note,
            'client':order.client.id,
            'salsemanid':order.salseman.id if order.salseman else None,
            'date':order.date,
            'items':[]
        }
        orderstosend.append(orderdata)
        orderitems=Orderitem.objects.filter(order=order)
        thisordertotal=0
        for item in orderitems:
            thisordertotal+=round(item.product.sellprice*item.qty, 2)
            orderitemsdata={
                'ordernumber':item.order.order_no,
                #uniqcode will be the connection
                'uniqcode':item.product.uniqcode,
                'qty':item.qty,
                'price':item.price,
                'total':item.total,
            }
            orderdata['items'].append(orderitemsdata)
            #orderitemsstosend.append(orderitemsdata)
        orderdata['total']=thisordertotal
    print('orders', orderstosend)
    return JsonResponse({
        'success':True,
        'orders':orderstosend,
        #'items':orderitemsstosend
    })